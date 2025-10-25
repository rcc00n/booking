from bisect import bisect_left
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Dict, Any, List, Sequence, Iterable, Tuple, Optional
from dataclasses import dataclass, field
from urllib.parse import urlencode

from django.contrib.admin import DateFieldListFilter
from django.contrib.admin.options import IS_POPUP_VAR, TO_FIELD_VAR
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.core.exceptions import FieldError, PermissionDenied
from django.db import transaction, IntegrityError
from django.template.loader import render_to_string
from django.template.response import TemplateResponse
from django.conf import settings
from django.contrib import admin, messages
from django.db.models import Sum, Count, Q, F, ExpressionWrapper, IntegerField, Prefetch, OuterRef, Subquery, Exists
from itertools import cycle

from django.utils.formats import number_format
from django.utils.timezone import localtime, datetime, make_aware, localdate, get_current_timezone, make_naive, is_aware
from django.utils.html import escape
from core.utils.admin_perms import is_master, master_obj
from django.shortcuts import redirect, get_object_or_404
from django.utils.html import format_html, format_html_join
from django.utils.safestring import mark_safe
from datetime import date, timedelta, time as time_cls
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseRedirect, HttpResponseBadRequest
from django.templatetags.static import static
from django.contrib.auth.models import Permission
from django.db.models.functions import Coalesce, Concat, Greatest
from django.db.models import DecimalField, Value
from django.contrib.contenttypes.models import ContentType
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
import re
from django.utils.dateparse import parse_date
from django.utils.translation import gettext_lazy as _
from django.urls import path, reverse, NoReverseMatch, re_path
from django.http import HttpResponse
from django.shortcuts import render
from .filters import *
from .models import *
from core.models import Notification
from .forms import *
from .forms import ProductImportUploadForm
from .validators import *
from core.services.user_import import (
    import_users_from_file,
    UserImportError,
    UserImportSchemaError,
)
from core.services.product_import import import_products_from_file, ProductImportError
from core.services.pricing import compute_appointment_pricing, PricingComputationError
from core.services import payments as payment_services
from core.utils.fees import CARD_PROCESSING_PERCENT, CARD_PROCESSING_FIXED, card_processing_fee
from core.tasks import generate_payment_receipt_task, email_payment_receipt_task

PAID_BADGE_ICON_URL = static("admin/icons/paid.png")
PARTIAL_BADGE_ICON_URL = static("admin/icons/partially-paid.png")
NOTES_BADGE_ICON_URL = static("admin/icons/message.png")

# -----------------------------
# Custom filter for filtering users by Role
# -----------------------------

# Переопределение index view
from datetime import timedelta
from django.contrib import admin
from django.db.models import Count, Sum, Q, F
from django.template.response import TemplateResponse
from django.utils import timezone
from django.utils.timezone import localdate
from datetime import timezone as py_tz

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers


def _autosize_columns(ws):
    """Resize worksheet columns based on content length with sane bounds."""
    if ws.max_column == 0:
        return

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for column_cells in ws.iter_cols(
            min_col=col_idx,
            max_col=col_idx,
            min_row=1,
            max_row=ws.max_row or 1,
        ):
            for cell in column_cells:
                value = cell.value
                text_length = 0 if value is None else len(str(value))
                if text_length > max_len:
                    max_len = text_length
        adjusted_width = min(60, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def _write_xlsx(headers, rows, *, filename="export.xlsx", sheet_name="Export"):
    """Build an XLSX response preserving header order and basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Export")[:31]

    if headers:
        ws.append(list(headers))

    for row in rows:
        if isinstance(row, dict):
            ordered = [row.get(h) for h in headers] if headers else list(row.values())
        elif isinstance(row, (list, tuple)):
            ordered = list(row)
        else:
            try:
                ordered = list(row)
            except TypeError:
                ordered = [row]

        current = []
        for value in ordered:
            if value is None:
                current.append(None)
                continue

            if isinstance(value, datetime):
                processed = value
                if is_aware(processed):
                    try:
                        processed = localtime(processed)
                    except Exception:
                        pass
                processed = processed.replace(tzinfo=None)
                current.append(processed)
                continue

            if isinstance(value, date):
                current.append(value)
                continue

            if isinstance(value, time_cls):
                current.append(value)
                continue

            if isinstance(value, (int, float, bool, Decimal)):
                current.append(value)
                continue

            if isinstance(value, str):
                current.append(value)
                continue

            current.append(str(value))

        ws.append(current)

    if headers:
        for col_idx, header in enumerate(headers, start=1):
            key = (header or "").lower()
            if "date" in key or "time" in key or key.endswith("_at") or key == "at":
                for column_cells in ws.iter_cols(
                    min_col=col_idx,
                    max_col=col_idx,
                    min_row=2,
                    max_row=ws.max_row,
                ):
                    for cell in column_cells:
                        value = cell.value
                        if value is None:
                            continue
                        if isinstance(value, datetime):
                            cell.number_format = "YYYY-MM-DD HH:MM"
                        elif isinstance(value, date):
                            cell.number_format = "YYYY-MM-DD"

    _autosize_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _coerce_json(value):
    if isinstance(value, dict):
        return {k: _coerce_json(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_coerce_json(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time_cls):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


@dataclass(frozen=True)
class ExportDashboardEntry:
    key: str
    label: str
    url: str
    file_format: str = "XLSX"
    group: str = "Data"
    description: Optional[str] = None
    params: Dict[str, Any] = field(default_factory=dict)
    supports_range: bool = False
    requires_range: bool = False

# --- Preset Date Range Filter (factory) ---
def _start_of_week(d: date) -> date:
    """Return Monday for the given date."""
    return d - timedelta(days=d.weekday())


def _start_of_month(d: date) -> date:
    return d.replace(day=1)


def _month_bounds_for_last_month(d: date) -> Tuple[date, date]:
    first_this_month = _start_of_month(d)
    last_month_end = first_this_month - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    return last_month_start, last_month_end


def make_preset_date_filter(
    *, date_field: str, title: str = _("Period"), param: str = "period"
):
    """
    Factory returning a SimpleListFilter bound to a given date/datetime field.
    Usage: list_filter += [make_preset_date_filter(date_field="created_at")]
    """

    filter_title = title
    filter_param = param
    filter_date_field = date_field

    class PresetDateRangeFilter(admin.SimpleListFilter):
        title = filter_title
        parameter_name = filter_param

        def lookups(self, request, model_admin):
            return [
                ("today", _("Сегодня")),
                ("week_to_date", _("С начала недели")),
                ("month_to_date", _("С начала месяца")),
                ("last_week", _("Прошлая неделя (Пн–Вс)")),
                ("last_month", _("Прошлый календарный месяц")),
            ]

        def queryset(self, request, queryset):
            value = self.value()
            if not value:
                return queryset

            today = timezone.localdate()
            start: Optional[date] = None
            end: Optional[date] = None

            if value == "today":
                start = end = today
            elif value == "week_to_date":
                start = _start_of_week(today)
                end = today
            elif value == "month_to_date":
                start = _start_of_month(today)
                end = today
            elif value == "last_week":
                start = _start_of_week(today) - timedelta(days=7)
                end = start + timedelta(days=6)
            elif value == "last_month":
                start, end = _month_bounds_for_last_month(today)

            if start is None or end is None:
                return queryset

            # Prefer __date lookup for DateTime fields; fallback to plain lookup for DateField.
            try:
                return queryset.filter(
                    **{
                        f"{filter_date_field}__date__gte": start,
                        f"{filter_date_field}__date__lte": end,
                    }
                )
            except FieldError:
                return queryset.filter(
                    **{
                        f"{filter_date_field}__gte": start,
                        f"{filter_date_field}__lte": end,
                    }
                )

    return PresetDateRangeFilter

from core.models import (
    Appointment, AppointmentItem, Payment,
    AppointmentStatus, Role, MasterProfile, Service,
    ClientIntakeForm, ClientIntakeFormSubmission,
)


def custom_index(request):
    today = localdate()
    now = timezone.now()
    week_ago = today - timedelta(days=6)
    week_days = [today - timedelta(days=6 - i) for i in range(7)]
    first_day = today.replace(day=1)

    # Базовые QS
    appts_7d = Appointment.objects.filter(start_time__date__range=[week_ago, today])
    payments_7d = Payment.objects.filter(appointment__start_time__date__range=[week_ago, today])

    # Роль/профиль мастера
    userprof = getattr(request.user, "userprofile", None)

    master_profile = getattr(userprof, "master_profile", None) if userprof else None

    # График продаж/записей за 7 дней
    chart_data, total_sales = [], 0.0
    for i in range(7):
        day = today - timedelta(days=6 - i)
        sales = payments_7d.filter(appointment__start_time__date=day) \
                    .aggregate(total=Sum("amount"))["total"] or 0
        appts = appts_7d.filter(start_time__date=day).distinct().count()
        total_sales += float(sales)

        chart_data.append({"day": day.strftime("%a %d"), "sales": float(sales), "appointments": appts})

    # Статусы и счётчики на ближайшие 7 дней

    confirmed = AppointmentStatus.objects.filter(name="Confirmed").first()
    cancelled = AppointmentStatus.objects.filter(name="Cancelled").first()

    # Подзапрос «последний статус для визита»
    last_status_sq_items = (
        AppointmentStatusHistory.objects
        .filter(appointment_id=OuterRef("appointment_id"))
        .order_by("-set_at")
        .values("status_id")[:1]
    )

    # берём позиции ближайшей недели (как и было), но аннотируем последний статус визита
    upcoming_items = (
        AppointmentItem.objects
        .filter(start_time__date__range=(today, today + timedelta(days=7)))
        .annotate(last_status_id=Subquery(last_status_sq_items))
    )

    # считаем КОЛ-ВО ВИЗИТОВ (distinct по appointment_id), где последний статус == нужному
    confirmed_count = (
        upcoming_items
        .filter(last_status_id=getattr(confirmed, "id", None))
        .values("appointment_id")
        .distinct()
        .count()
    ) if confirmed else 0

    cancelled_count = (
        upcoming_items
        .filter(last_status_id=getattr(cancelled, "id", None))
        .values("appointment_id")
        .distinct()
        .count()
    ) if cancelled else 0


# Top services (текущий месяц) — считаем позиции у Service через обратную связь "appointmentitem"
    top_services = (
        Service.objects.annotate(
            count=Count(
                "appointmentitem",
                filter=Q(appointmentitem__appointment__start_time__date__gte=first_day),
            )
        )
        .order_by("-count")[:10]
    )
    first_day = localdate().replace(day=1)
    if first_day.month == 12:
        first_day_next = date(first_day.year + 1, 1, 1)
    else:
        first_day_next = date(first_day.year, first_day.month + 1, 1)

    # 2) Агрегация денег по мастеру из AppointmentItem.final_price
    paid_by_appt = dict(
        Payment.objects
        .filter(created_at__date__gte=first_day, created_at__date__lt=first_day_next)
        .values("appointment_id")
        .annotate(total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))))
        .values_list("appointment_id", "total")
    )

    master_month_totals = {}
    top_masters = []
    if paid_by_appt:
        rows = (
            AppointmentItem.objects
            .filter(appointment_id__in=paid_by_appt.keys())
            .values("appointment_id", "master_id")
            .annotate(msum=Coalesce(Sum("final_price"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))))
        )

        appt_total = {}
        for r in rows:
            aid = r["appointment_id"]
            appt_total[aid] = (appt_total.get(aid, 0) or 0) + (r["msum"] or 0)

        master_totals = {}
        for r in rows:
            aid = r["appointment_id"]
            mid = r["master_id"]
            msum = r["msum"] or 0
            paid = paid_by_appt.get(aid, 0) or 0
            total = appt_total.get(aid, 0) or 0
            if paid and total > 0 and msum > 0:
                part = paid * (msum / total)
                master_totals[mid] = (master_totals.get(mid, 0) or 0) + part

        master_month_totals = master_totals
        top_ids = sorted(master_month_totals.keys(), key=lambda k: master_month_totals[k], reverse=True)[:10]
        top_masters_qs = MasterProfile.objects.filter(pk__in=top_ids).select_related("user")
        top_masters = list(top_masters_qs)
        for m in top_masters:
            m.total = master_month_totals.get(m.pk, Decimal("0"))

        top_masters = sorted(top_masters, key=lambda m: m.total or 0, reverse=True)[:10]

    current_month_label = first_day.strftime("%B %Y")
    month_targets = (
        MasterMonthlySalesTarget.objects
        .filter(month__year=first_day.year, month__month=first_day.month)
        .select_related("master__user__user")
    )
    targets_by_master = {t.master_id: t for t in month_targets}
    master_target_rows = []
    master_target_for_current_user = None
    masters_qs = MasterProfile.objects.select_related("user__user").order_by(
        "user__user__first_name", "user__user__last_name"
    )
    for master in masters_qs:
        target_obj = targets_by_master.get(master.id)
        achieved_amount = master_month_totals.get(master.id) or Decimal("0")
        target_amount = getattr(target_obj, "target_amount", None)
        remaining_amount = None
        if target_amount is not None:
            remaining_amount = target_amount - achieved_amount
            if remaining_amount < Decimal("0"):
                remaining_amount = Decimal("0")
        entry = {
            "master": master,
            "target": target_obj,
            "target_amount": target_amount,
            "achieved_amount": achieved_amount,
            "remaining_amount": remaining_amount,
        }
        master_target_rows.append(entry)
        if master_profile and master.id == master_profile.id:
            master_target_for_current_user = entry

    # Недавние встречи (20) с префетчем позиций
    recent_appointments = (
        AppointmentItem.objects.select_related("appointment__client__user").order_by("-start_time")[:20]
    )

    # Сегодняшние предстоящие встречи (Appointment + items); мастеру — только его
    today_appointments = (
        AppointmentItem.objects.filter(start_time__date=today, start_time__gte=now)
        .select_related("appointment__client__user")
        .order_by("start_time")
    )
    if is_master(request.user) and master_profile:
        today_appointments = today_appointments.filter(master=master_profile).distinct()

    # Ежедневная разбивка Confirmed/Cancelled (на 7 дней вперёд)
    daily_counts = []
    for day in week_days:
        c = Appointment.objects.filter(start_time__date=day,
                                       appointmentstatushistory__status=confirmed) \
            .distinct().count() if confirmed else 0
        x = Appointment.objects.filter(start_time__date=day,
                                       appointmentstatushistory__status=cancelled) \
            .distinct().count() if cancelled else 0
        daily_counts.append({"day": day.strftime("%a %d"), "confirmed": c, "cancelled": x})

    context = admin.site.each_context(request)
    context.update({
        "is_master": is_master(request.user),
        "daily_appointments": daily_counts,
        "chart_data": chart_data,
        "total_sales": total_sales,
        "upcoming_total": upcoming_items.count(),
        "confirmed_count": confirmed_count,
        "cancelled_count": cancelled_count,
        "top_services": top_services,      # Service с .name и .count
        "top_masters": top_masters,        # MasterProfile с .total
        "today": today,
        "recent_appointments": recent_appointments,
        "today_appointments": today_appointments,
        "master_target_rows": master_target_rows,
        "master_target_month_label": current_month_label,
        "master_target_for_current_user": master_target_for_current_user,
    })
    return TemplateResponse(request, "admin/index.html", context)

# Переопределить главную страницу
admin.site.index = custom_index
class ExportCsvMixin:
    export_fields = None  # список полей; можно переопределить в admin
    export_dashboard_label = None
    export_dashboard_group = None
    export_dashboard_description = None
    export_dashboard_supports_range = True
    export_dashboard_requires_range = False
    export_dashboard_params = None

    def get_urls(self):
        opts = self.model._meta
        return [
            path(
                "export-csv/",
                self.admin_site.admin_view(self.export_all_csv),
                name=f"{opts.app_label}_{opts.model_name}_export_csv"
            )
        ] + super().get_urls()

    def export_all_csv(self, request):
        queryset = self.get_queryset(request)

        fields = self.export_fields or [field.name for field in self.model._meta.fields]
        filename = f"{self.model._meta.model_name}.xlsx"

        def iter_rows():
            for obj in queryset:
                if hasattr(self, "get_export_row"):
                    row = self.get_export_row(obj)
                else:
                    row = [getattr(obj, field) for field in fields]

                if isinstance(row, dict):
                    yield [row.get(field) for field in fields]
                elif isinstance(row, (list, tuple)):
                    yield list(row)
                else:
                    yield [row]

        return _write_xlsx(fields, iter_rows(), filename=filename)

    def changelist_view(self, request, extra_context=None):
        # Попробуем reverse без краша
        extra_context = extra_context or {}
        try:
            opts = self.model._meta
            export_url = reverse(f'admin:{opts.app_label}_{opts.model_name}_export_csv')
            export_url += f"?{request.GET.urlencode()}"
            extra_context['export_url'] = export_url
        except NoReverseMatch:
            extra_context['export_url'] = None

        return super().changelist_view(request, extra_context=extra_context)



class ExportXlsxMixin:
    """
    Универсальный XLSX-экспорт для Django Admin.
    Дает:
      - URL 'export-xlsx/' для полной выгрузки текущего changelist с учетом фильтров
      - Экшены: export_appointments_xlsx / export_appointment_items_xlsx
    Подключение: наследуй нужный ModelAdmin от этого миксина.
    """

    export_dashboard_label = None
    export_dashboard_group = None
    export_dashboard_description = None
    export_dashboard_supports_range = True
    export_dashboard_requires_range = False
    export_dashboard_params = None

    # ============ ПУБЛИЧНЫЕ ТОЧКИ ============

    def get_urls(self):
        opts = self.model._meta
        return [
            path(
                "export-xlsx/",
                self.admin_site.admin_view(self._export_all_xlsx_view),
                name=f"{opts.app_label}_{opts.model_name}_export_xlsx",
            ),
        ] + super().get_urls()

    # Экшен: 1 строка = 1 Appointment
    def export_appointments_xlsx(self, request, queryset):
        qs = self._qs_for_export(request, queryset)
        headers = [
            "Appointment ID",
            "Start",
            "Client",
            "Status",
            "Payment Status",
            "Personal Discount",
            "Items Count",
            "Total (sum of items)",
            "Items (preview)",
        ]
        rows = []
        for appt in qs:
            items = getattr(appt, "appointmentitem_set").all()
            total = sum([(getattr(it, "final_price", None) or 0) for it in items])
            preview = " | ".join(
                f"{getattr(getattr(it, 'service', None), 'name', getattr(it, 'service', ''))} ×{getattr(it, 'quantity', 1)}"
                for it in items[:6]
            )
            if items.count() > 6:
                preview += " …"
            rows.append([
                str(getattr(appt, "pk", "")),
                self._to_naive_dt(getattr(appt, "start_time", None)),
                self._client_name(getattr(appt, "client", None)),
                self._safe_str(getattr(appt, "status", "")),
                self._safe_str(getattr(appt, "payment_status", "")),
                getattr(appt, "personal_discount", None),
                items.count(),
                self._as_decimal(total),
                preview,
            ])

        # Start = колонка 2 (datetime), Total = колонка 8 (денежный)
        return self._xlsx_response(
            "appointments.xlsx", "Appointments", headers, rows,
            money_cols={8}, datetime_cols={2}
        )

    export_appointments_xlsx.short_description = "Export Appointments (XLSX, 1 row per appointment)"

    # Экшен: 1 строка = 1 Appointment Item
    def export_appointment_items_xlsx(self, request, queryset):
        dataset = self._appointment_item_export_dataset(self._qs_for_export(request, queryset))
        return self._xlsx_response(
            "appointment_items.xlsx",
            "Items",
            dataset["headers"],
            dataset["rows"],
            money_cols=dataset["money_cols"],
            datetime_cols=dataset["datetime_cols"],
        )

    export_appointment_items_xlsx.short_description = "Export Appointment Items (XLSX, 1 row per item)"

    # Зарегистрируй экшены в админ-классе:
    # actions = ["export_appointments_xlsx", "export_appointment_items_xlsx"]

    # ============ ВСПОМОГАТЕЛЬНОЕ ============

    def _export_all_xlsx_view(self, request):
        """
        Полная выгрузка текущего списка (как в changelist, с учетом фильтров).
        По умолчанию — 1 строка = 1 объект self.model с его ._meta.fields.
        """
        queryset = self._qs_for_export(request)
        model_meta = getattr(getattr(self, "model", None), "_meta", None)
        model_name = getattr(model_meta, "model_name", "").lower() if model_meta else ""

        if model_name == "appointment":
            dataset = self._appointment_item_export_dataset(queryset)
            filename = f"{self.model._meta.model_name}.xlsx"
            return self._xlsx_response(
                filename,
                "Export",
                dataset["headers"],
                dataset["rows"],
                money_cols=dataset["money_cols"],
                datetime_cols=dataset["datetime_cols"],
            )

        fields = [f.name for f in self.model._meta.fields]
        headers = fields
        rows = ([self._xlsx_safe(getattr(obj, f)) for f in fields] for obj in queryset)
        return self._xlsx_response(f"{self.model._meta.model_name}.xlsx", "Export", headers, rows)

    def _qs_for_export(self, request, queryset=None):
        """
        Оптимизация выборки для экшенов по встречам и их позициям.
        Переопредели под свою модель при необходимости.
        """
        qs = (queryset or self.get_queryset(request))

        model_meta = getattr(getattr(self, "model", None), "_meta", None)
        model_name = getattr(model_meta, "model_name", "").lower() if model_meta else ""
        if model_name != "appointment":
            return qs

        qs = qs.select_related("client__user")

        item_qs = AppointmentItem.objects.select_related(
            "service",
            "service__category",
            "master__user__user",
        ).order_by("start_time")

        history_qs = (
            AppointmentStatusHistory.objects
            .select_related("status", "set_by__user", "cancellation_reason")
            .order_by("-set_at")
        )

        payment_qs = Payment.objects.select_related("method")

        return qs.prefetch_related(
            Prefetch("items", queryset=item_qs, to_attr="_export_items"),
            Prefetch(
                "appointmentstatushistory_set",
                queryset=history_qs,
                to_attr="_export_status_history",
            ),
            Prefetch("payments", queryset=payment_qs, to_attr="_export_payments"),
        )

    def _appointment_item_export_dataset(self, qs):
        """
        Построение набора данных для XLSX с 1 строкой на AppointmentItem.
        """
        headers = [
            "Appt. ref.",
            "Client",
            "Team member",
            "Status",
            "Created date",
            "Scheduled date",
            "Cancelled date",
            "Category",
            "Service",
            "Duration (mins)",
            "Appt. slot",
            "Created by",
            "Cancelled by",
            "Net sales",
            "Cancellation reason",
            "Fees charged",
            "Upfront payments",
        ]

        card_method_keywords = ("card", "credit", "debit", "terminal", "stripe")

        def person_label(entity):
            if entity is None:
                return ""
            user_obj = None
            candidate = getattr(entity, "user", None)
            if candidate is not None:
                nested = getattr(candidate, "user", None)
                user_obj = nested or candidate
            elif hasattr(entity, "get_full_name"):
                name = entity.get_full_name()
                if name:
                    return name
                return getattr(entity, "username", "") or ""
            if user_obj is None:
                return ""
            full_name = user_obj.get_full_name()
            if full_name:
                return full_name
            return getattr(user_obj, "username", "") or ""

        def localize_dt(value):
            if not value:
                return None
            try:
                return localtime(value)
            except Exception:
                return value

        def to_excel_dt(value):
            dt = localize_dt(value)
            if isinstance(dt, datetime) and is_aware(dt):
                return dt.replace(tzinfo=None)
            return dt

        def coalesce_decimal(*values):
            for candidate in values:
                if candidate in (None, ""):
                    continue
                if isinstance(candidate, Decimal):
                    return candidate.quantize(TWOPLACES)
                try:
                    return Decimal(candidate).quantize(TWOPLACES)
                except (InvalidOperation, TypeError, ValueError):
                    continue
            return Decimal("0.00")

        def derive_status_bundle(appt):
            status_name = getattr(getattr(appt, "status", None), "name", None)
            history = getattr(appt, "_export_status_history", None)
            if history is None:
                history = list(
                    appt.appointmentstatushistory_set.select_related(
                        "status", "set_by__user", "cancellation_reason"
                    ).order_by("-set_at")
                )
            cancelled_dt = None
            cancelled_by = ""
            cancel_reason = ""
            if history:
                latest = history[0]
                if not status_name:
                    status_name = getattr(getattr(latest, "status", None), "name", None)
                for entry in history:
                    entry_status = (getattr(getattr(entry, "status", None), "name", "") or "").lower()
                    if entry_status in {"cancelled", "canceled"}:
                        cancelled_dt = getattr(entry, "set_at", None)
                        cancelled_by = person_label(getattr(entry, "set_by", None))
                        cancel_reason = getattr(getattr(entry, "cancellation_reason", None), "name", "") or ""
                        break
            return status_name or "New", cancelled_dt, cancelled_by, cancel_reason

        def uses_card_fee(appt):
            if getattr(appt, "apply_card_processing_fee", False):
                return True
            payments = getattr(appt, "_export_payments", None)
            if payments is None:
                payments = appt.payments.select_related("method").all()
            for payment in payments:
                method_name = (getattr(getattr(payment, "method", None), "name", "") or "").lower()
                if any(keyword in method_name for keyword in card_method_keywords):
                    return True
            return False

        def iter_items(appt):
            items = getattr(appt, "_export_items", None)
            if items is not None:
                return items
            return appt.items.select_related(
                "service",
                "service__category",
                "master__user__user",
            ).order_by("start_time")

        def short_reference(appt):
            value = getattr(appt, "id", None) or getattr(appt, "pk", None)
            if not value:
                return ""
            hex_value = getattr(value, "hex", None)
            if hex_value:
                return hex_value[:8].upper()
            return str(value).replace("-", "")[:8].upper()

        rows = []
        for appt in qs:
            status_name, cancelled_dt, cancelled_by, cancel_reason = derive_status_bundle(appt)
            created_by_display = person_label(getattr(appt, "created_by", None))
            client_display = person_label(getattr(appt, "client", None))
            fee_applicable = uses_card_fee(appt)
            appt_ref = short_reference(appt)
            created_dt = to_excel_dt(getattr(appt, "created_at", None))

            for item in iter_items(appt):
                service = getattr(item, "service", None)
                category = getattr(getattr(service, "category", None), "name", "") if service else ""
                service_name = getattr(service, "name", "") or ""
                duration_override = getattr(item, "duration_override_min", None)
                if duration_override is not None:
                    duration_minutes = int(duration_override)
                    buffer_minutes = 0
                else:
                    base_duration = getattr(service, "duration_min", None) if service else None
                    duration_minutes = int(base_duration or 0)
                    buffer_minutes = int(getattr(service, "extra_time_min", 0) or 0) if service else 0
                slot_minutes = duration_minutes + buffer_minutes

                start_dt_local = localize_dt(getattr(item, "start_time", None))
                scheduled_dt = to_excel_dt(getattr(item, "start_time", None))
                end_dt_local = start_dt_local + timedelta(minutes=slot_minutes) if start_dt_local else None
                appt_slot = ""
                if start_dt_local and end_dt_local:
                    appt_slot = f"{start_dt_local.strftime('%H:%M:%S')}-{end_dt_local.strftime('%H:%M:%S')}"

                net_sales = coalesce_decimal(
                    getattr(item, "final_price", None),
                    getattr(item, "unit_price", None),
                    getattr(service, "base_price", None) if service else None,
                )
                tax_component = None
                try:
                    tax_component = getattr(item, "tax_amount", None)
                except Exception:
                    tax_component = None
                if tax_component not in (None, ""):
                    net_sales = (net_sales + coalesce_decimal(tax_component)).quantize(TWOPLACES)
                fees_charged = Decimal("0.00")
                if fee_applicable:
                    fees_charged = (net_sales * CARD_PROCESSING_PERCENT + CARD_PROCESSING_FIXED).quantize(TWOPLACES)
                upfront_payments = Decimal("0.00")

                rows.append([
                    appt_ref,
                    client_display,
                    person_label(getattr(item, "master", None)),
                    status_name,
                    created_dt,
                    scheduled_dt,
                    to_excel_dt(cancelled_dt),
                    category or "",
                    service_name,
                    duration_minutes,
                    appt_slot,
                    created_by_display,
                    cancelled_by,
                    net_sales,
                    cancel_reason,
                    fees_charged,
                    upfront_payments,
                ])

        return {
            "headers": headers,
            "rows": rows,
            "money_cols": {14, 16, 17},
            "datetime_cols": {5, 6, 7},
        }

    # ============ НИЗКИЙ УРОВЕНЬ (XLSX) ============

    def _xlsx_response(
            self,
            filename: str,
            sheet_title: str,
            headers: Sequence[str],
            rows: Iterable[Sequence],
            *,
            money_cols: set[int] | None = None,
            datetime_cols: set[int] | None = None,
    ) -> HttpResponse:
        """
        Возвращает HttpResponse с XLSX.
        money_cols / datetime_cols — 1-based индексы колонок для number_format.
        """
        money_cols = money_cols or set()
        datetime_cols = datetime_cols or set()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31] if sheet_title else "Export"

        # Заголовок
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        # Данные
        row_idx = 1
        for row in rows:
            row_idx += 1
            safe_row = [self._xlsx_safe(v) for v in row]
            ws.append(safe_row)

            # Форматы
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in money_cols and isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                if col_idx in datetime_cols and isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"

        # Авто-ширина
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)

        resp = HttpResponse(
            buff.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ============ ПРИВЕДЕНИЕ ТИПОВ ============

    def _xlsx_safe(self, v):
        """Конвертирует произвольные значения в Excel-дружелюбные."""
        if v is None:
            return ""
        # datetime → naive (локальная TZ), time → без tzinfo
        if isinstance(v, datetime):
            return self._to_naive_dt(v)
        if isinstance(v, time_cls):
            # Excel не поддерживает tz-aware time; сбрасываем tzinfo
            return time_cls(v.hour, v.minute, v.second, v.microsecond)
        if isinstance(v, (date, )):
            return v
        if isinstance(v, (str, int, float, bool, Decimal)):
            return v
        # Модели/Enums/Objects → строка (использует __str__)
        return str(v)

    def _to_naive_dt(self, dt: datetime) -> datetime | None:
        """Excel не любит tz-aware; переводим в naive (settings.TIME_ZONE)."""
        if not isinstance(dt, datetime):
            return dt
        return make_naive(dt) if is_aware(dt) else dt

    def _as_decimal(self, v):
        if v is None:
            return None
        if isinstance(v, Decimal):
            return v
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        try:
            return Decimal(str(v))
        except Exception:
            return str(v)

    def _safe_str(self, v, default=""):
        return default if v in (None, "") else str(v)

    def _client_name(self, client):
        if client is None:
            return ""
        full = getattr(client, "full_name", "")
        if full:
            return full
        user = getattr(client, "user", None)
        return getattr(user, "username", "") if user else ""

    def _wb_new(self) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)  # удаляем дефолтный пустой лист
        return wb

    def _append_sheet(self, wb: Workbook, title: str, headers, rows, *, money_cols=None, datetime_cols=None):
        money_cols = set(money_cols or [])
        datetime_cols = set(datetime_cols or [])

        ws = wb.create_sheet(title=title[:31] or "Export")
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        row_idx = 1
        for row in rows:
            row_idx += 1
            safe_row = [self._xlsx_safe(v) for v in row]
            ws.append(safe_row)

            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in money_cols and isinstance(cell.value, (int, float, Decimal)):
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                if col_idx in datetime_cols and isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
                if col_idx in datetime_cols and isinstance(cell.value, date) and not isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd"

        # авто-ширина
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for c in ws[letter]:
                s = "" if c.value is None else str(c.value)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        return ws

    def _wb_response(self, wb: Workbook, filename: str) -> HttpResponse:
        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)
        resp = HttpResponse(
            buff.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    def build_statistics_workbook(
            self,
            *,
            start, end,
            kpi_total_revenue, kpi_total_appointments, kpi_with_discount,
            top_bookers, top_spenders, ds_totals,
            discount_breakdown, promo_breakdown, client_source_table,
    ):
        """
        Собирает многолистовую книгу XLSX из данных статистики.
        На вход подаются готовые QuerySet/списки, которые уже посчитаны в stats_view.
        """
        wb = self._wb_new()

        # KPI
        self._append_sheet(
            wb, "KPI",
            headers=["Start", "End", "Total Revenue", "Appointments", "With Discount"],
            rows=[[start, end, kpi_total_revenue, kpi_total_appointments, kpi_with_discount]],
            money_cols={3}, datetime_cols={1, 2},
        )

        # Top Bookers
        self._append_sheet(
            wb, "Top Bookers",
            headers=["Client ID", "First Name", "Last Name", "Email", "Appointments", "Spent"],
            rows=[
                [
                    r["client_id"],
                    r["client__user__first_name"],
                    r["client__user__last_name"],
                    r["client__user__email"],
                    r["appt_count"],
                    r["spent"],
                ] for r in top_bookers
            ],
            money_cols={6},
        )

        # Top Spenders
        self._append_sheet(
            wb, "Top Spenders",
            headers=["Client ID", "First Name", "Last Name", "Email", "Spent", "Appointments"],
            rows=[
                [
                    r["client_id"],
                    r["client__user__first_name"],
                    r["client__user__last_name"],
                    r["client__user__email"],
                    r["spent"],
                    r["appt_count"],
                ] for r in top_spenders
            ],
            money_cols={5},
        )

        # Discount Sources totals
        self._append_sheet(
            wb, "Discount Sources",
            headers=["Discount Source", "Appointments (distinct)", "Revenue"],
            rows=[[r["discount_source"], r["cnt"], r["revenue"]] for r in ds_totals],
            money_cols={3},
        )

        # ServiceDiscount breakdown
        self._append_sheet(
            wb, "ServiceDiscount",
            headers=["ServiceDiscount", "Service", "Uses", "Revenue(by items)", "Saved"],
            rows=[
                [str(d), getattr(d.service, "name", ""), d.uses, d.revenue, d.saved]
                for d in discount_breakdown
            ],
            money_cols={4, 5},
        )

        # Promo breakdown
        self._append_sheet(
            wb, "Promo",
            headers=["Code", "Uses", "Revenue (share)"],
            rows=[[r["code"], r["uses"], r["revenue"]] for r in promo_breakdown],
            money_cols={3},
        )

        # Client Source
        self._append_sheet(
            wb, "Client Source",
            headers=["Key", "Label", "Revenue", "Appointments", "Clients"],
            rows=[
                [r["key"], r["label"], r["revenue"], r["appts"], r["clients"]]
                for r in client_source_table
            ],
            money_cols={3},
        )

        return wb

    def build_statistics_flat_workbook(
            self,
            *,
            start, end,
            kpi_total_revenue, kpi_total_appointments, kpi_with_discount,
            top_bookers, top_spenders, ds_totals,
            discount_breakdown, promo_breakdown, client_source_table,
    ) -> Workbook:
        """Строит однолистовую книгу с блоками статистики друг за другом."""
        source_wb = self.build_statistics_workbook(
            start=start,
            end=end,
            kpi_total_revenue=kpi_total_revenue,
            kpi_total_appointments=kpi_total_appointments,
            kpi_with_discount=kpi_with_discount,
            top_bookers=top_bookers,
            top_spenders=top_spenders,
            ds_totals=ds_totals,
            discount_breakdown=discount_breakdown,
            promo_breakdown=promo_breakdown,
            client_source_table=client_source_table,
        )

        combined = Workbook()
        ws = combined.active
        ws.title = "Statistics"

        sheets = list(source_wb.worksheets)
        for idx, sheet in enumerate(sheets):
            if sheet.max_row == 0:
                continue

            ws.append([sheet.title])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                ws.append(list(row))
                if row_index == 1:
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(vertical="center")

            if idx != len(sheets) - 1:
                ws.append([])

        _autosize_columns(ws)
        return combined


class CustomUserAdmin(ExportCsvMixin ,BaseUserAdmin):
    """
    Custom admin interface for Django's User model, enhanced with roles and profile fields.
    """
    add_form = CustomUserCreationForm
    form = CustomUserChangeForm
    change_list_template = "admin/users/changelist_cards.html"
    export_fields = ['username', 'email', 'first_name', 'last_name', 'phone', 'birth_date', 'address', 'postal_code', 'is_staff', 'is_superuser', 'is_active', 'source', 'consent']
    list_per_page = 10
    readonly_fields = getattr(BaseUserAdmin, "readonly_fields", tuple()) + ("password_change_link",)

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': (
                'usable_password', 'password1', 'password2',
                'email', 'first_name', 'last_name',
                'phone', 'birth_date', 'address', 'postal_code',
                'personal_discount_percent', 'how_heard', 'email_marketing_consent',
                'notes',
                'is_active', 'is_staff', 'is_superuser', 'groups',
            ),
        }),
        ('Health', {
            'classes': ('collapse',),
            'fields': (
                'has_allergies', 'allergies_text', 'gender',
                'chronic_conditions', 'medications', 'pregnant',
                'skin_sensitivity', 'recent_procedures',
                'contraindications', 'health_notes',
            ),
        }),
    )

    def get_export_row(self, obj):
        phone = obj.userprofile.phone if hasattr(obj, 'userprofile') else ''
        birth_date = obj.userprofile.birth_date if hasattr(obj, 'userprofile') else ''
        address = obj.userprofile.address if hasattr(obj, 'userprofile') else ''
        postal_code = obj.userprofile.postal_code if hasattr(obj, 'userprofile') else ''
        source = obj.userprofile.source if hasattr(obj, 'userprofile') else ''
        consent = obj.userprofile.email_marketing_consent if hasattr(obj, 'userprofile') else ''

        return [
            obj.username,
            obj.email,
            obj.first_name,
            obj.last_name,
            phone,
            birth_date,
            address,
            postal_code,
            obj.is_staff,
            obj.is_superuser,
            obj.is_active,
            source,
            consent
        ]
    # Fields shown when adding a new user


    # Fields shown in user list
    list_display = ('username', 'email', 'first_name', 'last_name', 'staff_status', 'phone', 'birth_date', 'source', 'client_status_col')
    list_filter = ('is_superuser', 'userprofile__how_heard', ClientStatusFilter)
    search_fields = ('username', 'email', 'first_name', 'last_name')
    ordering = ('-date_joined',)

    # Field layout when editing a user
    fieldsets = (
        (None, {'fields': ('email', 'password_change_link', 'personal_discount_percent')}),
        ('Personal Info', {'fields': (
            'first_name', 'last_name', 'phone', 'birth_date', 'address',
            'postal_code', 'how_heard', 'email_marketing_consent'
        )}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Files', {'fields': ('files',)}),
        ('Notes', {'fields': ('notes',)}),
        ('Health', {
            'classes': ('collapse', 'wide'),
            'fields': (
                'has_allergies', 'allergies_text', 'gender',
                'chronic_conditions', 'medications',
                'pregnant', 'skin_sensitivity',
                'recent_procedures', 'contraindications',
                'health_notes',
            ),
        }),
    )

    def password_change_link(self, obj):
        if not obj or not getattr(obj, "pk", None):
            return _("Change password after saving the user.")
        try:
            url = reverse("admin:auth_user_password_change", args=[obj.pk])
        except NoReverseMatch:
            return _("Change password")
        return mark_safe(f'<a href="{url}">{_("Change password")}</a>')

    password_change_link.short_description = _("Password")

    def get_queryset(self, request):
        # Prefetch the attached profile to avoid N+1 lookups.
        qs = super().get_queryset(request).select_related('userprofile')

        universal_pending_subquery = ClientIntakeAssignment.objects.filter(
            client__user_id=OuterRef('pk'),
            form__is_universal=True,
            form__is_active=True,
            completed_at__isnull=True,
        )
        qs = qs.annotate(universal_intake_pending=Exists(universal_pending_subquery))

        status_key = request.GET.get('client_status')
        if status_key:
            status_label = ClientStatusFilter.LOOKUP_TO_LABEL.get(status_key)
            if status_label:
                user_ids = list(qs.values_list('id', flat=True))
                if user_ids:
                    profiles = (
                        UserProfile.objects.select_related('user')
                        .filter(user_id__in=user_ids)
                    )
                    matching_ids = [
                        profile.user_id
                        for profile in profiles
                        if profile.client_status == status_label
                    ]
                    qs = qs.filter(id__in=matching_ids) if matching_ids else qs.none()
                else:
                    qs = qs.none()

        user_order = request.GET.get('user_order')
        if user_order == 'oldest':
            qs = qs.order_by('date_joined')
        else:
            qs = qs.order_by('-date_joined')

        # Stash the choice and drop the param before default admin validation runs.
        request._user_order_choice = user_order or 'newest'
        if "user_order" in request.GET:
            mutable_get = request.GET.copy()
            mutable_get.pop("user_order", None)
            request.GET = mutable_get

        return qs

    def history_view(self, request, object_id, extra_context=None):
        if request.GET.get("mode") == "log":
            return super().history_view(request, object_id, extra_context)

        user_model = get_user_model()
        user_obj = get_object_or_404(user_model, pk=object_id)
        profile = getattr(user_obj, "userprofile", None)

        from core.models import Appointment, AppointmentItem, AppointmentStatusHistory

        current_status_sq = (
            AppointmentStatusHistory.objects
            .filter(appointment_id=OuterRef("pk"))
            .order_by("-set_at")
            .values("status__name")[:1]
        )

        if profile is None:
            visits_qs = Appointment.objects.none()
        else:
            visits_qs = (
                Appointment.objects.filter(client=profile)
                .annotate(current_status=Subquery(current_status_sq))
                .select_related("client__user", "payment_status")
                .prefetch_related(
                    Prefetch(
                        "items",
                        queryset=AppointmentItem.objects.select_related(
                            "service",
                            "master__user",
                        ).order_by("start_time"),
                    )
                )
                .order_by("-start_time")
            )

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": user_obj,
            "title": _("History"),
            "user_obj": user_obj,
            "visits": list(visits_qs[:200]),
            "change_url_name": "admin:core_appointment_change",
        }

        if extra_context:
            context.update(extra_context)

        return TemplateResponse(request, "admin/users/history_visits.html", context)

    @staticmethod
    def _normalize_phone_digits(value: str) -> str:
        return re.sub(r"\D+", "", value or "")

    @classmethod
    def _phone_regex_for_digits(cls, digits: str) -> str:
        if not digits:
            return ""
        # Allow any number of non-digit separators between the input digits.
        return r"\D*".join(re.escape(d) for d in digits)

    def get_search_results(self, request, queryset, search_term):
        search_term = (search_term or "").strip()
        if not search_term:
            return queryset, False

        text_q = (
            Q(first_name__icontains=search_term)
            | Q(last_name__icontains=search_term)
            | Q(email__icontains=search_term)
            | Q(username__icontains=search_term)
        )
        digits = self._normalize_phone_digits(search_term)
        if digits:
            regex_pattern = self._phone_regex_for_digits(digits)
            text_q |= Q(userprofile__phone__iregex=regex_pattern)

        return queryset.filter(text_q), False

    def get_ordering(self, request):
        user_order = getattr(request, "_user_order_choice", None) or request.GET.get("user_order")
        if user_order == "oldest":
            return ("date_joined",)
        if user_order == "newest" or not user_order:
            return ("-date_joined",)
        return super().get_ordering(request)

    def lookup_allowed(self, lookup, value):
        if lookup == "user_order":
            return True
        return super().lookup_allowed(lookup, value)

    @admin.display(description="Status")
    def client_status_col(self, obj):
        up = getattr(obj, 'userprofile', None)
        return getattr(up, 'client_status', '-') if up else ('-')

    @admin.display(description="Source")
    def source(self, obj):
        up = getattr(obj, 'userprofile', None)
        return getattr(up, 'source', '-') if up else '-'

    def get_fieldsets(self, request, obj=None):
        # Allow Django to use default fieldsets logic
        return super().get_fieldsets(request, obj)

    def get_form(self, request, obj=None, **kwargs):
        # Return different form on add vs change
        return self.add_form if obj is None else self.form



    def save_model(self, request, obj, form, change):
        # Persist the user first using the parent logic
        with transaction.atomic():
            super().save_model(request, obj, form, change)

            if not hasattr(form, "cleaned_data"):
                return

            profile_data = {}
            for form_field, profile_field in (
                ("phone", "phone"),
                ("birth_date", "birth_date"),
                ("address", "address"),
                ("postal_code", "postal_code"),
            ):
                if form_field in form.cleaned_data:
                    profile_data[profile_field] = form.cleaned_data.get(form_field)

            if not profile_data:
                return

            profile, _ = UserProfile.objects.select_for_update().get_or_create(user=obj)

            phone_value = profile_data.get("phone")
            profile.phone = phone_value or profile.phone
            profile.birth_date = profile_data.get("birth_date", profile.birth_date)
            profile.address = profile_data.get("address", profile.address) or ""
            profile.postal_code = profile_data.get("postal_code", profile.postal_code) or ""
            profile.save()

    # Custom display methods for user profile fields
    def phone(self, instance):
        return instance.userprofile.phone if hasattr(instance, 'userprofile') else '-'

    def birth_date(self, instance):
        return instance.userprofile.birth_date if hasattr(instance, 'userprofile') else '-'

    def _is_popup(self, request):
        return IS_POPUP_VAR in request.POST or IS_POPUP_VAR in request.GET

    def _profile_popup_response(self, request, profile, *, action):
        if profile is None:
            return None
        to_field = request.POST.get(TO_FIELD_VAR) or request.GET.get(TO_FIELD_VAR)
        if to_field:
            try:
                field = profile._meta.get_field(to_field)
            except Exception:
                value = getattr(profile, to_field, profile.pk)
            else:
                value = field.value_from_object(profile)
        else:
            value = profile.pk
        popup_response_data = {
            "action": action,
            "value": str(value),
            "obj": str(profile),
        }
        return TemplateResponse(
            request,
            self.popup_response_template,
            {
                **self.admin_site.each_context(request),
                "popup_response_data": json.dumps(popup_response_data),
            },
        )

    def _maybe_redirect_back(self, request, response):
        """
        Если пришли из Appointment (по нашему параметру _from_appointment),
        вернёмся на календарь визитов.
        """
        if "_from_appointment" in request.GET or "_from_appointment" in request.POST:

            return redirect(reverse("admin:core_appointment_changelist"))
        return response

    def response_add(self, request, obj, post_url_continue=None):
        if self._is_popup(request):
            profile = getattr(obj, "userprofile", None)
            popup_resp = self._profile_popup_response(request, profile, action="add")
            if popup_resp is not None:
                return popup_resp
        resp = super().response_add(request, obj, post_url_continue)
        return self._maybe_redirect_back(request, resp)

    def response_change(self, request, obj):
        if self._is_popup(request):
            profile = getattr(obj, "userprofile", None)
            popup_resp = self._profile_popup_response(request, profile, action="change")
            if popup_resp is not None:
                return popup_resp
        resp = super().response_change(request, obj)
        return self._maybe_redirect_back(request, resp)

    def response_delete(self, request, obj_display, obj_id):
        resp = super().response_delete(request, obj_display, obj_id)
        return self._maybe_redirect_back(request, resp)

    @admin.display(description="")
    def send_notify_button(self, obj):
        return mark_safe(
            f'<button type="button" class="send-notify-btn" '
            f'data-user-id="{obj.id}" '
            f'data-user-name="{obj.get_full_name() or obj.username}">Send Notification</button>'
        )
    def get_urls(self):
        urls = super().get_urls()
        opts = self.model._meta
        custom_urls = [
            path(
                'import-users/',
                self.admin_site.admin_view(self.import_users_view),
                name=f'{opts.app_label}_{opts.model_name}_import',
            ),
            path('send_notification/', self.admin_site.admin_view(self.send_notification_view), name='send_notification'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        try:
            opts = self.model._meta
            extra_context.setdefault(
                'import_url',
                reverse(f'admin:{opts.app_label}_{opts.model_name}_import'),
            )
            extra_context.setdefault('import_label', 'Import Users')
        except NoReverseMatch:
            pass
        extra_context['client_status_options'] = [
            {'value': key, 'label': label}
            for key, label in ClientStatusFilter.LOOKUP_TO_LABEL.items()
        ]
        current_status = request.GET.get('client_status', '')
        extra_context['client_status_current'] = current_status
        extra_context['client_status_current_label'] = (
            ClientStatusFilter.LOOKUP_TO_LABEL.get(current_status, '') if current_status else ''
        )
        user_order = request.GET.get('user_order') or 'newest'
        if user_order not in ('newest', 'oldest'):
            user_order = 'newest'
        extra_context['user_order_current'] = user_order
        extra_context.setdefault(
            "user_pagination",
            {
                "has_previous": False,
                "has_next": False,
                "previous_page": None,
                "next_page": None,
                "current_page": 1,
                "total_pages": 1,
            },
        )
        response = super().changelist_view(request, extra_context=extra_context)
        if hasattr(response, "context_data"):
            context = response.context_data
            cl = context.get("cl")
            pagination = {
                "has_previous": False,
                "has_next": False,
                "previous_page": None,
                "next_page": None,
                "current_page": 1,
                "total_pages": 1,
                "start_index": 0,
                "end_index": 0,
            }
            if cl is not None:
                paginator = getattr(cl, "paginator", None)
                total_pages = getattr(paginator, "num_pages", 1) or 1
                current_page = getattr(cl, "page_num", 1) or 1
                has_previous = current_page > 1
                has_next = total_pages and current_page < total_pages
                per_page = getattr(cl, "list_per_page", self.list_per_page)
                result_count = getattr(getattr(cl, "paginator", None), "count", 0) or 0
                start_index = 0
                end_index = 0
                if result_count:
                    start_index = ((current_page - 1) * per_page) + 1
                    end_index = min(start_index + per_page - 1, result_count)
                pagination.update(
                    {
                        "has_previous": has_previous,
                        "has_next": has_next,
                        "previous_page": current_page - 1 if has_previous else None,
                        "next_page": current_page + 1 if has_next else None,
                        "current_page": current_page,
                        "total_pages": total_pages,
                        "start_index": start_index,
                        "end_index": end_index,
                    }
                )
            context["user_pagination"] = pagination
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" and hasattr(response, "context_data"):
            fragment_html = render_to_string(
                "admin/users/includes/user_list_fragment.html",
                response.context_data,
                request=request,
            )
            cl = response.context_data.get("cl")
            pagination = response.context_data.get("user_pagination", {})
            result_count = getattr(getattr(cl, "paginator", None), "count", 0) or 0
            meta = {
                "result_count": result_count,
                "status_current": request.GET.get("client_status", "") or "",
                "status_label": ClientStatusFilter.LOOKUP_TO_LABEL.get(request.GET.get("client_status", ""), ""),
                "order_current": getattr(request, "_user_order_choice", "newest"),
                "page": pagination.get("current_page", 1),
                "has_next": pagination.get("has_next", False),
                "has_previous": pagination.get("has_previous", False),
                "next_page": pagination.get("next_page"),
                "previous_page": pagination.get("previous_page"),
                "start_index": pagination.get("start_index", 0),
                "end_index": pagination.get("end_index", 0),
                "total_pages": pagination.get("total_pages", 1),
            }
            return JsonResponse({"html": fragment_html, "meta": meta})
        return response

    def import_users_view(self, request):
        opts = self.model._meta
        form = UserImportUploadForm()
        if request.method == "POST":
            form = UserImportUploadForm(request.POST, request.FILES)
            if form.is_valid():
                uploaded = form.cleaned_data["import_file"]
                try:
                    result = import_users_from_file(uploaded)
                except UserImportSchemaError as exc:
                    form.add_error("import_file", str(exc))
                except UserImportError as exc:
                    form.add_error(None, str(exc))
                else:
                    if result.created:
                        messages.success(request, f"Imported {result.created} user(s).")
                    if result.errors:
                        snippet = result.errors[:10]
                        list_html = format_html_join(
                            '',
                            '<li>Row {0} ({1}): {2}</li>',
                            (
                                (
                                    message.row_number,
                                    message.username or '-',
                                    message.message,
                                )
                                for message in snippet
                            ),
                        )
                        remaining = len(result.errors) - len(snippet)
                        tail = ""
                        if remaining > 0:
                            tail = format_html('<p>... and {} more row(s) with issues.</p>', remaining)
                        messages.error(
                            request,
                            format_html('Some rows could not be imported:<ul>{}</ul>{}', list_html, tail),
                        )

                    changelist_url = reverse(f'admin:{opts.app_label}_{opts.model_name}_changelist')
                    return redirect(changelist_url)

        context = {
            **self.admin_site.each_context(request),
            "opts": opts,
            "form": form,
            "title": "Import users",
        }
        return TemplateResponse(request, "admin/user_import.html", context)

    @method_decorator(csrf_exempt)
    def send_notification_view(self, request):
        if request.method == "POST":
            data = json.loads(request.body)
            user_id = data.get("user_id")
            message = data.get("message")

            user = UserProfile.objects.filter(id=user_id).first()
            if user:
                Notification.objects.create(
                    user=user,
                    message=message,
                    channel="email"  # или sms
                )
                return JsonResponse({"status": "ok"})

        return JsonResponse({"status": "error"}, status=400)
    @admin.display(boolean=True, description="Staff Status")
    def staff_status(self, instance):
        return instance.is_staff
    staff_status.boolean = True


# Unregister the default User admin and re-register with our custom one
admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)


@admin.register(MasterAvailability)
class MasterAvailabilityAdmin(ExportCsvMixin, admin.ModelAdmin):
    form = MasterAvailabilityForm
    list_display = ("master", "start_time", "end_time", "reason")
    list_filter = ("master",)
    search_fields = ("master__first_name", "master__last_name", "reason")
    export_fields = ["master", "start_time", "end_time", "reason"]

    def _redirect_to_appointments(self, request):
        url = reverse("admin:core_appointment_changelist")
        # переносим дату (и любые будущие параметры из календаря)
        passthrough = {}
        for key in ("date",):
            val = request.GET.get(key) or request.POST.get(key)
            if val:
                passthrough[key] = val
        if passthrough:
            url = f"{url}?{urlencode(passthrough)}"
        return redirect(url)

    # ---- после добавления ----
    def response_add(self, request, obj, post_url_continue=None):
        # всегда назад в календарь записей
        return self._redirect_to_appointments(request)

    # ---- после изменения ----
    def response_change(self, request, obj):
        # даже если нажали "Сохранить и продолжить" — уводим в календарь
        return self._redirect_to_appointments(request)

    # ---- после удаления ----
    def response_delete(self, request, obj_display, obj_id):
        return self._redirect_to_appointments(request)

    def has_add_permission(self, request):
        if request.user.is_superuser or request.user.is_staff:
            return True
        return hasattr(request.user, "master_profile")

    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.is_staff:
            return True
        if hasattr(request.user, "master_profile"):
            # список — разрешаем открыть; конкретный объект — только свой
            return True if obj is None else (obj.master_id == request.user.master_profile.id)
        return False

    # --- удаление: только свои (админ/стфф — любые)
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.is_staff:
            return True
        if hasattr(request.user, "master_profile"):
            return True if obj is None else (obj.master_id == request.user.master_profile.id)
        return False

    # --- только свои time off ---
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return qs.filter(master=request.user.master_profile)
        return qs
    def get_form(self, request, obj=None, **kwargs):
        BaseForm = super().get_form(request, obj, **kwargs)

        class WrappedForm(BaseForm):
            def __init__(self, *args, **kw):
                super().__init__(*args, **kw)
                # Только для мастеров (не суперюзеров/стффа)
                if hasattr(request.user, "master_profile") and not request.user.is_superuser:
                    if "master" in self.fields:
                        # визуально фиксируем поле
                        self.fields["master"].disabled = True
        return WrappedForm

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)

        date_str = request.GET.get("date")
        time_str = request.GET.get("time")
        const_master = request.GET.get("master")

        if date_str and time_str:
            try:
                combined = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                initial["start_time"] = make_aware(combined)

            except ValueError:
                pass
        if const_master:
            initial["master"] = const_master
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            initial["master"] = request.user.master_profile.id

        return initial


# -----------------------------
# Appointment Admin
# -----------------------------


def _can_override_discount_rule(user) -> bool:
    """Право на одновременную скидку услуги и промокод в одной позиции."""
    return bool(
        getattr(user, "is_superuser", False)
        or user.has_perm("core.bypass_service_discount_rule")  # NOTE: при желании задай такой perm
    )


def _can_edit_unit_price(user) -> bool:
    """Право редактировать индивидуальную цену позиции (unit_price)."""
    return bool(
        getattr(user, "is_superuser", False)
        or user.has_perm("core.can_edit_unit_price")  # NOTE: при желании задай такой perm
    )


# ──────────────────────────────────────────────────────────────────────────────
# Inline: AppointmentItem
# ──────────────────────────────────────────────────────────────────────────────

class AppointmentItemInline(admin.TabularInline):
    model = AppointmentItem
    form = AppointmentItemInlineForm
    fk_name = "appointment"
    extra = 0
    # autocomplete_fields = ["service", "master", "promocode"]  # если используете автокомплит

    def get_formset(self, request, obj=None, **kwargs):
        """
        obj — это родитель (Appointment). Передадим его в каждую форму через kwargs.
        А ещё скорректируем queryset для промокода, чтобы уже выбранный (даже неактивный) отображался.
        """
        parent = obj
        FormSet = super().get_formset(request, obj, **kwargs)

        class PrefillFormSet(FormSet):
            def __init__(self, *a, **kw):
                super().__init__(*a, **kw)
                # Если у вас в форме есть поле 'promocode' и вы фильтруете только активные,
                # добиваемся, чтобы выбранный ранее промокод тоже был в queryset:
                for form in self.forms:
                    if hasattr(form, "fix_promocode_queryset"):
                        form.fix_promocode_queryset()

            def _construct_form(self, i, **kw):
                # Передаём родителя в конструктор формы
                kw.setdefault("parent_obj", parent)
                return super()._construct_form(i, **kw)

        return PrefillFormSet

    def has_delete_permission(self, request, obj=None):

        return True


class AppointmentProductSaleInline(admin.StackedInline):
    model = ProductSale
    form = AppointmentProductSaleForm
    fk_name = "appointment"
    extra = 0
    autocomplete_fields = ("product", "sold_by", "client")
    verbose_name = "Product sale"
    verbose_name_plural = "Product sales"
    prefix = "product_sales"

    def formfield_for_foreignkey(self, db_field, request=None, **kwargs):
        if db_field.name == "client":
            kwargs["queryset"] = (
                UserProfile.objects.filter(userrole__role__name="Client")
                .select_related("user")
                .order_by("user__first_name", "user__last_name", "user__username")
                .distinct()
            )
        if db_field.name == "sold_by" and request is not None:
            profile = getattr(request.user, "userprofile", None)
            filters = Q(user__is_superuser=True)
            if profile:
                filters |= Q(pk=profile.pk)
            kwargs["queryset"] = (
                UserProfile.objects.select_related("user")
                .filter(filters)
                .order_by("user__first_name", "user__last_name", "user__username")
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        base_kwargs = getattr(formset, "form_kwargs", {}) or {}
        formset.form_kwargs = {**base_kwargs, "request": request, "appointment": obj}

        price_url = reverse("admin:core_productsale_product_price")
        product_field = formset.form.base_fields.get("product")
        if product_field:
            product_field.widget.attrs["data-price-endpoint"] = price_url
        return formset


# ──────────────────────────────────────────────────────────────────────────────
# CSV helpers
# ──────────────────────────────────────────────────────────────────────────────

def _money(x):
    return f"{x:.2f}" if x is not None else ""


# ──────────────────────────────────────────────────────────────────────────────
# AppointmentAdmin
# ──────────────────────────────────────────────────────────────────────────────


@admin.register(Appointment)
class AppointmentAdmin(ExportXlsxMixin, admin.ModelAdmin):
    """
    Полнофункциональная админка:
      • обычный список + календарный вид (?view=calendar)
      • AJAX JSON эндпоинт для событий календаря
      • CSV-экшены (по приёмам и по позициям)
      • inline позиций с валидаторами/правами
      • режим мастера: видит только свои записи, read-only, без действий
    """
    form = AppointmentAdminForm
    inlines = [AppointmentItemInline, AppointmentProductSaleInline]
    add_form = AppointmentAddForm
    # NOTE: если хочешь отдельный шаблон для календаря — задай его здесь
    change_list_template = "admin/appointments_calendar.html"  # твой базовый шаблон списка
    change_form_template = "admin/custom_edit_appointment.html"
    date_hierarchy = "start_time"  # поправь, если поле называется иначе

    list_select_related = ("client",)
    search_fields = (
        "id",
        "client__user__first_name",
        "client__user__last_name",
        "client__user__username",
    )

    ordering = ("-start_time",)
    autocomplete=["promocode",]
    readonly_fields = ("final_price", "card_processing_fee", "discount_source", "personal_discount_percent", "computed_total_readonly", "items_preview",)

    fieldsets = (
        (None, {
            "fields": (
                "client",
                "start_time",
                "end_time",           # если есть
                "status",             # если есть
                "payment_status",     # если есть
                "personal_discount",  # если есть
                "card_processing_fee",
                "computed_total_readonly",
                "items_preview",
                "notes",
            )
        }),
    )

    # ── РЕЖИМ МАСТЕРА: права и доступ ────────────────────────────────────────

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # if is_master(request.user):
        #     # Если мастер хранится в AppointmentItem (мульти-услуги)
        #     return qs.filter(items__master=master_obj(request.user)).distinct()
            # Если мастер хранится прямо в Appointment (одиночная услуга):
        # return qs.filter(master=master_obj(request.user))
        return qs


    def get_actions(self, request):
        actions = super().get_actions(request)
        if is_master(request.user):
            # мастерам скрываем CSV-экшены и любые массовые действия
            return {}
        return actions

    # ── QS и агрегаты ────────────────────────────────────────────────────────


    def get_changeform_initial_data(self, request):
        """
        Для страницы создания (add) — дать начальные значения.
        Для редактирования (change) Django сам подставит instance-данные.
        """
        from django.utils import timezone
        data = super().get_changeform_initial_data(request)

        # пример: округлим старт к ближайшим 15 минутам вперёд
        now = timezone.now()
        minute = (now.minute // 15 + 1) * 15
        if minute == 60:
            from datetime import timedelta
            now = (now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1))
        else:
            now = now.replace(minute=minute, second=0, microsecond=0)
        data.setdefault("start_time", now)

        # если у пользователя есть привязанный мастер — подставим его в initial
        # (мастер не сможет выбрать другого, это дополнительно контролируется на фронте и в POST)
        mp = MasterProfile.objects.filter(user=UserProfile.objects.filter(user=request.user).first()).first()
        if mp:
            data.setdefault("master", mp)

        return data

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}
        obj = None
        if object_id:
            obj = self.get_object(request, object_id)

        def _safe_reverse(name, *, kwargs=None):
            try:
                return reverse(name, kwargs=kwargs) if kwargs else reverse(name)
            except NoReverseMatch:
                return ""

        extra_context.setdefault("terminal_conn_token_url", _safe_reverse("terminal-conn-token"))
        if obj:
            extra_context.setdefault(
                "terminal_start_url",
                _safe_reverse("api-terminal-start", kwargs={"appt_id": obj.pk}),
            )
            extra_context.setdefault(
                "payment_verify_url",
                _safe_reverse("api-payment-verify", kwargs={"appt_id": obj.pk}),
            )
        else:
            extra_context.setdefault("terminal_start_url", "")
            extra_context.setdefault("payment_verify_url", "")

        return super().changeform_view(request, object_id, form_url, extra_context)

    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):

        ctx = dict(context)
        adminform = ctx.get("adminform")
        if adminform:
            ctx["form"] = adminform.form

        # отдаём инлайн-формсет позиций в шаблон
        # ── НАЙТИ inline formsets НАДЁЖНО ─────────────────
        items_inline = None
        sales_inline = None
        for inline in ctx.get("inline_admin_formsets", []):
            model = getattr(inline.opts, "model", None)
            if model is AppointmentItem:
                items_inline = inline
            elif model is ProductSale:
                sales_inline = inline

        if items_inline is not None:
            ctx["items_formset"] = items_inline.formset
            ctx["items_inline"] = items_inline
        if sales_inline is not None:
            ctx["product_sales_formset"] = sales_inline.formset
            ctx["product_sales_inline"] = sales_inline
            ctx["product_sales_prefix"] = sales_inline.formset.prefix
            ctx["product_sales_can_delete"] = sales_inline.formset.can_delete

        sale_forms_to_render = []
        if sales_inline is not None:
            for form in sales_inline.formset.forms:
                instance_exists = getattr(form.instance, "pk", None)
                if form.is_bound:
                    if form.errors or not form.empty_permitted or form.has_changed():
                        sale_forms_to_render.append(form)
                elif instance_exists:
                    sale_forms_to_render.append(form)
        ctx["product_sales_forms"] = sale_forms_to_render

        sale_defaults = {"quantity": 1}
        if obj and getattr(obj, "client", None):
            client_obj = obj.client
            client_user = getattr(client_obj, "user", None)
            raw_client_label = (
                getattr(client_user, "get_full_name", lambda: "")() or getattr(client_user, "username", "") or str(client_obj)
            )
            client_label = (str(raw_client_label) or "").strip() or str(client_obj)
            sale_defaults["client"] = {
                "id": str(obj.client_id),
                "label": client_label,
            }
        profile = getattr(request.user, "userprofile", None)
        if profile:
            user = getattr(profile, "user", None)
            raw_sold_by_label = (
                getattr(user, "get_full_name", lambda: "")() or getattr(user, "username", "") or str(profile)
            )
            sold_by_label = (str(raw_sold_by_label) or "").strip() or str(profile)
            sale_defaults["sold_by"] = {
                "id": str(profile.pk),
                "label": sold_by_label,
            }
        ctx["product_sale_defaults"] = sale_defaults
        ctx["product_sale_price_endpoint"] = reverse("admin:core_productsale_product_price")
        # ===== данные для кастомных селектов =====
        # ограничим пул мастеров, если текущий пользователь — мастер
        mp = MasterProfile.objects.filter(user=UserProfile.objects.filter(user=request.user).first()).first()

        # мастера (id, название)
        # if is_master(request.user):
        #     masters_qs = MasterProfile.objects.select_related("user").filter(pk=mp.pk)
        # else:
        masters_qs = MasterProfile.objects.select_related("user").all()

        masters_qs = masters_qs.order_by("user__user__first_name", "user__user__last_name")
        masters = [{"id": str(m.id), "name": str(m)} for m in masters_qs]

        today = timezone.now().date()
        svc_discounts = {}
        for sd in ServiceDiscount.objects.filter(start_date__lte=today, end_date__gte=today).select_related("service"):
            svc_discounts[str(sd.service_id)] = int(sd.discount_percent)

        # карта мастер → [услуги]
        ms_map = defaultdict(list)
        for sm in ServiceMaster.objects.select_related("service", "service__category", "master").order_by("service__name"):
            sid = str(sm.service_id)
            service = sm.service
            duration_min = service.duration_min or 0
            extra_min = service.extra_time_min or 0
            total_duration = duration_min + extra_min
            ms_map[str(sm.master_id)].append({
                "id": str(sm.service_id),
                "name": service.name,
                "base_price": str(service.base_price),
                "duration_min": duration_min,
                "extra_time_min": extra_min,
                "total_duration_min": total_duration,
                "svc_disc": svc_discounts.get(sid, 0),  # %
                "is_taxable": bool(service.is_taxable),
                "category": service.category.name if service.category_id else "Other",
            })

        # промокоды
        promos_by_service = defaultdict(list)
        promos_global = []
        qs = PromoCode.objects.filter(active=True, start_date__lte=today, end_date__gte=today).prefetch_related("applicable_services")
        for pc in qs:
            payload = {"id": str(pc.pk), "text": pc.code, "discount": int(pc.discount_percent)}
            services = list(pc.applicable_services.all())
            if services:
                for s in services:
                    promos_by_service[str(s.pk)].append(payload)
            else:
                promos_global.append(payload)

        appointment_pricing_snapshot = None
        item_pricing_map: Dict[str, Any] = {}
        if obj:
            try:
                appointment_pricing_snapshot = compute_appointment_pricing(obj)
            except PricingComputationError:
                appointment_pricing_snapshot = None

        if appointment_pricing_snapshot:
            ctx["appointment_pricing_snapshot"] = appointment_pricing_snapshot
            if items_inline is not None:
                lookup = {
                    item_data.get("id"): item_data
                    for item_data in appointment_pricing_snapshot.get("items", [])
                    if item_data.get("id")
                }
                mapped: Dict[str, Any] = {}
                for form in items_inline.formset.forms:
                    inst_pk = getattr(form.instance, "pk", None)
                    if inst_pk:
                        data = lookup.get(str(inst_pk))
                        if data:
                            mapped[form.prefix] = data
                item_pricing_map = mapped
            ctx["item_pricing_map"] = item_pricing_map
            ctx["currency_code"] = appointment_pricing_snapshot.get("currency") or getattr(settings, "CURRENCY_CODE", "CAD")
            ctx["currency_symbol"] = appointment_pricing_snapshot.get("currency_symbol") or "CA$"
        else:
            ctx["item_pricing_map"] = {}
            currency_code_current = ctx.get("currency_code") or getattr(settings, "CURRENCY_CODE", "CAD")
            ctx["currency_symbol"] = {
                "cad": "CA$",
                "usd": "$",
            }.get(str(currency_code_current).lower(), f"{str(currency_code_current).upper()} ")

        paid_total = Decimal("0.00")
        if obj:
            paid_agg = obj.payments.filter(status="succeeded").aggregate(
                total=Coalesce(
                    Sum(
                        F("amount_received") - F("amount_refunded"),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    ),
                    Value(Decimal("0.00")),
                )
            )
            paid_total = Decimal(paid_agg.get("total") or Decimal("0.00")).quantize(TWOPLACES)
        ctx["paid_total"] = paid_total

        appointment_total_amount = Decimal("0.00")
        if obj:
            if appointment_pricing_snapshot:
                totals_payload = appointment_pricing_snapshot.get("totals") or {}
                total_value = (
                    totals_payload.get("grand_total")
                    or totals_payload.get("final_price_recorded")
                    or totals_payload.get("pre_fee_total")
                    or Decimal("0.00")
                )
                appointment_total_amount = Decimal(str(total_value))
            else:
                appointment_total_amount = Decimal(getattr(obj, "final_price", Decimal("0.00")) or Decimal("0.00"))
        ctx["appointment_total_amount"] = appointment_total_amount.quantize(TWOPLACES)
        outstanding_amount = (appointment_total_amount - paid_total).quantize(TWOPLACES)
        if outstanding_amount < Decimal("0.00"):
            outstanding_amount = Decimal("0.00")
        ctx["appointment_outstanding_amount"] = outstanding_amount

        ctx["card_fee_percent"] = str(CARD_PROCESSING_PERCENT)
        ctx["card_fee_fixed"] = str(CARD_PROCESSING_FIXED)

        try:
            ctx["payment_add_url"] = reverse("admin:core_payment_add")
        except NoReverseMatch:
            ctx["payment_add_url"] = ""

        intake_forms_overview: List[Dict[str, Any]] = []
        intake_required_ids: List[str] = []
        notifications = []
        if obj:
            submissions_map = {
                str(sub.form_id): sub
                for sub in ClientIntakeFormSubmission.objects.filter(appointment=obj).select_related("form")
            }
            required_forms_qs = (
                ClientIntakeForm.objects
                .filter(services__appointmentitem__appointment=obj)
                .distinct()
            )
            for intake_form in required_forms_qs:
                fid = str(intake_form.pk)
                intake_required_ids.append(fid)
                submission = submissions_map.get(fid)
                service_names = list(
                    intake_form.services
                    .filter(appointmentitem__appointment=obj)
                    .order_by("name")
                    .values_list("name", flat=True)
                    .distinct()
                )
                intake_forms_overview.append({
                    "id": fid,
                    "name": intake_form.name,
                    "description": intake_form.description,
                    "services": service_names,
                    "submitted": bool(submission),
                    "submitted_at": submission.submitted_at if submission else None,
                    "manage_url": reverse("admin:core_appointment_manage_form", args=[obj.pk, intake_form.pk]),
                })
            notifications = (
                Notification.objects
                .filter(appointment=obj)
                .select_related("user__user")
                .order_by("-sent_at", "-id")
            )

        ctx.update({
            "masters_data": masters,
            "ms_map_data": dict(ms_map),
            "svc_discounts_data": svc_discounts,
            "promos_by_service_data": dict(promos_by_service),
            "promos_global_data": promos_global,
            "APPT_FIELDS_1": ("client", "start_time", "payment_status", "current_status"),
            "intake_forms_overview": intake_forms_overview,
            "intake_required_ids": intake_required_ids,

            # === важные флаги для шаблонов/JS ===
            "is_master": is_master(request.user),
            "current_master_id": mp.id if mp else None,
            "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
            "gst_enabled": getattr(settings, "GST_ENABLED", True),
        })
        ctx.setdefault("currency_code", getattr(settings, "CURRENCY_CODE", "CAD"))
        ctx.setdefault("currency_symbol", {
            "cad": "CA$",
            "usd": "$",
        }.get(str(ctx["currency_code"]).lower(), f"{str(ctx['currency_code']).upper()} "))
        ctx["notifications"] = notifications
        return super().render_change_form(request, ctx, add=add, change=change, form_url=form_url, obj=obj)

    def manage_intake_form_view(self, request, appointment_id, form_id):
        appointment = get_object_or_404(Appointment, pk=appointment_id)
        if not self.has_change_permission(request, appointment):
            raise PermissionDenied

        intake_form = get_object_or_404(ClientIntakeForm, pk=form_id)
        submission = (
            ClientIntakeFormSubmission.objects
            .filter(appointment=appointment, form=intake_form)
            .first()
        )

        client_profile = getattr(appointment, "client", None)
        if submission:
            initial_data = submission.data or submission.raw_payload or {}
        else:
            initial_data = {}

        if request.method == "POST":
            bound_form = intake_form.build_bound_form(
                data=request.POST,
                files=request.FILES,
                client=client_profile,
            )
            if bound_form.is_valid():
                cleaned = _coerce_json(bound_form.cleaned_data)
                ClientIntakeFormSubmission.objects.update_or_create(
                    appointment=appointment,
                    client=client_profile,
                    form=intake_form,
                    defaults={
                        "submitted_by": request.user if getattr(request.user, "is_authenticated", False) else None,
                        "data": cleaned,
                        "raw_payload": cleaned,
                        "form_schema_snapshot": intake_form.normalized_schema(),
                        "schema_version": intake_form.schema_version,
                        "is_complete": True,
                    },
                )
                messages.success(request, f"{intake_form.name} saved.")
                return redirect("admin:core_appointment_change", appointment.pk)
        else:
            bound_form = intake_form.build_bound_form(
                data=None,
                files=None,
                initial=initial_data,
                client=client_profile,
            )

        related_services = (
            intake_form.services.filter(appointmentitem__appointment=appointment)
            .order_by("name")
            .values_list("name", flat=True)
            .distinct()
        )

        context = {
            **self.admin_site.each_context(request),
            "title": f"{intake_form.name} — {appointment}",
            "appointment": appointment,
            "intake_form": intake_form,
            "submission": submission,
            "form": bound_form,
            "related_services": list(related_services),
            "back_url": reverse("admin:core_appointment_change", args=[appointment.pk]),
        }
        return TemplateResponse(request, "admin/appointment_manage_form.html", context)



    def save_model(self, request, obj, form, change):
        # Админка валидирует формы, но мы дополнительно страхуемся:
        obj.full_clean()  # вызывает Appointment.clean()
        super().save_model(request, obj, form, change)
        new_status = form.cleaned_data.get("current_status")
        if not new_status:
            return

        last = (AppointmentStatusHistory.objects
                .filter(appointment=obj)
                .order_by("-set_at")             # или -created_at
                .values_list("status_id", flat=True)
                .first())
        if last != new_status.id:
            AppointmentStatusHistory.objects.create(
                appointment=obj,
                status=new_status,
                set_by=self._actor(request.user),
                set_at=timezone.now(),
            )

    def _actor(self, user):
        # Подгони под свой профиль:

        p = getattr(user, "userprofile", None)
        if p is not None:
            return p
        return user

    def save_formset(self, request, form, formset, change):
        # Забираем инстансы без сохранения
        instances = formset.save(commit=False)
        # Удаления — отдельно
        for deleted in formset.deleted_objects:
            deleted.delete()

        # Прогоняем full_clean() на каждом дочернем объекте
        for inst in instances:
            if isinstance(inst, ProductSale):
                if form.instance and not inst.client_id and getattr(form.instance, "client_id", None):
                    inst.client_id = form.instance.client_id
                if not inst.sold_by_id:
                    profile = getattr(request.user, "userprofile", None)
                    if profile:
                        inst.sold_by = profile
                if not inst.sold_at:
                    base_dt = getattr(form.instance, "start_time", None)
                    inst.sold_at = base_dt or timezone.now()
            inst.full_clean()  # вызывает AppointmentItem.clean()
            inst.save()

        formset.save_m2m()
    def get_form(self, request, obj=None, **kwargs):
        if obj is None:
            kwargs["form"] = self.add_form
        else:
            kwargs["form"] = self.form
        return super().get_form(request, obj, **kwargs)


    def add_view(self, request, form_url='', extra_context=None):
        return self.custom_create_view(request)

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            # только клиент при создании
            return (("Client", {"fields": ("client",)}),)
        # на редактировании — ваша стандартная форма
        return (
            (None, {"fields": ("client", "start_time", "payment_status", "notes")}),
            ("Totals", {"fields": ("final_price", "discount_source", "personal_discount_percent"),
                        "classes": ("collapse",)}),
        )

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<uuid:appointment_id>/forms/<uuid:form_id>/",
                self.admin_site.admin_view(self.manage_intake_form_view),
                name="core_appointment_manage_form",
            ),
            path(
                "create/custom/",
                self.admin_site.admin_view(self.custom_create_view),
                name="core_appointment_custom_create",
            ),
            path(
                "api/appointment/<uuid:appointment_id>/enable-card-fee/",
                self.admin_site.admin_view(self.enable_card_fee_view),
                name="core_appointment_enable_card_fee",
            ),
        ]
        return custom + urls

    def enable_card_fee_view(self, request, appointment_id):
        if request.method != "POST":
            return HttpResponseBadRequest("POST required")
        if not self.has_change_permission(request):
            raise PermissionDenied

        appointment = get_object_or_404(Appointment, pk=appointment_id)

        try:
            snapshot = compute_appointment_pricing(appointment)
        except PricingComputationError:
            appointment.recompute_totals(save=True)
            snapshot = compute_appointment_pricing(appointment)

        totals = snapshot.get("totals") or {}
        pre_fee_raw = totals.get("pre_fee_total") or totals.get("final_subtotal") or Decimal("0.00")
        try:
            pre_fee_total = Decimal(pre_fee_raw)
        except Exception:
            pre_fee_total = Decimal("0.00")
        if pre_fee_total < Decimal("0.00"):
            pre_fee_total = Decimal("0.00")

        paid_total = payment_services.get_total_received_for_appointment(appointment)
        current_total = Decimal(appointment.final_price or Decimal("0.00"))
        outstanding_due = (current_total - paid_total).quantize(TWOPLACES)

        if outstanding_due <= Decimal("0.00"):
            fee = Decimal("0.00")
        else:
            fee = card_processing_fee(outstanding_due).quantize(TWOPLACES)

        existing_fee = Decimal(getattr(appointment, "card_processing_fee", Decimal("0.00")) or Decimal("0.00"))
        total_fee = (existing_fee + fee).quantize(TWOPLACES)

        appointment.apply_card_processing_fee = True
        appointment.card_processing_fee = total_fee
        appointment.final_price = (pre_fee_total + total_fee).quantize(TWOPLACES)
        appointment.save(update_fields=["apply_card_processing_fee", "card_processing_fee", "final_price"])

        return JsonResponse(
            {
                "ok": True,
                "fee": f"{total_fee:.2f}",
                "grand_total": f"{appointment.final_price or Decimal('0.00'):.2f}",
                "apply_card_processing_fee": True,
            }
        )


    def _calendar_date_for_obj(self, obj) -> str | None:
        start = getattr(obj, "start_time", None)
        if not start:
            return None
        try:
            localized = localtime(start)
        except Exception:
            localized = start
        try:
            return localized.date().isoformat()
        except Exception:
            return None

    def _redirect_to_calendar(self, request, *, date: str | None = None) -> HttpResponseRedirect:
        url = reverse("admin:core_appointment_changelist")
        passthrough: dict[str, str] = {}
        if date:
            passthrough["date"] = date
        else:
            existing = request.GET.get("date") or request.POST.get("date")
            if existing:
                passthrough["date"] = existing
        if passthrough:
            url = f"{url}?{urlencode(passthrough)}"
        return redirect(url)

    def _default_payment_status_id(self):
        obj, _ = PaymentStatus.objects.get_or_create(name="Not Paid")
        return obj.id

    def _context_lists(self):
        clients_raw = (
            UserProfile.objects.select_related("user")
            .annotate(
                first_name=Coalesce("user__first_name", Value("")),
                last_name=Coalesce("user__last_name", Value("")),
                username=Coalesce("user__username", Value("")),
            )
            .values("id", "first_name", "last_name", "username")
            .order_by("user__first_name", "user__last_name", "user__username")
        )
        clients = []
        for row in clients_raw:
            label = f"{row['first_name']} {row['last_name']}".strip()
            if not label:
                label = row["username"] or str(row["id"])
            clients.append({"id": row["id"], "label": label})
        masters = (
            MasterProfile.objects.select_related("user")
            .annotate(label=Concat("user__user__first_name", Value(" "), "user__user__last_name"))
            .values("id", "label")
            .order_by("label")
        )

        service_forms_map: Dict[str, List[str]] = {}
        intake_forms_map: Dict[str, ClientIntakeForm] = {}
        services_info: Dict[str, Dict[str, Any]] = {}

        services_qs = (
            Service.objects.filter(is_active=True)
            .select_related("category")
            .prefetch_related(
                Prefetch(
                    "pre_appointment_forms",
                    queryset=ClientIntakeForm.objects.filter(is_active=True),
                )
            )
        )
        for svc in services_qs:
            sid = str(svc.pk)
            active_forms = list(svc.active_forms())
            form_ids = [str(form.pk) for form in active_forms]
            category_name = svc.category.name if svc.category_id else "Other"
            service_forms_map[sid] = form_ids
            services_info[sid] = {
                "id": sid,
                "name": svc.name,
                "base_price": str(svc.base_price),
                "duration_min": svc.duration_min,
                "extra_time_min": svc.extra_time_min,
                "total_duration_min": (svc.duration_min or 0) + (svc.extra_time_min or 0),
                "forms": form_ids,
                "category": category_name,
                "is_taxable": bool(svc.is_taxable),
            }
            for form in active_forms:
                intake_forms_map[str(form.pk)] = form

        services_by_master: Dict[str, List[Dict[str, Any]]] = {}
        qs = (
            ServiceMaster.objects
            .filter(service__is_active=True)
            .select_related("service", "service__category", "master")
            .values(
                "master_id",
                "service__id",
                "service__name",
                "service__base_price",
                "service__duration_min",
                "service__extra_time_min",
                "service__is_taxable",
                "service__category__name",
                "service__category_id",
            )
        )
        for r in qs:
            sid = str(r["service__id"])
            base = services_info.get(sid)
            if not base:
                forms = service_forms_map.get(sid, [])
                category_name = r["service__category__name"] if r["service__category_id"] else "Other"
                base = {
                    "id": sid,
                    "name": r["service__name"],
                    "base_price": str(r["service__base_price"]),
                    "duration_min": r["service__duration_min"],
                    "extra_time_min": r["service__extra_time_min"],
                    "total_duration_min": (r["service__duration_min"] or 0) + (r["service__extra_time_min"] or 0),
                    "forms": forms,
                    "category": category_name,
                    "is_taxable": bool(r.get("service__is_taxable")),
                }
                services_info[sid] = base
            payload = {
                "id": base["id"],
                "name": base["name"],
                "base_price": base["base_price"],
                "duration_min": base["duration_min"],
                "extra_time_min": base.get("extra_time_min", 0),
                "total_duration_min": base.get("total_duration_min", (base.get("duration_min") or 0)),
                "forms": list(base.get("forms", [])),
                "category": base.get("category", "Other"),
                "is_taxable": bool(base.get("is_taxable", False)),
            }
            services_by_master.setdefault(str(r["master_id"]), []).append(payload)

        return (
            list(clients),
            list(masters),
            services_by_master,
            service_forms_map,
            intake_forms_map,
        )

    def _parse_start_dt(date_str: str | None, time_str: str | None):
        if not date_str or not time_str:
            return None
    # Популярные форматы
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                naive = datetime.strptime(f"{date_str} {time_str}", fmt)
                return timezone.make_aware(naive, timezone.get_current_timezone())
            except ValueError:
                pass
        # Fallback для ISO, если прилетит time с 'Z'
        try:
            t = time_str.replace("Z", "+00:00") if time_str.endswith("Z") else time_str
            dt = datetime.fromisoformat(f"{date_str}T{t}")
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except Exception:
            return None


    def custom_create_view(self, request):
        # ---------------- helpers ----------------
        def _parse_dt(date_str: str | None, time_str: str | None):
            if not date_str or not time_str:
                return None
            # популярные форматы
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
                try:
                    naive = datetime.strptime(f"{date_str} {time_str}", fmt)
                    return make_aware(naive, get_current_timezone())
                except ValueError:
                    pass
            # ISO fallback (если time приходит c 'Z' и т.п.)
            try:
                t = time_str.replace("Z", "+00:00") if time_str.endswith("Z") else time_str
                dt = datetime.fromisoformat(f"{date_str}T{t}")
                if dt.tzinfo is None:
                    dt = make_aware(dt, get_current_timezone())
                return dt
            except Exception:
                return None

        def _empty_error_bag():
            return {
                "__all__": [],                 # глобальные ошибки
                "fields": defaultdict(list),   # ошибки по полям формы (вне айтемов)
                "items": defaultdict(lambda: defaultdict(list)),  # ошибки по строкам айтемов
                "intake": defaultdict(lambda: defaultdict(list)),  # ошибки по формам-анкетам
            }

        def _serialize_validation_error(exc, bag):
            """
            Преобразует ValidationError (в т.ч. с message_dict) в наш bag.
            """
            if hasattr(exc, "message_dict"):
                for key, msgs in exc.message_dict.items():
                    msgs = msgs if isinstance(msgs, (list, tuple)) else [msgs]
                    if key.startswith("items-"):
                        # варианты: items-2-start_time или items-2
                        parts = key.split("-")
                        try:
                            idx = int(parts[1])
                        except Exception:
                            idx = None
                        if idx is not None:
                            if len(parts) >= 3:
                                field = "-".join(parts[2:])
                                bag["items"][idx][field].extend(msgs)
                            else:
                                bag["items"][idx]["__all__"].extend(msgs)
                        else:
                            bag["__all__"].extend(msgs)
                    elif key in ("__all__", "non_field_errors"):
                        bag["__all__"].extend(msgs)
                    elif key.startswith("intake:"):
                        parts = key.split(":")
                        if len(parts) >= 2:
                            fintake = parts[1]
                            field_name = parts[2] if len(parts) >= 3 else "__all__"
                            bag["intake"][fintake][field_name].extend(msgs)
                        else:
                            bag["__all__"].extend(msgs)
                    else:
                        bag["fields"][key].extend(msgs)
            else:
                msgs = exc.messages if hasattr(exc, "messages") else [str(exc)]
                bag["__all__"].extend(msgs)

        def _finalize_bag(bag):
            bag["fields"] = dict(bag["fields"])
            bag["items"] = {i: dict(fields) for i, fields in bag["items"].items()}
            bag["intake"] = {form_id: dict(fields) for form_id, fields in bag["intake"].items()}
            return bag

        def _context_lists():
            """Ваш существующий метод, оставляю как есть; если у вас уже есть — используйте его."""
            return self._context_lists()

        def _build_forms_catalog(forms_map):
            catalog = {}
            for fid, form in forms_map.items():
                catalog[fid] = {
                    "id": fid,
                    "name": form.name,
                    "slug": form.slug,
                    "description": form.description or "",
                    "version": form.schema_version,
                    "schema": form.normalized_schema(),
                    "updated_at": form.updated_at.isoformat() if form.updated_at else None,
                }
            return catalog

        # ---------------- GET: первичная отрисовка ----------------
        try:
            availability_url = reverse("api-availability")
        except NoReverseMatch:
            availability_url = ""

        mp = MasterProfile.objects.filter(user=UserProfile.objects.filter(user=request.user).first()).first()
        clients, masters, services_by_master, service_forms_map, intake_forms_map = _context_lists()
        intake_forms_catalog = _build_forms_catalog(intake_forms_map)

        if request.method == "GET":

            q_date = (request.GET.get("date") or "").strip()

            q_time = (request.GET.get("time") or "").strip()

            q_master = (request.GET.get("master") or "").strip()

            if q_master == "undefined":

                q_master = ""



            if is_master(request.user):

                q_master = str(mp.pk) if mp else ""

                masters = [m for m in masters if mp and str(m["id"]) == str(mp.pk)]



            initial_first_item = {}

            if q_master and MasterProfile.objects.filter(pk=q_master).exists():

                initial_first_item["master"] = str(q_master)



            dt = _parse_dt(q_date or None, q_time or None)

            if dt:

                initial_first_item["start_time_date"] = dt.strftime("%Y-%m-%d")

                initial_first_item["start_time_time"] = dt.strftime("%H:%M:%S" if dt.second else "%H:%M")
            else:

                if q_date:

                    initial_first_item.setdefault("start_time_date", q_date)

                if q_time:

                    initial_first_item.setdefault("start_time_time", q_time[:5])



            ctx = {

                **self.admin_site.each_context(request),

                "clients": list(clients),

                "masters": masters,

                "services_by_master": services_by_master,

                "initial_first_item": initial_first_item,

                "prefill_query": {"date": q_date or None, "time": q_time or None, "master": str(q_master) if q_master else None},

                "service_forms_map": service_forms_map,

                "intake_forms_catalog": intake_forms_catalog,

                "posted_intake_payload": {"forms": []},

                "intake_error_map": {},

                "form_errors": [],

                "field_errors": {},

                "item_errors": {},

                "posted_items": [],

                "posted_client": "",

                "is_master": is_master(request.user),

                "current_master_id": mp.id if mp else None,

                "availability_url": availability_url,
                "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
                "gst_enabled": getattr(settings, "GST_ENABLED", True),
                "currency_code": getattr(settings, "CURRENCY_CODE", "CAD"),
            }

            return TemplateResponse(request, "admin/custom_create_appointment.html", ctx)



        # ---------------- POST: ??????? ?????? ----------------

        clients, masters, services_by_master, service_forms_map, intake_forms_map = _context_lists()
        intake_forms_catalog = _build_forms_catalog(intake_forms_map)

        if is_master(request.user):
            masters = [m for m in masters if str(m["id"]) == str(mp.pk)]
        # соберём промокоды по сервисам (как у вас)
        promos_by_service: dict[str, list[dict]] = {}
        try:
            fk_qs = PromoCode.objects.filter(~Q(service=None)).select_related("service")
        except Exception:
            fk_qs = PromoCode.objects.none()

        for pc in fk_qs:
            sid = str(pc.service_id)
            promos_by_service.setdefault(sid, []).append({
                "id": str(pc.pk),
                "text": getattr(pc, "code", str(pc.pk)),
                "discount": getattr(pc, "discount_percent", None),
            })

        try:
            m2m_qs = PromoCode.objects.prefetch_related("applicable_services")
        except Exception:
            m2m_qs = PromoCode.objects.none()

        for pc in m2m_qs:
            services = getattr(pc, "applicable_services", None)
            if not services:
                continue
            for svc in services.all():
                sid = str(svc.pk)
                promos_by_service.setdefault(sid, []).append({
                    "id": str(pc.pk),
                    "text": getattr(pc, "code", str(pc.pk)),
                    "discount": getattr(pc, "discount_percent", None),
                })

        # dedup + sort
        for sid, items in promos_by_service.items():
            seen, uniq = set(), []
            for it in items:
                if it["id"] in seen:
                    continue
                seen.add(it["id"])
                uniq.append(it)
            promos_by_service[sid] = sorted(uniq, key=lambda x: (x["text"] or "").lower())

        intake_payload_raw = (request.POST.get("intake_payload") or "").strip()

        # state + первичная валидация
        bag = _empty_error_bag()
        intake_payload = {"forms": []}
        submitted_intake_map: dict[str, dict] = {}
        if intake_payload_raw:
            try:
                intake_payload = json.loads(intake_payload_raw) or {"forms": []}
            except json.JSONDecodeError:
                bag["__all__"].append("Unable to parse submitted intake forms.")
                intake_payload = {"forms": []}
        if isinstance(intake_payload, dict):
            entries = intake_payload.get("forms", [])
            if isinstance(entries, list):
                for entry in entries:
                    if not isinstance(entry, dict):
                        continue
                    fid = entry.get("id")
                    if fid:
                        submitted_intake_map[str(fid)] = entry
        else:
            bag["__all__"].append("Invalid intake forms payload structure.")
            intake_payload = {"forms": []}
        client_id = (request.POST.get("client") or "").strip()
        try:
            total_forms = int(request.POST.get("items-TOTAL_FORMS", 0))
        except Exception:
            total_forms = 0

        posted_items = []  # чтобы вернуть в шаблон ровно то, что вводили
        for i in range(total_forms):
            pref = f"items-{i}-"
            posted_items.append({
                "master":      (request.POST.get(pref + "master") or "").strip(),
                "service":     (request.POST.get(pref + "service") or "").strip(),
                "start_time_0": (request.POST.get(pref + "start_time_0") or "").strip(),  # date
                "start_time_1": (request.POST.get(pref + "start_time_1") or "").strip(),  # time
                "unit_price":  (request.POST.get(pref + "unit_price") or "").strip(),
                "promocode":   (request.POST.get(pref + "promocode") or "").strip(),
                "duration_override_min": (request.POST.get(pref + "duration_override_min") or "").strip(),
                "manual_discount_percent": (request.POST.get(pref + "manual_discount_percent") or "").strip(),
            })

        # обязательные поля верхнего уровня
        if not client_id:
            bag["fields"]["client"].append("Client is required.")
        if total_forms < 1:
            bag["__all__"].append("Add at least one service.")

        # построчная валидация ещё до создания объектов
        valid_rows = []
        for idx, row in enumerate(posted_items):
            # пропускаем пустые
            if not any(row.values()):
                continue

            master_id  = row["master"]
            service_id = row["service"]
            date_str   = row["start_time_0"]
            time_str   = row["start_time_1"]

            if is_master(request.user):
                if master_id and str(master_id) != str(mp.pk):
                    # фиксируем в UI, но и ошибку подсветим, чтобы было наглядно, почему перезатираем
                    bag["items"][idx]["master"].append("You can assign items only to yourself.")
                master_id = str(mp.pk)

            if not master_id:
                bag["items"][idx]["master"].append("Select master.")
            if not service_id:
                bag["items"][idx]["service"].append("Select service.")
            if not date_str:
                bag["items"][idx]["start_time_0"].append("Select date.")
            if not time_str:
                bag["items"][idx]["start_time_1"].append("Select time.")

            dt = _parse_dt(date_str, time_str)
            if date_str and time_str and not dt:
                bag["items"][idx]["start_time_1"].append("Invalid date/time.")

            duration_raw = row.get("duration_override_min", "").strip()
            duration_override = None
            if duration_raw:
                try:
                    duration_override = int(duration_raw)
                    if duration_override < 1:
                        raise ValueError
                except (TypeError, ValueError):
                    bag["items"][idx]["duration_override_min"].append("Enter a positive duration (minutes).")
                    duration_override = None

            manual_raw = row.get("manual_discount_percent", "").strip()
            manual_discount = None
            if manual_raw:
                try:
                    manual_discount = int(manual_raw)
                except (TypeError, ValueError):
                    bag["items"][idx]["manual_discount_percent"].append("Enter discount between 0 and 100.")
                    manual_discount = None
                else:
                    if manual_discount < 0 or manual_discount > 100:
                        bag["items"][idx]["manual_discount_percent"].append("Enter discount between 0 and 100.")
                        manual_discount = None
            else:
                manual_discount = 0

            valid_rows.append({
                "idx": idx,
                "master_id": master_id or None,
                "service_id": service_id or None,
                "dt": dt,
                "unit_price": (row["unit_price"] or None),
                "promocode_id": (row["promocode"] or None),
                "duration_override": duration_override,
                "manual_discount": manual_discount,
            })

        required_form_ids: set[str] = set()
        for row in valid_rows:
            service_id = row.get("service_id")
            if not service_id:
                continue
            required_form_ids.update(service_forms_map.get(str(service_id), []))

        client_profile = None
        if client_id:
            client_profile = UserProfile.objects.filter(pk=client_id).select_related("user").first()

        validated_intake: dict[str, dict] = {}
        for form_id in required_form_ids:
            form_obj = intake_forms_map.get(form_id)
            if not form_obj:
                continue
            submission_entry = submitted_intake_map.get(form_id)
            if not submission_entry:
                bag["intake"][form_id]["__all__"].append("This form must be completed before booking.")
                continue
            payload = submission_entry.get("data") or {}
            if not isinstance(payload, dict):
                bag["intake"][form_id]["__all__"].append("Invalid payload for this form.")
                continue
            bound_form = form_obj.build_bound_form(data=payload, client=client_profile)
            if not bound_form.is_valid():
                for field_name, errors in bound_form.errors.items():
                    target = field_name if field_name != "__all__" else "__all__"
                    bag["intake"][form_id][target].extend([str(err) for err in errors])
                continue
            validated_intake[form_id] = {
                "cleaned_data": bound_form.cleaned_data,
                "raw_data": payload,
                "form": form_obj,
            }

        # если уже есть ошибки — просто показать страницу с ними
        has_errors = bool(bag["__all__"] or bag["fields"] or bag["items"] or bag["intake"])
        if has_errors:
            ctx = {
                **self.admin_site.each_context(request),
                "clients": list(clients),
                "masters": masters,
                "services_by_master": services_by_master,
                "promos_by_service_json": json.dumps(promos_by_service),
                "form_errors": bag["__all__"],
                "field_errors": dict(bag["fields"]),
                "item_errors": {i: dict(v) for i, v in bag["items"].items()},
                "intake_error_map": {fid: dict(fields) for fid, fields in bag["intake"].items()},
                "posted_items": posted_items,
                "posted_client": client_id,
                "service_forms_map": service_forms_map,
                "intake_forms_catalog": intake_forms_catalog,
                "posted_intake_payload": intake_payload,


                "is_master": is_master(request.user),
                "current_master_id": mp.id if mp else None,
                "availability_url": availability_url,
                "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
                "gst_enabled": getattr(settings, "GST_ENABLED", True),
                "currency_code": getattr(settings, "CURRENCY_CODE", "CAD"),
            }
            return TemplateResponse(request, "admin/custom_create_appointment.html", ctx)

        # всё валидно — создаём
        try:
            with transaction.atomic():
                appt = Appointment(
                    client_id=client_id,
                    payment_status_id=self._default_payment_status_id(),
                )
                appt.full_clean()
                appt.save()


                row_errs = {}  # message_dict для последующего ValidationError
                prebuilt_items = []  # сюда сложим валидные, чтобы потом сохранить



                for row in valid_rows:
                    idx = row["idx"]
                    if not (row["master_id"] and row["service_id"] and row["dt"]):
                        # (сюда попадём только если кто-то внезапно пустой — но мы отфильтровали выше)
                        key = f"items-{idx}"
                        row_errs.setdefault(key, []).append("Incomplete row.")
                        continue

                    if is_master(request.user) and str(row["master_id"]) != str(mp.pk):
                        key = f"items-{idx}-master"
                        row_errs.setdefault(key, []).append("You can assign items only to yourself.")
                        continue
                    item = AppointmentItem(
                        appointment=appt,
                        master_id=row["master_id"],
                        service_id=row["service_id"],
                        start_time=row["dt"],
                        unit_price=row["unit_price"] or None,
                        duration_override_min=row.get("duration_override"),
                        manual_discount_percent=row.get("manual_discount", 0) or 0,
                    )
                    try:
                        item.full_clean()
                        prebuilt_items.append((idx, item, row.get("promocode_id") or ""))
                        item.save()
                    except ValidationError as e:
                        if hasattr(e, "message_dict"):
                            for field, msgs in e.message_dict.items():
                                msgs = msgs if isinstance(msgs, (list, tuple)) else [msgs]
                                if field in ("__all__", "non_field_errors"):
                                    key = f"items-{idx}"
                                    row_errs.setdefault(key, []).extend(msgs)
                                else:
                                    key = f"items-{idx}-{field}"
                                    row_errs.setdefault(key, []).extend(msgs)
                        else:
                            msgs = e.messages if hasattr(e, "messages") else [str(e)]
                            key = f"items-{idx}"
                            row_errs.setdefault(key, []).extend(msgs)

                    # Если хотя бы у одной строки есть ошибки — откатываем и показываем их в форме
                if row_errs:
                    raise ValidationError(row_errs)
                first_start = None
                for idx, item, promocode_id in prebuilt_items:
                    item.save()
                    if first_start is None or item.start_time < first_start:
                        first_start = item.start_time

                    promo_id = (promocode_id or "").strip()
                    if promo_id:
                        promo_obj = PromoCode.objects.filter(pk=promo_id).first()
                        if promo_obj:
                            AppointmentItemPromoCode.objects.update_or_create(
                                item=item, defaults={"promocode": promo_obj}
                            )
                    else:
                        AppointmentItemPromoCode.objects.filter(item=item).delete()

                if first_start:
                    appt.start_time = first_start

                if hasattr(appt, "recompute_totals"):
                    appt.recompute_totals(save=True)
                else:
                    appt.save(update_fields=["start_time"])

                status = AppointmentStatus.objects.filter(name="Booked").first()
                if status:
                    # аккуратно получаем профиль, если он есть
                    set_by = getattr(getattr(request.user, "userprofile", None), "pk", None)
                    AppointmentStatusHistory.objects.create(
                        appointment=appt,
                        status=status,
                        set_by=request.user.userprofile if set_by else None,
                    )

                if validated_intake:
                    profile_for_submission = getattr(appt, "client", None) or client_profile
                    if profile_for_submission is None and client_id:
                        profile_for_submission = UserProfile.objects.filter(pk=client_id).first()
                    for payload in validated_intake.values():
                        form_obj = payload["form"]
                        if profile_for_submission is None:
                            continue
                        ClientIntakeFormSubmission.objects.create(
                            form=form_obj,
                            client=profile_for_submission,
                            appointment=appt,
                            submitted_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                            data=_coerce_json(payload["cleaned_data"]),
                            raw_payload=_coerce_json(payload["raw_data"]),
                            form_schema_snapshot=form_obj.normalized_schema(),
                            schema_version=form_obj.schema_version,
                            is_complete=True,
                        )

            messages.success(request, "Appointment created.")
            target_date = self._calendar_date_for_obj(appt)
            return self._redirect_to_calendar(request, date=target_date)

        except ValidationError as ve:
            # Переносим ошибки из моделей в наш мешок и показываем в той же форме
            _serialize_validation_error(ve, bag)
            bag = _finalize_bag(bag)
            ctx = {
                **self.admin_site.each_context(request),
                "clients": clients,
                "masters": masters,
                "services_by_master": services_by_master,
                "promos_by_service_json": json.dumps(promos_by_service),
                "form_errors": bag["__all__"],
                "field_errors": bag["fields"],
                "item_errors": bag["items"],
                "intake_error_map": bag["intake"],
                "posted_items": posted_items,
                "posted_client": client_id,
                "service_forms_map": service_forms_map,
                "intake_forms_catalog": intake_forms_catalog,
                "posted_intake_payload": intake_payload,
                "is_master": is_master(request.user),
                "current_master_id": mp.id if mp else None,
                "availability_url": availability_url,
                "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
                "gst_enabled": getattr(settings, "GST_ENABLED", True),
                "currency_code": getattr(settings, "CURRENCY_CODE", "CAD"),
            }
            return TemplateResponse(request, "admin/custom_create_appointment.html", ctx)

        except IntegrityError as ie:
            # Ошибка БД — показываем как глобальную и не падаем 500
            bag["__all__"].append(f"Database error: {ie}")
            ctx = {
                **self.admin_site.each_context(request),
                "clients": clients,
                "masters": masters,
                "services_by_master": services_by_master,
                "promos_by_service_json": json.dumps(promos_by_service),
                "form_errors": bag["__all__"],
                "field_errors": dict(bag["fields"]),
                "item_errors": {i: dict(v) for i, v in bag["items"].items()},
                "intake_error_map": {fid: dict(fields) for fid, fields in bag["intake"].items()},
                "posted_items": posted_items,
                "posted_client": client_id,
                "service_forms_map": service_forms_map,
                "intake_forms_catalog": intake_forms_catalog,
                "posted_intake_payload": intake_payload,
                "is_master": is_master(request.user),
                "current_master_id": mp.id if mp else None,
                "availability_url": availability_url,
                "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
                "gst_enabled": getattr(settings, "GST_ENABLED", True),
                "currency_code": getattr(settings, "CURRENCY_CODE", "CAD"),
            }
            return TemplateResponse(request, "admin/custom_create_appointment.html", ctx)

        except Exception as e:
            # На проде — лог, а пользователю безопасно
            bag["__all__"].append("Unexpected error while creating appointment.")
            print("Error" + str(e))
            ctx = {
                **self.admin_site.each_context(request),
                "clients": clients,
                "masters": masters,
                "services_by_master": services_by_master,
                "promos_by_service_json": json.dumps(promos_by_service),
                "form_errors": bag["__all__"],
                "field_errors": dict(bag["fields"]),
                "item_errors": {i: dict(v) for i, v in bag["items"].items()},
                "intake_error_map": {fid: dict(fields) for fid, fields in bag["intake"].items()},
                "posted_items": posted_items,
                "posted_client": client_id,
                "service_forms_map": service_forms_map,
                "intake_forms_catalog": intake_forms_catalog,
                "posted_intake_payload": intake_payload,
                "is_master": is_master(request.user),
                "current_master_id": mp.id if mp else None,
                "availability_url": availability_url,
                "gst_percent": str(getattr(settings, "GST_PERCENT", Decimal("5.0"))),
                "gst_enabled": getattr(settings, "GST_ENABLED", True),
                "currency_code": getattr(settings, "CURRENCY_CODE", "CAD"),
            }
            return TemplateResponse(request, "admin/custom_create_appointment.html", ctx)

    def response_add(self, request, obj, post_url_continue=None):
        response = super().response_add(request, obj, post_url_continue)
        if IS_POPUP_VAR in request.POST or IS_POPUP_VAR in request.GET:
            return response
        return self._redirect_to_calendar(request, date=self._calendar_date_for_obj(obj))

    def response_post_save_add(self, request, obj):
        response = super().response_post_save_add(request, obj)
        if IS_POPUP_VAR in request.POST or IS_POPUP_VAR in request.GET:
            return response
        return self._redirect_to_calendar(request, date=self._calendar_date_for_obj(obj))

    def response_change(self, request, obj):
        response = super().response_change(request, obj)
        if IS_POPUP_VAR in request.POST or IS_POPUP_VAR in request.GET:
            return response
        if "_saveasnew" in request.POST:
            return self._redirect_to_calendar(request, date=self._calendar_date_for_obj(obj))
        return response

    @admin.display(description=_("Позиций"), ordering="_items_count")
    def items_count_display(self, obj):
        return getattr(obj, "_items_count", 0)

    @admin.display(description=_("Итого, $"), ordering="_total")
    def total_price_display(self, obj):
        return getattr(obj, "_total", Decimal("0"))

    @admin.display(description=_("Статус"))
    def status_display(self, obj):
        return getattr(obj, "status", None) or "—"

    @admin.display(description=_("Оплата"))
    def payment_status_display(self, obj):
        return getattr(obj, "payment_status", None) or "—"

    @admin.display(description=_("Состав"))
    def items_preview(self, obj):
        items_mgr = getattr(obj, "appointmentitem_set", None) or getattr(obj, "items", None)
        if not items_mgr:
            return "-"
        parts = []
        for it in items_mgr.all()[:6]:
            s_name = getattr(it.service, "name", str(it.service))
            m_name = getattr(it.master, "short_name", None) or getattr(it.master, "user", None) or "—"
            start = getattr(it, "start_time", None)
            fp = getattr(it, "final_price", None)
            frag = f"{s_name} | {m_name}"
            if start:
                frag += f" @ {start:%Y-%m-%d %H:%M}"
            if fp is not None:
                frag += f" — ${fp}"
            parts.append(frag)
        if items_mgr.count() > 6:
            parts.append("…")
        return " | ".join(parts)

    @admin.display(description=_("Итого (расчёт)"), ordering="_total")
    def computed_total_readonly(self, obj):
        return self.total_price_display(obj)

    # ── Сохранение и жёсткие проверки ────────────────────────────────────────

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        appt = form.instance
        mp = MasterProfile.objects.filter(user=UserProfile.objects.filter(user=request.user).first()).first()
        if is_master(request.user):
            for formset in formsets:
                if not isinstance(formset, BaseInlineFormSet):
                    continue
                model = getattr(getattr(formset, "model", None), "__name__", "")
                # интересует только инлайн с AppointmentItem
                if model != "AppointmentItem":
                    continue

        appt.recompute_totals(save=True)
        # Бизнес-правила (как и раньше — строгость сохранили):
        validate_appointment_has_items_on_save(appt)
        validate_items_prices_nonnegative(appt)
        # Если у тебя запрещены дубли услуг в одном приёме — оставь проверку:
        # (иначе — закомментируй следующую строку)
        # validate_no_duplicate_services_in_items(appt)
        validate_no_time_overlap_for_same_master(appt)


    # ── CSV действия (через твой миксин) ─────────────────────────────────────

    actions = ["export_appointments_xlsx", "export_appointment_items_xlsx"]



    # ── AJAX КАЛЕНДАРЬ (список → календарь + JSON) ──────────────────────────


    def changelist_view(self, request, extra_context=None):
        selected_date = request.GET.get('date')
        if selected_date:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            selected_date = timezone.localdate()

        services = Service.objects.all()
        appointment_statuses = AppointmentStatus.objects.all()
        payment_statuses = PaymentStatus.objects.all()

        appointments = AppointmentItem.objects.select_related(
            'appointment__client', 'service', 'master'
        ).prefetch_related(
            'appointment__items__service',
            'appointment__product_sales',
            Prefetch(
                'appointment__payments',
                queryset=Payment.objects.filter(status="succeeded").only("amount_received", "amount_refunded", "status"),
                to_attr='prefetched_succeeded_payments',
            ),
        )

        start_of_day = make_aware(datetime.combine(selected_date, datetime.min.time()))
        end_of_day = make_aware(datetime.combine(selected_date, datetime.max.time()))

        availabilities = MasterAvailability.objects.filter(
            start_time__lte=end_of_day,
            end_time__gte=start_of_day
        )

        if request.GET.get("service"):
            appointments = appointments.filter(service_id=request.GET["service"])
        if request.GET.get("status"):
            appointments = appointments.filter(
                appointment__appointmentstatushistory__status_id=request.GET["status"]
            )
        if request.GET.get("payment_status"):
            appointments = appointments.filter(
                appointment__payment_status_id__in=request.GET.getlist("payment_status")
            )
        master_ids = request.GET.getlist("master")
        if master_ids:
            appointments = appointments.filter(master_id__in=master_ids)

        masters = MasterProfile.objects.filter(
            id__in=appointments.values_list('master_id', flat=True)
        ).distinct()

        # Слоты по 15 минут
        start_hour = 8
        end_hour = 21
        slot_times = []
        time_pointer = datetime(2000, 1, 1, start_hour, 0)
        end_time = datetime(2000, 1, 1, end_hour, 0)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            action = request.GET.get("action")
            calendar_table = createTable(
                selected_date, time_pointer, end_time, slot_times, appointments, masters, availabilities
            )

            if action == "filter":
                html = render_to_string('admin/appointments_calendar_partial.html', {
                    "calendar_table": calendar_table,
                    'masters': masters,
                    'selected_masters': master_ids,
                })
                return JsonResponse({"html": html})

            elif action == "calendar":
                html = render_to_string('admin/appointments_calendar_partial.html', {
                    'calendar_table': calendar_table,
                    'masters': masters,
                    'selected_masters': master_ids,
                }, request=request)
                return JsonResponse({'html': html})

        calendar_table = createTable(
            selected_date, time_pointer, end_time, slot_times, appointments, masters, availabilities
        )

        response = super().changelist_view(request, extra_context=extra_context)

        if hasattr(response, "context_data"):
            context = response.context_data

            # Собираем URL для XLSX-экспорта (с сохранением текущих фильтров)
            try:
                opts = self.model._meta
                q = f"?{request.GET.urlencode()}" if request.GET else ""
                export_url_xlsx = reverse(
                    f'admin:{opts.app_label}_{opts.model_name}_export_xlsx'
                ) + q
            except NoReverseMatch:
                export_url_xlsx = None

            context.update({
                "calendar_table": calendar_table,
                "masters": masters,
                "selected_masters": master_ids,
                "selected_date": selected_date,
                "prev_date": (selected_date - timedelta(days=1)).strftime("%Y-%m-%d"),
                "next_date": (selected_date + timedelta(days=1)).strftime("%Y-%m-%d"),
                "today": timezone.localdate().strftime("%Y-%m-%d"),
                "services": services,
                "appointment_statuses": appointment_statuses,
                "payment_statuses": payment_statuses,
                "export_url_xlsx": export_url_xlsx,  # <-- вот это добавили
            })

        return response


# -----------------------------
# Appointment Status History Admin
# -----------------------------
@admin.register(AppointmentStatusHistory)
class AppointmentStatusHistoryAdmin(ExportCsvMixin,admin.ModelAdmin):
    """
    Admin interface for tracking status changes of appointments.
    """
    exclude = ('set_by',)
    list_display = ('appointment', 'status', 'set_by', 'set_at')
    list_filter = ('appointment', StaffSetByFilter)
    export_fields = ['appointment', 'status', 'set_by', 'set_at']
    def has_delete_permission(self, request, obj=None):
        # Суперадмин может всегда
        if request.user.is_superuser:
            return True
        # Мастер может удалять
        if hasattr(request.user, "master_profile"):
            return True
        return False


    def save_model(self, request, obj, form, change):
        if not obj.set_by_id:
            profile = getattr(request.user, "userprofile", None)
            if profile is None:
                profile, _ = UserProfile.objects.get_or_create(user=request.user)
            obj.set_by = profile
        super().save_model(request, obj, form, change)


# -----------------------------
# Payment Admin
# -----------------------------
@admin.register(Payment)
class PaymentAdmin(ExportCsvMixin ,admin.ModelAdmin):
    """
    Admin interface for payments.
    """
    list_display = (
        "appointment",
        "services_done_column",
        "amount",
        "status",
        "method",
        "receipt_column",
        "created_at",
    )
    list_filter = (
        "method",
        "status",
        "livemode",
        make_preset_date_filter(date_field="created_at", title=_("Period")),
    )
    search_fields = (
        'appointment__client__user__first_name',
        'appointment__client__user__last_name',
        'appointment__client__user__email',
        'appointment__items__master__user__user__first_name',
        'appointment__items__master__user__user__last_name',
        'appointment__items__service__name',
        'stripe_payment_intent_id',
        'stripe_charge_id',
    )
    readonly_fields = (
        'created_at', 'updated_at', 'stripe_payment_intent_id', 'stripe_charge_id',
        'stripe_payment_method_id', 'receipt_url', 'raw_response', 'metadata',
        'amount_received', 'amount_refunded', 'captured_at', 'livemode',
        'receipt_pdf', 'receipt_sent_at', 'resend_receipt_action',
    )
    export_fields = [
        'appointment', 'amount', 'currency', 'status', 'amount_received',
        'amount_refunded', 'method', 'livemode', 'stripe_payment_intent_id',
        'stripe_charge_id', 'created_at',
    ]
    actions = ["action_generate_receipts", "action_send_receipts"]

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        appointment_id = request.GET.get("appointment")
        amount = request.GET.get("amount")
        method_hint = (request.GET.get("method_hint") or "").lower()
        status_hint = (request.GET.get("status_hint") or "").strip().lower()

        if appointment_id:
            try:
                initial["appointment"] = Appointment.objects.get(pk=appointment_id)
            except Appointment.DoesNotExist:
                pass

        if amount:
            try:
                initial["amount"] = Decimal(amount)
            except (InvalidOperation, TypeError, ValueError):
                pass

        if method_hint:
            normalized = method_hint.replace("-", "").replace("_", "")
            if normalized == "cash":
                target_name = "Cash"
            elif "transfer" in normalized:
                target_name = "E-transfer"
            else:
                target_name = None

            if target_name:
                method = PaymentMethod.objects.filter(name__iexact=target_name).first()
                if method:
                    initial["method"] = method

        if status_hint:
            valid_statuses = {choice[0] for choice in Payment.STRIPE_STATUS_CHOICES}
            if status_hint in valid_statuses:
                initial["status"] = status_hint

        return initial

    def get_fields(self, request, obj=None):
        fields = list(super().get_fields(request, obj))
        if "resend_receipt_action" not in fields:
            fields.append("resend_receipt_action")
        return fields

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<uuid:payment_id>/resend-receipt/",
                self.admin_site.admin_view(self.resend_receipt_view),
                name="core_payment_resend_receipt",
            ),
        ]
        return custom_urls + urls

    def resend_receipt_view(self, request, payment_id):
        payment = self.get_object(request, str(payment_id))
        if not payment:
            self.message_user(request, "Payment not found.", level=messages.ERROR)
            return redirect("admin:core_payment_changelist")
        if not self.has_change_permission(request, payment):
            raise PermissionDenied
        email_payment_receipt_task.delay(str(payment.pk), force=True)
        self.message_user(request, "Receipt email queued for delivery.", level=messages.SUCCESS)
        return redirect("admin:core_payment_change", payment.pk)

    @admin.display(description="Receipt")
    def receipt_column(self, obj):
        if getattr(obj, "receipt_pdf", None):
            try:
                url = obj.receipt_pdf.url
            except Exception:
                url = None
            if url:
                return format_html('<a href="{}" target="_blank">Download</a>', url)
        if obj.receipt_url:
            return format_html('<a href="{}" target="_blank">Stripe</a>', obj.receipt_url)
        return "—"

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related("appointment", "method")
        item_qs = AppointmentItem.objects.select_related("service").order_by("start_time")
        return qs.prefetch_related(
            Prefetch("appointment__items", queryset=item_qs, to_attr="_admin_prefetched_items")
        )

    @admin.display(description="Services", ordering="appointment__items__service__name")
    def services_done_column(self, obj):
        appointment = getattr(obj, "appointment", None)
        if appointment is None:
            return "-"
        items = getattr(appointment, "_admin_prefetched_items", None)
        if items is None:
            items = appointment.items.select_related("service").all()
        names = [
            getattr(getattr(item, "service", None), "name", "") or ""
            for item in items
            if getattr(getattr(item, "service", None), "name", "")
        ]
        return ", ".join(dict.fromkeys(names)) if names else "-"

    @admin.display(description="Resend receipt")
    def resend_receipt_action(self, obj):
        if not obj or not obj.pk or obj.status != "succeeded":
            return "—"
        url = reverse("admin:core_payment_resend_receipt", args=[obj.pk])
        return format_html('<a class="button" href="{}">Send again</a>', url)

    @admin.action(description="Generate PDF receipts")
    def action_generate_receipts(self, request, queryset):
        succeeded = queryset.filter(status="succeeded")
        count = 0
        for payment in succeeded:
            generate_payment_receipt_task.delay(str(payment.pk))
            count += 1
        if count:
            self.message_user(request, f"Queued receipt generation for {count} payments.")
        else:
            self.message_user(request, "No succeeded payments selected.", level=messages.WARNING)

    @admin.action(description="Send/Resend receipt emails")
    def action_send_receipts(self, request, queryset):
        succeeded = queryset.filter(status="succeeded")
        count = 0
        for payment in succeeded:
            email_payment_receipt_task.delay(str(payment.pk), force=True)
            count += 1
        if count:
            self.message_user(request, f"Queued receipt emails for {count} payments.")
        else:
            self.message_user(request, "No succeeded payments selected.", level=messages.WARNING)


# -----------------------------
# Appointment Prepayment Admin
# -----------------------------
@admin.register(AppointmentPrepayment)
class AppointmentPrepaymentAdmin(ExportCsvMixin,admin.ModelAdmin):
    """
    Admin interface for prepayment options tied to appointments.
    """
    list_display = ('appointment', 'option')
    export_fields = ['appointment', 'option']

# -----------------------------
# Hidden Proxy Admin for CustomUserDisplay
# -----------------------------
@admin.register(CustomUserDisplay)
class CustomUserDisplayAdmin(admin.ModelAdmin):
    def _redirect_to_appointments(self, request):
        url = reverse("admin:core_appointment_changelist")
        # переносим дату (и любые будущие параметры из календаря)
        passthrough = {}
        for key in ("date",):
            val = request.GET.get(key) or request.POST.get(key)
            if val:
                passthrough[key] = val
        if passthrough:
            url = f"{url}?{urlencode(passthrough)}"
        return redirect(url)

    # ---- после добавления ----
    def response_add(self, request, obj, post_url_continue=None):
        # всегда назад в календарь записей
        return self._redirect_to_appointments(request)

    # ---- после изменения ----
    def response_change(self, request, obj):
        # даже если нажали "Сохранить и продолжить" — уводим в календарь
        return self._redirect_to_appointments(request)

    # ---- после удаления ----
    def response_delete(self, request, obj_display, obj_id):
        return self._redirect_to_appointments(request)


# -----------------------------
# Service Master Admin
# -----------------------------
@admin.register(ServiceMaster)
class ServiceMasterAdmin(ExportCsvMixin, admin.ModelAdmin):
    """
    Admin interface to assign masters to services.
    """
    form = ServiceMasterAdminForm
    list_display = ('master', 'service')
    search_fields = ('master__user__user__first_name', 'master__user__user__last_name', 'service__name')
    export_fields = ['master', 'service']

# -----------------------------
# Service Admin
# -----------------------------
@admin.register(Service)
class ServiceAdmin(ExportCsvMixin, admin.ModelAdmin):
    """
    Admin interface for services.
    """
    change_list_template = "admin/service/changelist_table.html"
    change_form_template = "admin/service/change_form.html"
    list_display = (
        'name',
        'base_price',
        'category',
        'allowed_rooms_display',
        'duration_min',
        'is_taxable',
        'is_active',
        'image_admin_thumb',
    )
    search_fields = ('name',)
    list_filter = (
        'is_active',
        'is_taxable',
        'category',
        ('allowed_rooms', admin.RelatedOnlyFieldListFilter),
    )
    filter_horizontal = ("pre_appointment_forms", "allowed_rooms")
    readonly_fields = ("image_preview",)
    list_per_page = 10
    fieldsets = (
        (None, {
            "fields": (
                "name",
                "description",
                "category",
                "allowed_rooms",
                "is_active",
                "is_taxable",
                "base_price",
                "duration_min",
                "extra_time_min",
            )
        }),
        ("Media", {
            "fields": (
                "image",
                "image_alt_text",
                "image_preview",
            )
        }),
        ("Pre-appointment forms", {
            "fields": ("pre_appointment_forms",),
        }),
    )
    actions = ["mark_active", "mark_inactive"]
    export_fields = [
        'name',
        'description',
        'base_price',
        'category',
        'allowed_rooms_display',
        'duration_min',
        'extra_time_min',
        'is_taxable',
    ]

    @admin.action(description="Mark selected services as active")
    def mark_active(self, request, queryset):
        queryset.update(is_active=True)

    @admin.action(description="Mark selected services as inactive")
    def mark_inactive(self, request, queryset):
        queryset.update(is_active=False)

    @admin.display(description="Rooms")
    def allowed_rooms_display(self, obj):
        cache = getattr(obj, "_prefetched_objects_cache", {})
        rooms = cache.get("allowed_rooms")
        if rooms is None:
            rooms = list(obj.allowed_rooms.all())
        labels = ", ".join(filter(None, (r.room for r in rooms)))
        return labels or "—"

    def get_queryset(self, request):
        qs = (
            super()
            .get_queryset(request)
            .select_related("category")
            .prefetch_related("allowed_rooms")
        )

        category_value = getattr(request, "_svc_category_filter", None)
        if category_value is None:
            category_value = request.GET.get("svc_category")
        if category_value:
            if category_value == "none":
                qs = qs.filter(category__isnull=True)
            else:
                try:
                    qs = qs.filter(category_id=int(category_value))
                except (TypeError, ValueError):
                    qs = qs.none()

        return qs.order_by("name", "pk")

    def get_search_results(self, request, queryset, search_term):
        """
        Apply a broader search across service name, description, and category.
        """
        term = (request.GET.get("q") or search_term or "").strip()
        if not term:
            return queryset, False
        filters = (
            Q(name__icontains=term)
            | Q(description__icontains=term)
            | Q(category__name__icontains=term)
        )
        return queryset.filter(filters), False

    def get_ordering(self, request):
        return ("name", "pk")

    def change_view(self, request, object_id, form_url="", extra_context=None):
        if request.method == "POST" and "_duplicate" in request.POST:
            original = self.get_object(request, object_id)
            if original is None:
                raise Http404(_("Service not found."))

            if not self.has_change_permission(request, original) or not self.has_add_permission(request):
                raise PermissionDenied

            clone = None
            try:
                with transaction.atomic():
                    service_model = self.model
                    base_name = original.name
                    new_name = base_name + "- Copy"

                    manager = service_model._default_manager
                    if manager.filter(name=new_name).exclude(pk=original.pk).exists():
                        suffix_index = 1
                        while True:
                            suffix = " (copy)" if suffix_index == 1 else f" (copy {suffix_index})"
                            candidate = f"{base_name}{suffix}"
                            if not manager.filter(name=candidate).exists():
                                new_name = candidate
                                break
                            suffix_index += 1

                    clone = service_model(
                        name=new_name,
                        description=original.description,
                        base_price=original.base_price,
                        duration_min=original.duration_min,
                        extra_time_min=original.extra_time_min,
                        category=original.category,
                        image=original.image,
                        image_alt_text=original.image_alt_text,
                        is_active=getattr(original, "is_active", True),
                    )
                    clone.save()
                    clone.pre_appointment_forms.set(original.pre_appointment_forms.all())
                    clone.allowed_rooms.set(original.allowed_rooms.all())
            except Exception:
                messages.error(request, _("Could not duplicate service."))
            else:
                if clone is not None:
                    self.log_addition(request, clone, f"Duplicated from Service {original.pk}.")
                messages.success(request, _("Service duplicated."))
                return redirect(reverse("admin:core_service_changelist"))

        return super().change_view(request, object_id, form_url=form_url, extra_context=extra_context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        original_params = request.GET.copy()
        search_term = (original_params.get("q") or "").strip()

        categories = ServiceCategory.objects.for_catalog()
        category_options = [{"value": "", "label": _("All Categories")}]
        category_options.append({"value": "none", "label": _("Uncategorised")})
        category_options.extend(
            {"value": str(cat.pk), "label": cat.name}
            for cat in categories
        )

        current_category = original_params.get("svc_category", "")
        if current_category:
            request._svc_category_filter = current_category
        else:
            request._svc_category_filter = None

        currency_code = (getattr(settings, "STRIPE_CURRENCY", "USD") or "USD").upper()
        currency_symbol = {
            "CAD": "CA$",
            "USD": "$",
            "EUR": "\u20AC",
            "GBP": "\u00A3",
        }.get(currency_code, f"{currency_code} $")

        extra_context.update(
            {
                "category_options": category_options,
                "current_category": current_category,
                "currency_symbol": currency_symbol,
                "current_search": search_term,
            }
        )
        extra_context.setdefault(
            "svc_pagination",
            {
                "has_previous": False,
                "has_next": False,
                "previous_page": None,
                "next_page": None,
                "current_page": 1,
                "total_pages": 1,
            },
        )

        cleaned_params = original_params.copy()
        cleaned_params.pop("svc_category", None)
        request.GET = cleaned_params
        try:
            response = super().changelist_view(request, extra_context=extra_context)
        finally:
            request.GET = original_params

        try:
            opts = self.model._meta
            export_url = reverse(f'admin:{opts.app_label}_{opts.model_name}_export_csv')
            if original_params:
                export_url += f"?{original_params.urlencode()}"
        except NoReverseMatch:
            export_url = None

        if hasattr(response, "context_data"):
            context = response.context_data
            context["export_url"] = export_url
            context["category_options"] = category_options
            context["current_category"] = current_category
            context["currency_symbol"] = currency_symbol
            cl = context.get("cl")
            pagination = {
                "has_previous": False,
                "has_next": False,
                "previous_page": None,
                "next_page": None,
                "current_page": 1,
                "total_pages": 1,
                "start_index": 0,
                "end_index": 0,
            }
            if cl is not None:
                paginator = getattr(cl, "paginator", None)
                total_pages = getattr(paginator, "num_pages", 1) or 1
                current_page = getattr(cl, "page_num", 1) or 1
                has_previous = current_page > 1
                has_next = total_pages and current_page < total_pages
                per_page = getattr(cl, "list_per_page", self.list_per_page)
                result_count = getattr(getattr(cl, "paginator", None), "count", 0) or 0
                start_index = 0
                end_index = 0
                if result_count:
                    start_index = ((current_page - 1) * per_page) + 1
                    end_index = min(start_index + per_page - 1, result_count)
                pagination.update(
                    {
                        "has_previous": has_previous,
                        "has_next": has_next,
                        "previous_page": current_page - 1 if has_previous else None,
                        "next_page": current_page + 1 if has_next else None,
                        "current_page": current_page,
                        "total_pages": total_pages,
                        "start_index": start_index,
                        "end_index": end_index,
                    }
                )
            context["svc_pagination"] = pagination

        if request.headers.get("X-Requested-With") == "XMLHttpRequest" and hasattr(response, "context_data"):
            fragment_html = render_to_string(
                "admin/service/includes/service_list_fragment.html",
                response.context_data,
                request=request,
            )
            cl = response.context_data.get("cl")
            pagination = response.context_data.get("svc_pagination", {})
            result_count = getattr(getattr(cl, "paginator", None), "count", 0) or 0
            category_label = ""
            for option in category_options:
                if option["value"] == current_category:
                    category_label = option["label"]
                    break

            meta = {
                "result_count": getattr(cl, "result_count", result_count),
                "page": getattr(cl, "page_num", 0),
                "current_page": pagination.get("current_page", 1),
                "total_pages": pagination.get("total_pages", 1),
                "has_previous": pagination.get("has_previous", False),
                "has_next": pagination.get("has_next", False),
                "previous_page": pagination.get("previous_page"),
                "next_page": pagination.get("next_page"),
                "start_index": pagination.get("start_index", 0),
                "end_index": pagination.get("end_index", 0),
                "current_category": current_category,
                "current_category_label": category_label,
                "currency_symbol": currency_symbol,
                "search_term": search_term,
            }
            return JsonResponse({"html": fragment_html, "meta": meta})

        return response

    @admin.display(description="Preview", ordering=False)
    def image_admin_thumb(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="height:48px;width:85px;object-fit:cover;border-radius:6px;" alt="{}"/>',
                               obj.image.url,
                               obj.card_image_alt)
        return "-"

    @admin.display(description="Current preview")
    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-width:320px;border-radius:10px;" alt="{}"/>',
                               obj.image.url,
                               obj.card_image_alt)
        return "-"


@admin.register(ClientIntakeForm)
class ClientIntakeFormAdmin(admin.ModelAdmin):
    form = ClientIntakeFormAdminForm
    list_display = ("name", "slug", "is_active", "is_universal", "updated_at")
    list_filter = ("is_active", "is_universal")
    search_fields = ("name", "slug", "description")
    prepopulated_fields = {"slug": ("name",)}
    fieldsets = (
        (None, {"fields": ("name", "slug", "description", "is_active", "is_universal")}),
        (_("Builder"), {"fields": ("schema", "schema_version")} ),
        (_("Form class"), {"fields": ("form_class",), "classes": ("collapse",)}),
        (_("Timestamps"), {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )
    readonly_fields = ("created_at", "updated_at")


@admin.register(ClientIntakeFormSubmission)
class ClientIntakeFormSubmissionAdmin(admin.ModelAdmin):
    list_display = ("form", "client", "appointment", "assignment", "submitted_by", "submitted_at")
    list_filter = ("form", "submitted_at", "submitted_by", "assignment__form")
    search_fields = (
        "form__name",
        "client__user__first_name",
        "client__user__last_name",
        "appointment__id",
    )
    readonly_fields = (
        "form",
        "client",
        "appointment",
        "assignment",
        "submitted_by",
        "submitted_at",
        "data",
        "raw_payload",
        "form_schema_snapshot",
        "schema_version",
        "is_complete",
    )


@admin.register(ClientIntakeAssignment)
class ClientIntakeAssignmentAdmin(admin.ModelAdmin):
    list_display = ("form", "client", "is_completed", "assigned_by", "assigned_at", "completed_at")
    list_filter = ("form__is_universal", "form", "assigned_at", "completed_at")
    search_fields = (
        "client__user__first_name",
        "client__user__last_name",
        "client__user__email",
        "form__name",
    )
    autocomplete_fields = ("form", "client", "completed_by")
    readonly_fields = ("assigned_at", "assigned_by")

    @admin.display(boolean=True, description="Completed")
    def is_completed(self, obj):
        return obj.is_completed

    def save_model(self, request, obj, form, change):
        if not obj.assigned_by_id:
            obj.assigned_by = request.user
        super().save_model(request, obj, form, change)
# -----------------------------
# Notification Admin
# -----------------------------
@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    """
    Admin interface for notifications (email/SMS).
    """
    list_display = ('user', 'appointment', 'channel', 'short_message')
    list_filter = (('sent_at', DateFieldListFilter), 'channel')
    search_fields = ('user__user__first_name', 'user__user__last_name', 'appointment__items__service__name')
    ordering = ['-sent_at']

    @admin.display(description="message")
    def short_message(self, obj):
        """
        Truncates long messages to first 10 words.
        """
        words = obj.message.split()
        return ' '.join(words[:10]) + ('...' if len(words) > 10 else '')


@admin.register(ReminderSchedule)
class ReminderScheduleAdmin(admin.ModelAdmin):
    list_display = ("name", "is_active", "offset_amount", "offset_unit", "email_subject", "slug")
    list_filter  = ("is_active", "offset_unit")
    # search_fields = ("name", "slug", "email_subject", "email_template")
    fields = (
        "name", "slug", "is_active",
        "offset_amount", "offset_unit",
        "email_subject", "email_template",
    )

# -----------------------------
# Client File Admin
# -----------------------------
@admin.register(ClientFile)
class ClientFileAdmin(admin.ModelAdmin):
    """
    Admin interface for managing user-uploaded files.
    """
    list_display = ('user', "uploaded_by" ,'file_type', 'file')
    fields = ('user', 'file',"uploaded_by", 'file_type')
    readonly_fields = ('file_type',)  # 👈 делаем только для чтения
    exclude = ('file_type',)  # 👈 скрываем из формы создания
    list_filter = (('uploaded_at', DateFieldListFilter), 'file_type')
    search_fields = ('user__user__first_name', 'user__user__last_name')
    ordering = ['-uploaded_at']



# -----------------------------
# Client Review Admin
# -----------------------------
@admin.register(ClientReview)
class ClientReviewAdmin(ExportCsvMixin ,admin.ModelAdmin):
    list_display = ("appointment", "get_client", "get_master", "rating", "created_at")
    search_fields = ("appointment__client__user__first_name", "appointment__client__user__last_name", "comment")
    list_filter = ("rating", "created_at")
    export_fields = ["appointment", "get_client", "get_master", "rating", "created_at"]
    @admin.display(description="Client")
    def get_client(self, obj):
        return obj.appointment.client.get_full_name()

    @admin.display(description="Master")
    def get_master(self, obj):
        return obj.appointment.master.user.get_full_name()


#-----------------------------
# Discounts Admin
#-----------------------------
@admin.register(ServiceDiscount)
class ServiceDiscountAdmin(ExportCsvMixin ,admin.ModelAdmin):
    list_display = ('service', 'discount_percent', 'start_date', 'end_date', 'is_active')
    list_filter = ('start_date', 'end_date', 'service')
    search_fields = ('service__name',)
    form = ServiceDiscountAdminForm
    export_fields = ['service', 'discount_percent', 'start_date', 'end_date']
    @admin.display(boolean=True)
    def is_active(self, obj):
        return obj.is_active()

#-----------------------------
# Promocode Admin
#-----------------------------
@admin.register(PromoCode)
class PromoCodeAdmin(ExportCsvMixin ,admin.ModelAdmin):
    list_display = ('code', 'discount_percent', 'start_date', 'end_date',)
    list_filter = ('start_date', 'end_date')
    form = PromoCodeAdminForm

    export_fields = ['code', 'applicable_services', 'discount_percent', 'start_date', 'end_date']

    def get_export_row(self, obj):
        services_manager = getattr(obj, "applicable_services", None)
        services_display = ""
        if services_manager is not None:
            if hasattr(services_manager, "all"):
                services_display = ", ".join(
                    str(service) for service in services_manager.all() if service is not None
                )
            else:
                services_display = str(services_manager)

        return [
            getattr(obj, "code", ""),
            services_display,
            getattr(obj, "discount_percent", None),
            getattr(obj, "start_date", None),
            getattr(obj, "end_date", None),
        ]

    @admin.display(boolean=True)
    def is_active(self, obj):
        return obj.is_active()


# -----------------------------
# Retail products & sales
# -----------------------------


class LowStockFilter(admin.SimpleListFilter):
    title = "Low stock status"
    parameter_name = "low_stock"

    def lookups(self, request, model_admin):
        return (
            ("1", "Needs restock"),
            ("0", "Sufficient stock"),
        )

    def queryset(self, request, queryset):
        value = self.value()
        if value == "1":
            return queryset.filter(
                low_stock_threshold__gt=0,
                quantity_in_stock__lte=F("low_stock_threshold"),
            )
        if value == "0":
            return queryset.exclude(
                low_stock_threshold__gt=0,
                quantity_in_stock__lte=F("low_stock_threshold"),
            )
        return queryset


class ProductCategoryAdmin(admin.ModelAdmin):
    list_display = ("name", "is_active", "created_at", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("name",)
    ordering = ("name",)
    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        (None, {"fields": ("name", "description", "is_active")}),
        ("Timestamps", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )


class ProductAdmin(ExportXlsxMixin, admin.ModelAdmin):
    import_template_name = "admin/products/import.html"
    list_display = (
        "name",
        "category",
        "sku",
        "brand",
        "supplier",
        "price",
        "quantity_in_stock",
        "low_stock_indicator",
        "is_active",
        "updated_at",
    )
    list_filter = ("is_active", "category", LowStockFilter)
    search_fields = ("name", "sku", "brand", "supplier")
    ordering = ("name",)
    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        (
            None,
            {"fields": ("name", "sku", "category", "brand", "supplier", "description", "is_active")},
        ),
        ("Measure", {"fields": ("measure_type", "measure_value")}),
        (
            "Pricing & Inventory",
            {"fields": ("cost_price", "price", "quantity_in_stock", "low_stock_threshold")},
        ),
        ("Timestamps", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )

    @admin.display(description="Low stock", boolean=True)
    def low_stock_indicator(self, obj):
        return obj.is_low_on_stock

    def get_urls(self):
        urls = super().get_urls()
        opts = self.model._meta
        custom = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_products_view),
                name=f"{opts.app_label}_{opts.model_name}_import",
            ),
        ]
        return custom + urls

    def import_products_view(self, request):
        form = ProductImportUploadForm(request.POST or None, request.FILES or None)
        changelist_url = reverse(f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist")
        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "form": form,
            "title": "Import products",
            "changelist_url": changelist_url,
        }

        if request.method == "POST" and form.is_valid():
            uploaded = form.cleaned_data["import_file"]
            try:
                result = import_products_from_file(uploaded)
            except ProductImportError as exc:
                form.add_error("import_file", str(exc))
            else:
                if result.created or result.updated:
                    messages.success(
                        request,
                        f"Imported {result.created} new products and updated {result.updated}.",
                    )
                else:
                    messages.info(request, "No products were imported. The file did not contain new data.")

                if result.errors:
                    preview = "; ".join(
                        f"Row {msg.row_number}: {msg.message}"
                        for msg in result.errors[:3]
                    )
                    if len(result.errors) > 3:
                        preview += f" (+{len(result.errors) - 3} more rows)"
                    messages.warning(request, f"Some rows were skipped: {preview}")

                return HttpResponseRedirect(changelist_url)

        return TemplateResponse(request, self.import_template_name, context)

    def _export_all_xlsx_view(self, request):
        queryset = self.get_queryset(request).select_related("category")
        headers = [
            "ID",
            "Name",
            "SKU",
            "Category",
            "Description",
            "Measure Type",
            "Measure Value",
            "Brand",
            "Supplier",
            "Cost Price",
            "Price",
            "Quantity In Stock",
            "Low Stock Threshold",
            "Low Stock?",
            "Active",
            "Created At",
            "Updated At",
        ]
        rows = [
            [
                product.pk,
                product.name,
                product.sku or "",
                product.category.name if product.category else "",
                product.description,
                product.measure_type,
                product.measure_value,
                product.brand,
                product.supplier,
                product.cost_price,
                product.price,
                product.quantity_in_stock,
                product.low_stock_threshold,
                product.is_low_on_stock,
                product.is_active,
                product.created_at,
                product.updated_at,
            ]
            for product in queryset
        ]
        return self._xlsx_response(
            "products.xlsx",
            "Products",
            headers,
            rows,
            money_cols={10, 11},
            datetime_cols={16, 17},
        )

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        try:
            opts = self.model._meta
            query_string = f"?{request.GET.urlencode()}" if request.GET else ""
            export_url = reverse(f"admin:{opts.app_label}_{opts.model_name}_export_xlsx") + query_string
            extra_context["export_url"] = export_url
            extra_context["export_label"] = "📤 Export XLSX"
        except NoReverseMatch:
            extra_context.setdefault("export_url", None)

        try:
            import_url = reverse(f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import")
            extra_context["import_url"] = import_url
            extra_context["import_label"] = "📥 Import products"
        except NoReverseMatch:
            extra_context.setdefault("import_url", None)

        return super().changelist_view(request, extra_context=extra_context)


class ProductSaleAdmin(ExportXlsxMixin, admin.ModelAdmin):
    form = ProductSaleAdminForm
    list_display = (
        "sold_at",
        "product",
        "quantity",
        "unit_price",
        "total_amount",
        "sold_by",
        "client",
        "appointment",
    )
    list_filter = (
        "product__category",
        "product",
        "sold_by",
        make_preset_date_filter(date_field="sold_at", title=_("Sold period")),
    )
    search_fields = (
        "product__name",
        "product__sku",
        "appointment__id",
        "client__user__first_name",
        "client__user__last_name",
        "client__user__username",
        "notes",
    )
    ordering = ("-sold_at", "-id")
    readonly_fields = ("total_amount", "created_at", "updated_at")
    autocomplete_fields = ("product", "sold_by", "client", "appointment")
    list_select_related = (
        "product",
        "sold_by__user",
        "client__user",
        "appointment__client__user",
    )
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "product",
                    "sold_by",
                    "client",
                    "appointment",
                    "sold_at",
                    "quantity",
                    "unit_price",
                    "total_amount",
                    "notes",
                )
            },
        ),
        ("Timestamps", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if obj:
            readonly.extend(
                [
                    "product",
                    "sold_by",
                    "client",
                    # allow updating appointment if needed post creation
                    "sold_at",
                    "quantity",
                    "unit_price",
                    "total_amount",
                ]
            )
        return readonly

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "client":
            kwargs["queryset"] = (
                UserProfile.objects.filter(userrole__role__name="Client")
                .select_related("user")
                .order_by("user__first_name", "user__last_name", "user__username")
                .distinct()
            )
        if db_field.name == "sold_by":
            current_profile = getattr(request.user, "userprofile", None)
            sold_by_filters = Q(user__is_superuser=True)
            if current_profile:
                sold_by_filters |= Q(pk=current_profile.pk)
            kwargs["queryset"] = (
                UserProfile.objects.select_related("user")
                .filter(sold_by_filters)
                .order_by(
                    "user__first_name",
                    "user__last_name",
                    "user__username",
                )
            )
        if db_field.name == "appointment":
            kwargs["queryset"] = (
                Appointment.objects.select_related("client__user")
                .order_by("-start_time")
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        profile = getattr(request.user, "userprofile", None)
        if not obj and profile and "sold_by" in form.base_fields:
            form.base_fields["sold_by"].initial = profile.pk

        product_field = form.base_fields.get("product")
        if product_field:
            product_field.widget.attrs["data-price-endpoint"] = reverse(
                "admin:core_productsale_product_price"
            )
        return form

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "product-price/",
                self.admin_site.admin_view(self.product_price_lookup),
                name="core_productsale_product_price",
            ),
        ]
        return custom + urls

    def product_price_lookup(self, request: HttpRequest):
        product_id = request.GET.get("product")
        if not product_id:
            return JsonResponse({"error": "Missing product id"}, status=400)
        try:
            product = Product.objects.only("price").get(pk=product_id)
        except Product.DoesNotExist:
            return JsonResponse({"error": "Product not found"}, status=404)
        return JsonResponse({"unit_price": str(product.price)})

    def _export_all_xlsx_view(self, request):
        queryset = (
            self.get_queryset(request)
            .select_related(
                "product__category",
                "sold_by__user",
                "client__user",
                "appointment",
            )
        )
        headers = [
            "ID",
            "Sold At",
            "Product",
            "Product SKU",
            "Category",
            "Quantity",
            "Unit Price",
            "Total Amount",
            "Sold By",
            "Client",
            "Appointment ID",
            "Notes",
            "Created At",
            "Updated At",
        ]
        rows = []
        for sale in queryset:
            product = getattr(sale, "product", None)
            product_category = getattr(product, "category", None)
            sold_by_profile = getattr(sale, "sold_by", None)
            sold_by_user = getattr(sold_by_profile, "user", None)
            client_profile = getattr(sale, "client", None)
            client_user = getattr(client_profile, "user", None)
            sold_by_display = ""
            if sold_by_user:
                sold_by_display = sold_by_user.get_full_name() or sold_by_user.username
            client_display = ""
            if client_user:
                client_display = client_user.get_full_name() or client_user.username
            rows.append(
                [
                    sale.pk,
                    sale.sold_at,
                    product.name if product else "",
                    product.sku if product and product.sku else "",
                    product_category.name if product_category else "",
                    sale.quantity,
                    sale.unit_price,
                    sale.total_amount,
                    sold_by_display,
                    client_display,
                    sale.appointment.pk if sale.appointment else "",
                    sale.notes or "",
                    sale.created_at,
                    sale.updated_at,
                ]
            )
        return self._xlsx_response(
            "product_sales.xlsx",
            "Product Sales",
            headers,
            rows,
            money_cols={7, 8},
            datetime_cols={2, 13, 14},
        )

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        try:
            opts = self.model._meta
            query_string = f"?{request.GET.urlencode()}" if request.GET else ""
            export_url = reverse(f"admin:{opts.app_label}_{opts.model_name}_export_xlsx") + query_string
            extra_context["export_url"] = export_url
            extra_context["export_label"] = "📤 Export XLSX"
        except NoReverseMatch:
            extra_context.setdefault("export_url", None)
        return super().changelist_view(request, extra_context=extra_context)



    def save_model(self, request, obj, form, change):
        if not obj.sold_by_id:
            profile = getattr(request.user, "userprofile", None)
            if profile:
                obj.sold_by = profile
        super().save_model(request, obj, form, change)


# -----------------------------
# Register remaining models directly
# -----------------------------
admin.site.register(ProductCategory, ProductCategoryAdmin)
admin.site.register(Product, ProductAdmin)
admin.site.register(ProductSale, ProductSaleAdmin)
admin.site.register(Role)
admin.site.register(UserRole)
admin.site.register(AppointmentStatus)
admin.site.register(PaymentMethod)
admin.site.register(ClientSource)
admin.site.register(MasterRoom)
admin.site.register(PrepaymentOption)
admin.site.register(PaymentStatus)
admin.site.register(CancellationReason)


@admin.register(ServiceCategory)
class ServiceCategoryAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "slug",
        "only_discounted_services",
        "featured_rank",
        "catalog_order",
        "catalog_position_preview",
    )
    list_editable = ("featured_rank", "catalog_order")
    list_filter = ("only_discounted_services",)
    search_fields = ("name",)
    ordering = ("name",)
    readonly_fields = ("slug",)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.for_catalog()

    @admin.display(description="Catalog order")
    def catalog_position_preview(self, obj):
        if obj.featured_rank:
            mapping = dict(FEATURED_CATEGORY_RANKS)
            return mapping.get(obj.featured_rank, obj.featured_rank)
        if obj.catalog_order is not None:
            return obj.catalog_order
        return "\u2014"

@admin.register(AppointmentItemPromoCode)
class AppointmentItemPromoCodeAdmin(admin.ModelAdmin):
    list_display = ["item", "promocode", "promocode__discount_percent", "promocode__start_date", "promocode__end_date"]



@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    form = UserProfileChangeForm

    def _user_admin_url(self, suffix, *args):
        user_model = get_user_model()
        return reverse(f"admin:{user_model._meta.app_label}_{user_model._meta.model_name}_{suffix}", args=args)

    def _redirect_to_user_admin(self, request, suffix, *args):
        base_url = self._user_admin_url(suffix, *args)
        query_string = request.META.get("QUERY_STRING")
        if query_string:
            base_url = f"{base_url}?{query_string}"
        return redirect(base_url)

    def add_view(self, request, form_url="", extra_context=None):
        return self._redirect_to_user_admin(request, "add")

    def change_view(self, request, object_id, form_url="", extra_context=None):
        profile = self.get_object(request, object_id)
        if profile and profile.user_id:
            return self._redirect_to_user_admin(request, "change", profile.user_id)
        return super().change_view(request, object_id, form_url, extra_context)
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
    search_fields = (
        "user__first_name",
        "user__last_name",
        "user__username",
        "user__email",
        "phone",
    )
    # Явно перечислим поля на форме
    @admin.display(description="First name")
    def user_first_name(self, obj):
        return getattr(getattr(obj, "user", None), "first_name", "")

    @admin.display(description="Last name")
    def user_last_name(self, obj):
        return getattr(getattr(obj, "user", None), "last_name", "")

    @admin.display(description="Email")
    def user_email(self, obj):
        return getattr(getattr(obj, "user", None), "email", "")

    fieldsets = (
        (None, {
            "fields": (
                "user_first_name",
                "user_last_name"
            )
        }),
        ("Personal Info", {
            "fields": (
                "phone", "user_email","birth_date", "postal_code",
                "how_heard",
            )
        }),
        ("Notes", {"fields": ("notes",)}),

        # ТЕ САМЫЕ ПОЛЯ — но они теперь ПРИСУТСТВУЮТ в форме (как виртуальные),
        # поэтому Django их корректно отрендерит в админке UserProfile.
        ("Health", {
            "classes": ("collapse", "wide"),
            "fields": (
                "has_allergies", "allergies_text", "gender",
                "chronic_conditions", "medications",
                "pregnant", "skin_sensitivity",
                "recent_procedures", "contraindications",
                "health_notes",
            ),
        }),

    )

    readonly_fields = ("user_first_name", "user_last_name", "user_email")


    @admin.display(description="Status")
    def client_status_col(self, obj):
        return obj.client_status

    def response_change(self, request, obj):
        if "_from_appointment" in request.GET:
            return redirect("admin:core_appointment_changelist")
        return super().response_change(request, obj)

    def response_delete(self, request, obj_display, obj_id):
        if "_from_appointment" in request.GET:
            return redirect("admin:core_appointment_changelist")
        return super().response_delete(request, obj_display, obj_id)

    def response_post_save_change(self, request, obj):
        if "_from_appointment" in request.GET:
            return redirect("admin:core_appointment_changelist")
        return super().response_post_save_change(request, obj)

    def get_readonly_fields(self, request, obj=None):
        # Для мастера — все поля read-only, кроме 'notes'
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            # Берём список всех полей формы и исключаем 'notes'
            all_fields = list(self.fields) if self.fields else [f.name for f in self.model._meta.fields]
            return [f for f in all_fields if f != "notes"]
        return super().get_readonly_fields(request, obj)


    # Мастер может открывать и менять (только notes)
    def has_change_permission(self, request, obj=None):
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return True
        return super().has_change_permission(request, obj)

    # Мастеру — нельзя создавать профили
    def has_add_permission(self, request):
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return False
        return super().has_add_permission(request)

    # Мастеру — нельзя удалять профили
    def has_delete_permission(self, request, obj=None):
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return False
        return super().has_delete_permission(request, obj)


@admin.register(EmailVerification)
class EmailVerificationAdmin(admin.ModelAdmin):
    list_display = ("user", "purpose", "sent_to", "created_at", "expires_at", "last_sent_at", "attempts", "is_used")
    list_filter = ("purpose", "is_used", "created_at")
    search_fields = ("user__username", "user__email", "sent_to", "code")
    readonly_fields = ("user", "purpose", "code", "sent_to", "created_at", "expires_at", "attempts", "last_sent_at", "is_used")
    ordering = ("-created_at",)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

class MasterWorkDayInline(admin.TabularInline):
    model = MasterWorkDay
    extra = 7  # сразу 7 строк для всех дней




class MasterMonthlySalesTargetInline(admin.TabularInline):
    model = MasterMonthlySalesTarget
    extra = 0
    fields = ("month", "target_amount")
    ordering = ("-month",)


@admin.register(MasterMonthlySalesTarget)
class MasterMonthlySalesTargetAdmin(admin.ModelAdmin):
    list_display = ("master", "month", "target_amount", "updated_at")
    list_filter = ("month",)
    search_fields = ("master__user__user__first_name", "master__user__user__last_name")
    autocomplete_fields = ("master",)
    ordering = ("-month", "master__user__user__first_name")


@admin.register(MasterProfile)
class MasterProfileAdmin(ExportCsvMixin,admin.ModelAdmin):
    inlines = [MasterWorkDayInline, MasterMonthlySalesTargetInline]
    add_form = MasterCreateFullForm
    change_list_template = "admin/masters/changelist_cards.html"
    list_per_page = 24

    readonly_fields = ['password_display']
    export_fields = [
        "first_name",
        "last_name",
        "email",
        "username",
        "phone",
        "birth_date",
        "postal_code",
        "profession",
        "bio",
        "work_start",
        "work_end",
        "is_staff",
        "is_superuser",
        "is_active",
    ]
    search_fields = (
        "user__user__username",
        "user__user__first_name",
        "user__user__last_name",
        "user__user__email",
        "user__phone",
    )

    def get_queryset(self, request):
        # Prefetch related auth user to minimize queries in cards template.
        qs = super().get_queryset(request).select_related("user__user")

        name_order = getattr(request, "_master_order_choice", request.GET.get("name_order"))
        if name_order not in {"az", "za"}:
            cached_order = getattr(request, "_master_order_choice", None)
            if cached_order in {"az", "za"}:
                name_order = cached_order
            else:
                name_order = "az"
        request._master_order_choice = name_order

        if name_order == "za":
            qs = qs.order_by("-user__user__first_name", "-user__user__last_name", "-pk")
        else:
            qs = qs.order_by("user__user__first_name", "user__user__last_name", "pk")

        return qs

    def get_ordering(self, request):
        name_order = getattr(request, "_master_order_choice", None) or request.GET.get("name_order")
        if name_order == "za":
            return ("-user__user__first_name", "-user__user__last_name", "-pk")
        return ("user__user__first_name", "user__user__last_name", "pk")

    def lookup_allowed(self, lookup, value):
        if lookup in {"name_order"}:
            return True
        return super().lookup_allowed(lookup, value)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}

        professions = (
            self.model.objects.exclude(profession__isnull=True)
            .exclude(profession__exact="")
            .values_list("profession", flat=True)
            .distinct()
            .order_by("profession")
        )
        profession_options = [{"value": "", "label": _("All professions")}]
        profession_options.extend({"value": prof, "label": prof} for prof in professions)
        profession_current = request.GET.get("profession", "")
        profession_current_label = profession_current if profession_current else ""

        name_order = request.GET.get("name_order") or "az"
        if name_order not in {"az", "za"}:
            name_order = "az"

        request._master_profession_filter = profession_current
        request._master_order_choice = name_order

        extra_context.update(
            {
                "profession_options": profession_options,
                "profession_current": profession_current,
                "profession_current_label": profession_current_label,
                "name_order_current": name_order,
            }
        )

        try:
            opts = self.model._meta
            export_url = reverse(f'admin:{opts.app_label}_{opts.model_name}_export_csv')
            export_url += f"?{request.GET.urlencode()}"
        except NoReverseMatch:
            export_url = None
        extra_context["export_url"] = export_url

        original_get = request.GET
        cleaned_get = original_get.copy()
        if profession_current:
            cleaned_get.pop("profession", None)
            cleaned_get["profession__iexact"] = profession_current
        else:
            cleaned_get.pop("profession", None)
        cleaned_get.pop("name_order", None)
        request.GET = cleaned_get
        try:
            return super(ExportCsvMixin, self).changelist_view(request, extra_context=extra_context)
        finally:
            request.GET = original_get

    def get_export_row(self, obj):
        user_profile = getattr(obj, "user", None)
        auth_user = None

        if user_profile and hasattr(user_profile, "user"):
            auth_user = user_profile.user
        else:
            auth_user = user_profile
            if auth_user and hasattr(auth_user, "userprofile"):
                user_profile = auth_user.userprofile

        if not auth_user:
            auth_user = getattr(obj, "user", None)

        phone = getattr(user_profile, "phone", "")
        birth_date = getattr(user_profile, "birth_date", "")
        postal_code = getattr(user_profile, "postal_code", "")

        return [
            getattr(auth_user, "first_name", ""),
            getattr(auth_user, "last_name", ""),
            getattr(auth_user, "email", ""),
            getattr(auth_user, "username", ""),
            phone,
            birth_date,
            postal_code,
            getattr(obj, "profession", ""),
            getattr(obj, "bio", ""),
            getattr(obj, "work_start", ""),
            getattr(obj, "work_end", ""),
            getattr(auth_user, "is_staff", ""),
            getattr(auth_user, "is_superuser", ""),
            getattr(auth_user, "is_active", ""),
        ]
    form = MasterCreateFullForm  # на редактирование тоже можно оставить ту же

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related("user__user")

    list_display = ("get_name", "profession")

    def get_fieldsets(self, request, obj=None):
        form = self.form(instance=obj if obj else None)
        fields = list(form.fields.keys())

        if obj:
            # редактирование
            fields = [f for f in fields if f not in ['password1', 'password2', 'password']]  # ← обязательно убрать 'password'
            if 'email' in fields and 'password_display' not in fields:
                fields.insert(fields.index('email') + 1, 'password_display')
            elif 'password_display' not in fields:
                fields.append('password_display')
        else:
            # создание
            fields = [f for f in fields if f != 'password_display']

        return [(None, {'fields': fields})]

    def password_display(self, obj):
        from django.utils.html import format_html

        if not obj or not getattr(obj, "user", None) or not getattr(obj.user, "id", None):
            return _("Save the master to manage password.")

        try:
            url = reverse("admin:auth_user_password_change", args=[obj.user.id])
        except NoReverseMatch:
            return _("Change password")

        return format_html(
            '<a href="{}" class="button" style="color: #fff; background: #007bff; padding: 4px 8px; text-decoration: none; border-radius: 4px;">{}</a>',
            url,
            _("Change password"),
        )
    password_display.short_description = "Password"

    def has_change_permission(self, request, obj=None):
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if hasattr(request.user, "master_profile") and not request.user.is_superuser:
            return False
        return super().has_delete_permission(request, obj)

    def get_name(self, obj):
        return obj.user.get_full_name() or obj.user.username



    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        obj = self.get_object(request, object_id) if object_id else None
        if request.method == "POST":
            post = request.POST
            if hasattr(post, "_mutable"):
                mutable = post._mutable
                post._mutable = True
                for inline in self.get_inline_instances(request, obj):
                    formset_class = inline.get_formset(request, obj)
                    prefix = formset_class.get_default_prefix()
                    total_key = f"{prefix}-TOTAL_FORMS"
                    if total_key not in post:
                        post[total_key] = "0"
                        post[f"{prefix}-INITIAL_FORMS"] = "0"
                        post.setdefault(f"{prefix}-MIN_NUM_FORMS", "0")
                        max_num = getattr(formset_class, "max_num", None)
                        if max_num in (None, 0):
                            max_num = getattr(formset_class, "DEFAULT_MAX_NUM", "1000")
                        post.setdefault(f"{prefix}-MAX_NUM_FORMS", str(max_num))
                post._mutable = mutable
        return super().changeform_view(request, object_id, form_url, extra_context)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

        # 1. Забираем все текущие разрешения
        user = obj.user.user
        user.user_permissions.clear()

        # 2. Добавляем только view_appointment
        needed = [
            # Appointment
            "view_appointment", "add_appointment", "change_appointment", "delete_appointment",
            # MasterAvailability (time off)
            "view_masteravailability", "add_masteravailability", "change_masteravailability", "delete_masteravailability",

             "view_userprofile", "change_userprofile", "add_appointmentitem", "change_appointmentitem", "delete_appointmentitem"
        ]
        perms = Permission.objects.filter(codename__in=needed)
        user.user_permissions.add(*perms)

        ct_proxy = ContentType.objects.get_for_model(CustomUserDisplay)
        try:
            p_view_proxy = Permission.objects.get(content_type=ct_proxy, codename="view_customuserdisplay")
            user.user_permissions.add(p_view_proxy)
        except Permission.DoesNotExist:
            pass  # миграции ещё не применены — не падаем

        # (не обязательно, но можно) право на просмотр профилей клиентов
        ct_profile = ContentType.objects.get_for_model(UserProfile)
        try:
            p_view_profile = Permission.objects.get(content_type=ct_profile, codename="view_userprofile")
            user.user_permissions.add(p_view_profile)
        except Permission.DoesNotExist:
            pass

        user.save()


TZ = get_current_timezone()


# ──────────────────────────────────────────────────────────────────────────────
# helpers
# ──────────────────────────────────────────────────────────────────────────────

def _parse_period(request: HttpRequest) -> tuple[date, date]:
    """?start=YYYY-MM-DD&end=YYYY-MM-DD, по умолчанию последние 30 дней (включительно)."""
    today = date.today()
    start_s = request.GET.get("start") or ""
    end_s = request.GET.get("end") or ""
    try:
        start = date.fromisoformat(start_s) if start_s else today - timedelta(days=30)
    except Exception:
        start = today - timedelta(days=30)
    try:
        end = date.fromisoformat(end_s) if end_s else today
    except Exception:
        end = today
    if end < start:
        start, end = end, start
    return start, end


def _cancel_no_show_names() -> list[str]:
    """Фактические названия статусов в БД (fallback на строковые)."""
    cancelled = (AppointmentStatus.objects
                 .filter(name__iexact="Cancelled")
                 .values_list("name", flat=True).first()) or "Cancelled"
    no_show = (AppointmentStatus.objects
               .filter(name__iexact="No_Show")
               .values_list("name", flat=True).first()) or "No_Show"
    return [cancelled, no_show]

def _client_source_aggregation(qs):
    # пробуем client__source, если его нет — client__how_heard
    try:
        rows = (
            qs.values("client__source")
            .annotate(
                revenue=Coalesce(Sum("final_price"), Value(Decimal("0"))),
                appts=Count("id"),
                clients=Count("client_id", distinct=True),
            )
            .order_by("-revenue")
        )
        # нормализуем ключ + подпись
        return [{"label": (r["client__source"] or "unknown"), **{k: r[k] for k in ("revenue","appts","clients")}} for r in rows]
    except FieldError:
        rows = (
            qs.values("client__how_heard")
            .annotate(
                revenue=Coalesce(Sum("final_price"), Value(Decimal("0"))),
                appts=Count("id"),
                clients=Count("client_id", distinct=True),
            )
            .order_by("-revenue")
        )
        return [{"label": (r["client__how_heard"] or "unknown"), **{k: r[k] for k in ("revenue","appts","clients")}} for r in rows]

# ──────────────────────────────────────────────────────────────────────────────
# view
# ──────────────────────────────────────────────────────────────────────────────


def _collect_model_admin_exports(request: HttpRequest) -> List[ExportDashboardEntry]:
    entries: List[ExportDashboardEntry] = []

    for model, model_admin in admin.site._registry.items():
        try:
            if hasattr(model_admin, "has_view_permission") and not model_admin.has_view_permission(request):
                continue
        except Exception:
            continue

        provider = getattr(model_admin, "get_export_dashboard_entries", None)
        if callable(provider):
            custom = provider(request)
            if custom:
                entries.extend(custom)
            continue

        base_group = getattr(model_admin, "export_dashboard_group", None)
        if not base_group:
            base_group = model._meta.app_label.replace("_", " ").title() or "Data"

        base_label = getattr(model_admin, "export_dashboard_label", None)
        if not base_label:
            base_label = str(model._meta.verbose_name_plural).title()

        description = getattr(model_admin, "export_dashboard_description", None)
        supports_range = bool(getattr(model_admin, "export_dashboard_supports_range", False))
        requires_range = bool(getattr(model_admin, "export_dashboard_requires_range", False))
        params = getattr(model_admin, "export_dashboard_params", None) or {}

        app_label = model._meta.app_label
        model_name = model._meta.model_name

        if isinstance(model_admin, ExportCsvMixin):
            try:
                url = reverse(f"admin:{app_label}_{model_name}_export_csv")
            except NoReverseMatch:
                url = None
            if url:
                entries.append(
                    ExportDashboardEntry(
                        key=f"{app_label}.{model_name}.quick_xlsx",
                        label=f"{base_label} — Quick XLSX",
                        url=url,
                        description=description,
                        group=base_group,
                        params=dict(params),
                        supports_range=supports_range,
                        requires_range=requires_range,
                    )
                )

        if isinstance(model_admin, ExportXlsxMixin):
            try:
                url = reverse(f"admin:{app_label}_{model_name}_export_xlsx")
            except NoReverseMatch:
                url = None
            if url:
                entries.append(
                    ExportDashboardEntry(
                        key=f"{app_label}.{model_name}.full_xlsx",
                        label=f"{base_label} — Full XLSX",
                        url=url,
                        description=description,
                        group=base_group,
                        params=dict(params),
                        supports_range=supports_range,
                        requires_range=requires_range,
                    )
                )

    return entries


def _manual_export_entries(request: HttpRequest) -> List[ExportDashboardEntry]:
    entries: List[ExportDashboardEntry] = []
    try:
        stats_url = reverse("admin:stats")
    except NoReverseMatch:
        stats_url = None

    if stats_url:
        description = "Period-based performance metrics with breakdowns by clients and discounts."
        entries.append(
            ExportDashboardEntry(
                key="stats.multi_sheet",
                label="Statistics — Multi-sheet workbook",
                url=stats_url,
                description=description,
                group="Analytics",
                params={"export": "xlsx"},
                supports_range=True,
                requires_range=True,
            )
        )
        entries.append(
            ExportDashboardEntry(
                key="stats.combined_sheet",
                label="Statistics — Combined sheet",
                url=stats_url,
                description="All statistics sheets flattened into a single tab.",
                group="Analytics",
                params={"export": "xlsx_flat"},
                supports_range=True,
                requires_range=True,
            )
        )

    return entries


def _collect_export_dashboard_entries(request: HttpRequest) -> List[ExportDashboardEntry]:
    entries = _collect_model_admin_exports(request)
    entries.extend(_manual_export_entries(request))
    return entries


@staff_member_required
def exports_dashboard_view(request: HttpRequest) -> HttpResponse:
    entries = _collect_export_dashboard_entries(request)
    grouped: Dict[str, List[ExportDashboardEntry]] = defaultdict(list)
    for entry in entries:
        grouped[entry.group].append(entry)

    grouped_exports = [
        (group, sorted(items, key=lambda e: e.label.lower()))
        for group, items in sorted(grouped.items(), key=lambda item: item[0].lower())
    ]

    start, end = _parse_period(request)
    context = {
        "title": "Exports",
        "grouped_exports": grouped_exports,
        "default_start": start.isoformat(),
        "default_end": end.isoformat(),
    }
    return render(request, "admin/exports_dashboard.html", context)


@staff_member_required
def stats_view(request: HttpRequest) -> HttpResponse:
    # ── Период
    start, end = _parse_period(request)  # helper уже есть в файле

    # ── Отбрасываем отменённые/похожие статусы
    last_status_name = (
        AppointmentStatusHistory.objects
        .filter(appointment_id=OuterRef("pk"))
        .order_by("-set_at")
        .values("status__name")[:1]
    )
    cancelled_like = _cancel_no_show_names()  # helper уже есть

    # Валидные аппы по дате начала визита (для счётчиков штук)
    base_appt_q = (
        Appointment.objects
        .select_related("client__user")
        .annotate(last_status=Subquery(last_status_name))
        .exclude(last_status__in=cancelled_like)
        .filter(start_time__date__gte=start, start_time__date__lte=end)
    )

    # Реальные платежи в периоде (по дате платежа)
    payments_q = (
        Payment.objects
        .filter(
            appointment__in=base_appt_q,
            created_at__date__gte=start,
            created_at__date__lte=end,
        )
    )

    # ── KPI
    kpi_total_revenue = payments_q.aggregate(
        x=Coalesce(Sum("amount"), Value(Decimal("0.00")))
    )["x"]
    kpi_total_appointments = base_appt_q.count()
    kpi_with_discount = base_appt_q.exclude(discount_source="").count()

    # ── Топы (spent — по платежам периода)
    pay_sub = (
        payments_q
        .filter(appointment__client_id=OuterRef("client_id"))
        .values("appointment__client_id")
        .annotate(s=Coalesce(Sum("amount"), Value(Decimal("0.00"))))
        .values("s")[:1]
    )

    top_bookers = (
        base_appt_q.values(
            "client_id",
            "client__user__first_name",
            "client__user__last_name",
            "client__user__email",
        )
        .annotate(
            appt_count=Count("id"),
            spent=Coalesce(Subquery(pay_sub), Value(Decimal("0.00"))),
        )
        .order_by("-appt_count", "-spent")[:10]
    )

    top_spenders = (
        base_appt_q.values(
            "client_id",
            "client__user__first_name",
            "client__user__last_name",
            "client__user__email",
        )
        .annotate(
            spent=Coalesce(Subquery(pay_sub), Value(Decimal("0.00"))),
            appt_count=Count("id"),
        )
        .order_by("-spent", "-appt_count")[:10]
    )

    # ── Разбивка по discount_source (по платежам)
    ds_totals = (
        payments_q
        .values(discount_source=F("appointment__discount_source"))
        .annotate(
            cnt=Count("appointment", distinct=True),
            revenue=Coalesce(Sum("amount"), Value(Decimal("0.00"))),
        )
        .order_by("-revenue")
    )

    # ── Дневные суммы по источникам (stacked bar) — по платежам
    days = [start + timedelta(days=i) for i in range((end - start).days + 1)]
    daily = {d.isoformat(): {"general": 0.0, "personal": 0.0, "promocode": 0.0, "none": 0.0} for d in days}

    def _bucket(src: str) -> str:
        s = (src or "").strip().lower()
        if "promo" in s:    return "promocode"
        if "personal" in s: return "personal"
        if "general" in s:  return "general"
        if not s:           return "none"
        if "+" in s:  # комбинированные значения
            parts = s.split("+")
            if any("promo" in p for p in parts):    return "promocode"
            if any("personal" in p for p in parts): return "personal"
            if any("general" in p for p in parts):  return "general"
        return "none"

    per_day = (
        payments_q
        .annotate(day=F("created_at__date"))
        .values("day", "appointment__discount_source")
        .annotate(total=Coalesce(Sum("amount"), Value(Decimal("0.00"))))
    )
    for row in per_day:
        d = row["day"].isoformat() if hasattr(row["day"], "isoformat") else str(row["day"])
        b = _bucket(row.get("appointment__discount_source"))
        if d in daily:
            daily[d][b] += float(row["total"])

    chart = {
        "labels": list(daily.keys()),
        "general":  [daily[d]["general"]   for d in daily],
        "personal": [daily[d]["personal"]  for d in daily],
        "promocode":[daily[d]["promocode"] for d in daily],
        "none":     [daily[d]["none"]      for d in daily],
    }

    # ── Разбивка по ServiceDiscount (uses/saved по айтемам; revenue — как было от цен, по платежам можно сделать отдельно)
    # Использования: позиции в валидных аппоинтментах с активной скидкой услуги на дату визита
    discount_filters = (
            Q(service__appointmentitem__appointment__in=base_appt_q) &
            Q(service__appointmentitem__appointment__start_time__date__gte=F("start_date")) &
            Q(service__appointmentitem__appointment__start_time__date__lte=F("end_date")) &
            Q(service__appointmentitem__discount_source="service")
    )
    discount_breakdown = (
        ServiceDiscount.objects
        .select_related("service")
        .annotate(
            uses=Count("service__appointmentitem", filter=discount_filters),
            revenue=Coalesce(  # по позициям (как раньше)
                Sum("service__appointmentitem__final_price", filter=discount_filters),
                Value(Decimal("0"))
            ),
            saved=Coalesce(
                Sum(
                    (F("service__appointmentitem__unit_price") - F("service__appointmentitem__final_price")),
                    filter=discount_filters
                ),
                Value(Decimal("0"))
            ),
        )
        .filter(uses__gt=0)
        .order_by("-uses", "-revenue")
    )

    # ── Разбивка по промокодам: uses — по позициям; revenue — платежи визита,
    #     распределённые пропорционально сумме final_price позиций с этим кодом в визите.
    # Платежи по визиту
    pay_by_appt = dict(
        payments_q.values("appointment_id")
        .annotate(total=Coalesce(Sum("amount"), Value(Decimal("0"))))
        .values_list("appointment_id", "total")
    )

    # Сумма final по промо-позициям в визите и веса по конкретным кодам
    promo_rows = (
        AppointmentItemPromoCode.objects
        .select_related("item__appointment", "promocode")
        .filter(item__appointment__in=base_appt_q)
        .values("promocode__code", "item__appointment_id")
        .annotate(weight=Coalesce(Sum("item__final_price"), Value(Decimal("0"))),
                  uses=Count("id"))
    )

    weights_per_appt = defaultdict(Decimal)  # сумма весов по визиту
    rows_per_code_appt = defaultdict(list)   # [(code, appt_id, weight, uses)]
    for r in promo_rows:
        code = r["promocode__code"]
        aid  = r["item__appointment_id"]
        w    = r["weight"] or Decimal("0")
        u    = r["uses"] or 0
        weights_per_appt[aid] += w
        rows_per_code_appt[(code, aid)].append((w, u))

    promo_totals = defaultdict(lambda: {"uses": 0, "revenue": Decimal("0")})
    for (code, aid), lst in rows_per_code_appt.items():
        paid = pay_by_appt.get(aid, Decimal("0"))
        total_w = weights_per_appt.get(aid, Decimal("0")) or Decimal("0")
        # если нет весов — делим поровну между кодами визита
        denom = total_w if total_w > 0 else Decimal(len({c for (c, a) in rows_per_code_appt if a == aid}) or 1)
        share = paid * (sum(w for w, _ in lst) / denom) if total_w > 0 else paid / denom
        promo_totals[code]["uses"] += sum(u for _, u in lst)
        promo_totals[code]["revenue"] += share

    promo_breakdown = [
        {"code": code, "uses": data["uses"], "revenue": data["revenue"]}
        for code, data in promo_totals.items()
    ]
    promo_breakdown.sort(key=lambda x: (-x["uses"], -x["revenue"]))

    # ── Клиентские источники (stacked/таблица/бар) — всё по платежам
    src_attr, how_attr = "source", "how_heard"
    try:
        src_field = UserProfile._meta.get_field(src_attr)
    except Exception:
        src_attr = "how_heard"
        src_field = UserProfile._meta.get_field(src_attr)
    src_choices = list(getattr(src_field, "choices", [])) or []
    try:
        how_field = UserProfile._meta.get_field(how_attr)
        how_choices = list(getattr(how_field, "choices", [])) or []
    except Exception:
        how_choices = []
    if ("unknown", "Unknown") not in src_choices:
        src_choices.append(("unknown", "Unknown"))
    if how_choices and ("unknown", "Unknown") not in how_choices:
        how_choices.append(("unknown", "Unknown"))

    def _norm(v): return ("" if v is None else str(v)).strip()
    def _is_online(v): return _norm(v).lower() == "online"

    expanded_categories = []
    for s_val, s_label in src_choices:
        if _is_online(s_val) and how_choices:
            for h_val, h_label in how_choices:
                expanded_categories.append((f"online::{h_val}", f"{s_label} — {h_label}"))
        else:
            expanded_categories.append((s_val, s_label))

    labels = [d.isoformat() for d in days]
    label_index = {s: i for i, s in enumerate(labels)}
    series = {key: [0.0] * len(labels) for key, _ in expanded_categories}

    src_path = f"appointment__client__{src_attr}"
    how_path = f"appointment__client__{how_attr}"

    rows = (
        payments_q
        .annotate(day=F("created_at__date"))
        .values("day", src_path, how_path)
        .annotate(total=Coalesce(Sum("amount"), Value(Decimal("0"))))
    )

    def _key_for_row(src_val, how_val) -> str:
        s = _norm(src_val) or "unknown"
        if _is_online(s) and how_choices:
            h = _norm(how_val) or "unknown"
            return f"online::{h}"
        return s or "unknown"

    for r in rows:
        d = r["day"].isoformat() if hasattr(r["day"], "isoformat") else str(r["day"])
        j = label_index.get(d)
        if j is None:
            continue
        key = _key_for_row(r.get(src_path), r.get(how_path))
        if key not in series:
            # если всплыла новая how_heard — добавим
            label = key
            if key.startswith("online::"):
                h_val = key.split("::", 1)[1]
                h_label = next((lbl for val, lbl in how_choices if val == h_val), h_val.title().replace("_", " "))
                s_label = next((lbl for val, lbl in src_choices if _is_online(val)), "Online")
                label = f"{s_label} — {h_label}"
            expanded_categories.append((key, label))
            series[key] = [0.0] * len(labels)
        series[key][j] += float(r["total"])

    # Таблица totals по категориям (revenue по платежам, appts/clients — по платежам этого периода)
    totals = {key: {"revenue": 0.0, "appts": set(), "clients": set()} for key, _ in expanded_categories}
    rows_total = (
        payments_q
        .values(src_path, how_path, "appointment_id", "appointment__client_id")
        .annotate(revenue=Coalesce(Sum("amount"), Value(Decimal("0"))))
    )
    for r in rows_total:
        key = _key_for_row(r.get(src_path), r.get(how_path))
        t = totals.setdefault(key, {"revenue": 0.0, "appts": set(), "clients": set()})
        t["revenue"] += float(r["revenue"])
        if r.get("appointment_id"):
            t["appts"].add(r["appointment_id"])
        if r.get("appointment__client_id"):
            t["clients"].add(r["appointment__client_id"])

    client_source_table = [
        {
            "key": key,
            "label": label,
            "revenue": round(totals[key]["revenue"], 2),
            "appts": len(totals[key]["appts"]),
            "clients": len(totals[key]["clients"]),
        }
        for key, label in expanded_categories
    ]

    client_source_stacked = {
        "labels": labels,
        "datasets": [{"key": key, "label": label, "data": series[key]} for key, label in expanded_categories],
    }

    # Бар «total revenue by source» — по платежам
    src_bar_rows = (
        payments_q
        .values(src_path)
        .annotate(revenue=Coalesce(Sum("amount"), Value(Decimal("0"))))
        .order_by("-revenue")
    )
    client_source_stats = [{"label": (_norm(r.get(src_path)) or "unknown"), "revenue": r["revenue"]} for r in src_bar_rows]
    src_labels = [r["label"] for r in client_source_stats]
    src_revenue = [float(r["revenue"]) for r in client_source_stats]


    export_kind = request.GET.get("export")
    if export_kind in {"xlsx", "xlsx_flat"}:
        exporter = ExportXlsxMixin()  # просто используем миксин как helper
        builder = exporter.build_statistics_workbook
        filename = "statistics.xlsx"
        if export_kind == "xlsx_flat":
            builder = exporter.build_statistics_flat_workbook
            filename = "statistics_flat.xlsx"

        wb = builder(
            start=start, end=end,
            kpi_total_revenue=kpi_total_revenue,
            kpi_total_appointments=kpi_total_appointments,
            kpi_with_discount=kpi_with_discount,
            top_bookers=list(top_bookers),
            top_spenders=list(top_spenders),
            ds_totals=list(ds_totals),
            discount_breakdown=list(discount_breakdown),
            promo_breakdown=list(promo_breakdown),
            client_source_table=list(client_source_table),
        )
        return exporter._wb_response(wb, filename)

    # === иначе — рендерим страницу и даём ссылку на экспорт с текущими фильтрами ===
    params = request.GET.copy()
    params["export"] = "xlsx"
    export_xlsx_url = f"{request.path}?{params.urlencode()}"  # сохраняет все фильтры + export=xlsx

    params_flat = request.GET.copy()
    params_flat["export"] = "xlsx_flat"
    export_flat_xlsx_url = f"{request.path}?{params_flat.urlencode()}"
    # ── Контекст для шаблона
    context = {
        "title": "Statistics",
        "start": start,
        "end": end,
        "kpi": {
            "revenue": kpi_total_revenue,
            "appointments": kpi_total_appointments,
            "with_discount": kpi_with_discount,
        },
        "top_bookers": list(top_bookers),
        "top_spenders": list(top_spenders),
        "ds_totals": list(ds_totals),
        "discount_breakdown": list(discount_breakdown),
        "promo_breakdown": list(promo_breakdown),
        "chart": chart,

        "client_source_stats": client_source_stats,
        "client_source_chart": {"labels": src_labels, "revenue": src_revenue},

        "client_source_stacked": client_source_stacked,
        "client_source_table": client_source_table,
        "export_xlsx_url": export_xlsx_url,
        "export_flat_xlsx_url": export_flat_xlsx_url,
    }
    return render(request, "admin/statistics.html", context)

def _inject_admin_urls(original_get_urls):
    def get_urls():
        my = [
            path("exports/", admin.site.admin_view(exports_dashboard_view), name="exports"),
            path("stats/", admin.site.admin_view(stats_view), name="stats"),
        ]
        return my + original_get_urls()
    return get_urls

admin.site.get_urls = _inject_admin_urls(admin.site.get_urls)

def get_price_html(service):
    discount = service.get_active_discount()
    if discount:
        discounted = service.get_discounted_price()
        return format_html(
            '<span style="text-decoration: line-through; color: grey;">${}</span><br><strong>${}</strong>',
            service.base_price,
            discounted
        )
    return format_html("<strong>${}</strong>", service.base_price)



def _compute_paid_total(appointment_obj):
    paid_total_value = Decimal("0.00")
    prefetched = getattr(appointment_obj, "prefetched_succeeded_payments", None)
    if prefetched is not None:
        payments_iterable = prefetched
    else:
        payments_rel = getattr(appointment_obj, "payments", None)
        if payments_rel is None:
            return paid_total_value.quantize(TWOPLACES)
        try:
            payments_iterable = payments_rel.all()
        except AttributeError:
            payments_iterable = payments_rel or []
    for payment in payments_iterable or []:
        if getattr(payment, "status", "") != "succeeded":
            continue
        amount_received = getattr(payment, "amount_received", Decimal("0.00")) or Decimal("0.00")
        amount_refunded = getattr(payment, "amount_refunded", Decimal("0.00")) or Decimal("0.00")
        paid_total_value += Decimal(amount_received) - Decimal(amount_refunded)
    return paid_total_value.quantize(TWOPLACES)


def createTable(selected_date, time_pointer, end_time, slot_times, items, masters, availabilities):
    """
    items: QuerySet[AppointmentItem] с select_related('appointment__client','service','master')
    masters: список мастеров для колонок
    """
    COLOR_PALETTE = ["#E4D08A", "#EDC2A2", "#CEAEC6", "#A3C1C9", "#C3CEA3", "#E7B3C3"]
    master_ids = [m.id for m in masters]
    MASTER_COLORS = dict(zip(master_ids, cycle(COLOR_PALETTE)))

    # ───── badges для позиции ───────────────────────────────────────────────────
    def _corner_badges_for_item(item, meta):
        # скидка: если установлен промокод/персональная скидка на уровне позиции
        promo_html = ""
        base = item.unit_price if getattr(item, "unit_price", None) is not None else item.service.base_price
        final = getattr(item, "final_price", None)
        has_promo_rel = getattr(item, "promocode_link", None) is not None
        if has_promo_rel or (final is not None and str(final) != str(base)):
            promo_html = "<span class='badge badge--promo' title='Applied discount'>%</span>"

        # здоровье (по клиенту Appointment)
        def _health_flag_info(appt):
            prof = getattr(appt, "client", None)
            if not prof:
                return False, "", ""
            hc = getattr(prof, "health", None) or getattr(prof, "health_conditions", None) or {}
            def _to_str(v):
                if isinstance(v, (list, tuple)): return ", ".join(map(str, v))
                return (v or "").strip()
            has_all = bool(hc.get("has_allergies")) or bool(_to_str(hc.get("allergies_text")))
            has_med = bool(_to_str(hc.get("medications")))
            has_ctr = bool(_to_str(hc.get("contraindications")))
            has_ch = bool(_to_str(hc.get("chronic_conditions")))
            if not (has_all or has_med or has_ctr or has_ch):
                return False, "", ""
            try:
                url = reverse("health-view-master", args=[prof.id])
            except NoReverseMatch:
                url = ""
            return True, url, "Есть важные данные в анкете здоровья — нажмите, чтобы посмотреть"

        health_html = ""
        show_flag, flag_url, flag_title = _health_flag_info(item.appointment)

        payment_html = ""
        paid_total = meta.get("paid_total_decimal")
        grand_total = meta.get("grand_total_decimal")
        eps = Decimal("0.01")
        if paid_total is not None and grand_total is not None and (grand_total >= eps or paid_total >= eps):
            if paid_total >= grand_total - eps:
                payment_html = (
                    f"<span class='badge badge--paid'>"
                    f"<img src=\"{PAID_BADGE_ICON_URL}\" alt=\"Paid\" class=\"badge-icon\" height=\"24\" />"
                    f"</span>"
                )
            elif paid_total >= eps and paid_total < grand_total - eps:
                payment_html = (
                    f"<span class='badge badge--partial'>"
                    f"<img src=\"{PARTIAL_BADGE_ICON_URL}\" alt=\"Partially paid\" class=\"badge-icon\" height=\"24\"/>"
                    f"</span>"
                )

        if show_flag:
            ico = "⚕️"
            health_html = (
                f'<a class="badge badge--health" href="{flag_url}" title="{flag_title}">{ico}</a>'
                if flag_url else f'<span class="badge badge--health" title="{flag_title}">{ico}</span>'
            )
        note_html = ""
        if meta.get("has_note"):
            note_html = (
                f"<span class='badge badge--note'>"
                f"<img src=\"{NOTES_BADGE_ICON_URL}\" alt=\"Notes present\" class=\"badge-icon\" height=\"24\"/>"
                f"</span>")
        badges_html = "".join(filter(None, [promo_html, payment_html, health_html, note_html]))
        if not badges_html:
            return ""
        return f"<div class='corner-badges'>{badges_html}</div>"

    # ───── вспомогательные ──────────────────────────────────────────────────────
    def _item_meta(item, master_obj):
        s_local = localtime(item.start_time)
        # duration берём из Item (в нём уже может быть extra_time учтён)【:contentReference[oaicite:3]{index=3}】
        total_min = int(getattr(item, "duration_min", 0) or 0)
        e_local = s_local + timedelta(minutes=total_min)

        # Статус — по родительскому Appointment (последний из истории)
        last_status = item.appointment.appointmentstatushistory_set.order_by("-set_at").first()
        status_name = last_status.status.name if last_status else "Unknown"
        appointment_obj = item.appointment
        items = list(appointment_obj.items.all())
        items_count = len(items)
        base_price = appointment_obj.total_without_discounts(ignore_overrides=False)
        base_price_decimal = Decimal(base_price or Decimal("0.00")).quantize(TWOPLACES)

        service_discounted_subtotal = Decimal("0.00")
        service_tax_total = Decimal("0.00")
        for appt_item in items:
            final_val = getattr(appt_item, "final_price", None)
            if final_val is None:
                if appt_item.unit_price is not None:
                    final_val = appt_item.unit_price
                else:
                    final_val = getattr(appt_item.service, "base_price", Decimal("0.00"))
            service_discounted_subtotal += Decimal(final_val or Decimal("0.00"))
            service_tax_total += Decimal(getattr(appt_item, "tax_amount", Decimal("0.00")) or Decimal("0.00"))
        service_discounted_subtotal = service_discounted_subtotal.quantize(TWOPLACES)
        service_tax_total = service_tax_total.quantize(TWOPLACES)
        service_total_with_tax = (service_discounted_subtotal + service_tax_total).quantize(TWOPLACES)

        products_total_with_tax = Decimal("0.00")
        product_sales_rel = getattr(appointment_obj, "product_sales", None)
        if product_sales_rel is not None:
            for sale in product_sales_rel.all():
                subtotal = Decimal(getattr(sale, "total_amount", Decimal("0.00")) or Decimal("0.00"))
                tax_amount = Decimal(getattr(sale, "tax_amount", Decimal("0.00")) or Decimal("0.00"))
                products_total_with_tax += subtotal + tax_amount
        products_total_with_tax = products_total_with_tax.quantize(TWOPLACES)

        final_price = appointment_obj.final_price
        grand_total = Decimal(final_price or Decimal("0.00"))
        if final_price is None:
            grand_total = service_total_with_tax + products_total_with_tax
        grand_total = grand_total.quantize(TWOPLACES)

        paid_total_cached = getattr(appointment_obj, "_cached_paid_total", None)
        if paid_total_cached is None:
            paid_total_cached = _compute_paid_total(appointment_obj)
            setattr(appointment_obj, "_cached_paid_total", paid_total_cached)

        has_discount = bool(getattr(appointment_obj, "discount_source", ""))
        if not has_discount and base_price_decimal > Decimal("0.00"):
            if (base_price_decimal - service_discounted_subtotal) >= Decimal("0.01"):
                has_discount = True

        client = appointment_obj.client
        client_label = client.get_full_name() or client.user.username
        notes_value = getattr(appointment_obj, "notes", "") or ""
        has_note = bool(str(notes_value).strip())

        return {
            "s_local": s_local,
            "e_local": e_local,
            "status": status_name,
            "master_label": escape(str(master_obj)),
            "client_label": escape(client_label),
            "service_label": escape(item.service.name),
            "time_label": f"{s_local.strftime('%I:%M%p').lstrip('0')} - {e_local.strftime('%I:%M%p').lstrip('0')}",
            "duration_label": f"{total_min}min",
            "base_price": f"${base_price_decimal}",
            "base_price_raw": f"{base_price_decimal:.2f}",
            "items_count": items_count,
            "final_price": f"${grand_total}",
            "final_price_raw": f"{grand_total:.2f}",
            "service_total_raw": f"{service_total_with_tax:.2f}",
            "products_total_raw": f"{products_total_with_tax:.2f}",
            "has_discount": has_discount,
            "phone": escape(getattr(client, "phone", "") or ""),
            "grand_total_decimal": grand_total,
            "paid_total_decimal": paid_total_cached,
            "paid_total_raw": f"{paid_total_cached:.2f}",
            "has_note": has_note,
        }

    def _cell_html_item(item, meta, show_cancelled=False):
        cancelled_suffix = " (Cancelled)" if show_cancelled else ""
        opacity = ".7" if show_cancelled else "1"
        corner = _corner_badges_for_item(item, meta)
        footer = f"<div class='cell-mini-footer'>{meta['items_count']} services</div>" if meta.get('items_count', 1) > 1 else ""
        return f"""
        {corner}
        <div style="opacity:{opacity}; ">
          <div style="font-size:1.8vh;">
            {meta['s_local'].strftime('%I:%M').lstrip('0')} – {meta['e_local'].strftime('%I:%M').lstrip('0')}
            <strong>{meta['client_label']}</strong>
          </div>
          <div style="font-size:1.8vh;">
            {meta['service_label']}{cancelled_suffix}
          </div>
          {footer}
        </div>
        """

    def _make_item_cell(kind, item, rowspan, colspan, master_obj, bg, show_cancelled=False):
        meta = _item_meta(item, master_obj)
        return {
            "rowspan": rowspan,
            "colspan": colspan,
            "kind": kind,
            "appt_id": item.appointment_id,  # важно для ссылки клика по шаблону【:contentReference[oaicite:4]{index=4}】
            "html": _cell_html_item(meta=meta, item=item, show_cancelled=show_cancelled),
            "background": bg,
            "appointment": item,  # шаблон проверяет наличие cell.appointment для ветки рендера【:contentReference[oaicite:5]{index=5}】
            "client": meta["client_label"],
            "phone": meta["phone"],
            "service": meta["service_label"],
            "status": meta["status"],
            "master": meta["master_label"],
            "time_label": meta["time_label"],
            "duration": meta["duration_label"],
            "base_price": meta["base_price"],
            "base_price_raw": meta.get("base_price_raw", "0.00"),
            "final_price": meta["final_price"],
            "final_price_raw": meta.get("final_price_raw", "0.00"),
            "service_total_raw": meta.get("service_total_raw", "0.00"),
            "products_total_raw": meta.get("products_total_raw", "0.00"),
            "has_discount": meta.get("has_discount", False),
            "items_count": meta["items_count"],
            "paid_total_raw": meta.get("paid_total_raw", "0.00"),
            "has_note": meta.get("has_note", False),
        }

    def _make_unavail_cell(kind, rowsp, colspan, avail_id, reason, from_s, to_s, until_s):
        return {
            "rowspan": rowsp,
            "colspan": colspan,
            "kind": kind,
            "availability_id": avail_id,
            "html": f"""
                <div style="opacity:.85">
                  {escape(reason)}<br>
                  <small>{escape(from_s)}–{escape(to_s)}</small>
                </div>
            """,
            "unavailable": True,
            "reason": reason,
            "start": from_s,
            "end": to_s,
            "until": until_s,
        }

    # ───── сетка времени ────────────────────────────────────────────────────────
    while time_pointer <= end_time:
        slot_times.append(time_pointer.strftime("%H:%M"))
        time_pointer += timedelta(minutes=15)

    two_col_map = {}
    cancel_lanes = {}
    skip_two = {}
    skip_lane = {}
    for m in masters:
        mid = m.id
        two_col_map[mid] = {}
        cancel_lanes[mid] = {0: {}, 1: {}}
        skip_two[mid] = {}
        skip_lane[mid] = {0: {}, 1: {}}

    # статус Cancelled
    cancelled_status = AppointmentStatus.objects.filter(name="Cancelled").first()
    cancelled_id = getattr(cancelled_status, "id", None)

    # ───── позиции (AppointmentItem) ────────────────────────────────────────────
    for item in items:
        start_local = localtime(item.start_time)
        if start_local.date() != selected_date:
            continue

        mid = item.master_id
        slot_key = start_local.strftime("%H:%M")
        total_min = int(getattr(item, "duration_min", 0) or 0)
        rowspan = max(1, (-(-total_min // 15)))  # ceil

        last_status = item.appointment.appointmentstatushistory_set.order_by("-set_at").first()
        is_cancelled = bool(last_status and last_status.status_id == cancelled_id)

        if not is_cancelled:
            two_col_map[mid][slot_key] = {
                "kind": "appt_active",
                "rowspan": rowspan,
                "colspan": 2,
                "item": item,
            }
            for i in range(rowspan):
                t = (start_local + timedelta(minutes=15 * i)).strftime("%H:%M")
                skip_two[mid][t] = True
        else:
            lane0_busy = skip_lane[mid][0].get(slot_key) or (slot_key in cancel_lanes[mid][0])
            lane = 0 if not lane0_busy else 1
            cancel_lanes[mid][lane][slot_key] = {
                "kind": "appt_cancelled",
                "rowspan": rowspan,
                "colspan": 1,
                "item": item,
            }
            for i in range(rowspan):
                t = (start_local + timedelta(minutes=15 * i)).strftime("%H:%M")
                skip_lane[mid][lane][t] = True

    # ───── тайм-офф (перерывы/отпуска) ─────────────────────────────────────────
    for period in availabilities:
        mid = int(getattr(period.master, "id", period.master))
        start = localtime(period.start_time)
        end = localtime(period.end_time)
        if start.date() > selected_date or end.date() < selected_date:
            continue

        day_start = datetime.combine(selected_date, time(8, 0)).replace(tzinfo=start.tzinfo)
        day_end = datetime.combine(selected_date, time(21, 15)).replace(tzinfo=start.tzinfo)
        block_start = max(start, day_start)
        block_end = min(end, day_end)
        if block_start >= block_end:
            continue

        minutes = int((block_end - block_start).total_seconds() // 60)
        rowsp = max(1, (-(-minutes // 15)))  # ceil
        slot_key = block_start.strftime("%H:%M")

        if slot_key not in two_col_map.get(mid, set()):
            if mid not in two_col_map:
                two_col_map[mid] = {}
            two_col_map[mid][slot_key] = {
                "kind": "unavailable",
                "rowspan": rowsp,
                "colspan": 2,
                "reason": period.get_reason_display(),
                "from": block_start.strftime("%I:%M%p").lstrip("0"),
                "to": block_end.strftime("%I:%M%p").lstrip("0"),
                "until": period.end_time.strftime("%d %b %Y"),
                "availability_id": period.id,
            }
            for i in range(rowsp):
                t = (block_start + timedelta(minutes=15 * i)).strftime("%H:%M")
                if mid not in skip_two:
                    skip_two[mid] = {}
                skip_two[mid][t] = True

    # ───── финальная сборка строк ──────────────────────────────────────────────
    calendar_table = []

    for time_str in slot_times:
        row = {"time": time_str, "cells": []}

        for master in masters:
            mid = master.id

            # 1) старт двухколоночной?
            if time_str in two_col_map[mid]:
                cell = two_col_map[mid][time_str]

                # Проверяем конфликт с отменённой слева
                try:
                    start_idx = slot_times.index(time_str)
                except ValueError:
                    start_idx = 0
                span_times = slot_times[start_idx:start_idx + cell["rowspan"]]
                overlaps_cancel_left = any(skip_lane[mid][0].get(t) for t in span_times)

                if not overlaps_cancel_left:
                    if cell["kind"] == "appt_active":
                        row["cells"].append(
                            _make_item_cell(
                                kind="appt_active",
                                item=cell["item"],
                                rowspan=cell["rowspan"],
                                colspan=2,
                                master_obj=master,
                                bg=MASTER_COLORS.get(mid),
                                show_cancelled=False,
                            )
                        )
                    else:
                        row["cells"].append(
                            _make_unavail_cell(
                                kind="unavailable",
                                rowsp=cell["rowspan"],
                                colspan=2,
                                avail_id=cell["availability_id"],
                                reason=cell["reason"],
                                from_s=cell["from"],
                                to_s=cell["to"],
                                until_s=cell["until"],
                            )
                        )
                    continue

                # Перенос вправо (lane-right)
                for t in span_times:
                    skip_lane[mid][1][t] = True

                c0 = cancel_lanes[mid][0].get(time_str)
                if c0:
                    row["cells"].append(
                        _make_item_cell(
                            kind="appt_cancelled",
                            item=c0["item"],
                            rowspan=c0["rowspan"],
                            colspan=1,
                            master_obj=master,
                            bg=MASTER_COLORS.get(mid),
                            show_cancelled=True,
                        )
                    )
                elif not skip_lane[mid][0].get(time_str):
                    row["cells"].append({
                        "rowspan": 1, "colspan": 1, "kind": "free_half",
                        "master_id": mid, "html": "", "lane": "left"
                    })

                # правая половина переносимого блока
                row["cells"].append(
                    _make_item_cell(
                        kind="appt_active_right" if cell["kind"] == "appt_active" else "unavailable_right",
                        item=cell.get("item"),
                        rowspan=cell["rowspan"],
                        colspan=1,
                        master_obj=master,
                        bg=MASTER_COLORS.get(mid),
                        show_cancelled=False,
                    ) if cell["kind"] == "appt_active" else
                    _make_unavail_cell(
                        kind="unavailable_right",
                        rowsp=cell["rowspan"],
                        colspan=1,
                        avail_id=cell["availability_id"],
                        reason=cell["reason"],
                        from_s=cell["from"],
                        to_s=cell["to"],
                        until_s=cell["until"],
                    )
                )
                continue

            # 2) lane-режим (отменённые/перенесённые)
            lane0_start = time_str in cancel_lanes[mid][0]
            lane0_skip = bool(skip_lane[mid][0].get(time_str))
            lane1_skip = bool(skip_lane[mid][1].get(time_str))
            lane_mode = lane0_start or lane0_skip or lane1_skip

            if lane_mode:
                # левая половинка
                c0 = cancel_lanes[mid][0].get(time_str)
                if c0:
                    row["cells"].append(
                        _make_item_cell(
                            kind="appt_cancelled",
                            item=c0["item"],
                            rowspan=c0["rowspan"],
                            colspan=1,
                            master_obj=master,
                            bg=MASTER_COLORS.get(mid),
                            show_cancelled=True,
                        )
                    )
                elif not lane0_skip:
                    row["cells"].append({
                        "rowspan": 1, "colspan": 1, "kind": "free_half",
                        "master_id": mid, "html": "", "lane": "left"
                    })

                # правая половинка
                if not lane1_skip:
                    row["cells"].append({
                        "rowspan": 1, "colspan": 1, "kind": "free_half",
                        "master_id": mid, "html": "", "lane": "right"
                    })
                continue

            # 3) продолжающиеся двухколоночные
            if skip_two[mid].get(time_str):
                continue

            # 4) дефолтная свободная двухколоночная
            row["cells"].append({
                "rowspan": 1,
                "colspan": 2,
                "kind": "free",
                "master_id": mid,
                "html": "",
            })

        calendar_table.append(row)

    return calendar_table


def _health_flag_info(appt):
    """
    True/False + (url|""), title — нужно ли показывать флаг здоровья.
    Поддерживает и profile.health, и profile.health_conditions.
    """
    prof = getattr(appt, "client", None)
    if not prof:
        return False, "", ""

    hc = getattr(prof, "health", None) or getattr(prof, "health_conditions", None) or {}

    def _to_str(v):
        if isinstance(v, (list, tuple)):
            return ", ".join(map(str, v))
        return (v or "").strip()

    has_all = bool(hc.get("has_allergies")) or bool(_to_str(hc.get("allergies")))
    has_med = bool(_to_str(hc.get("medications")))
    has_ctr = bool(_to_str(hc.get("contraindications")))

    if not (has_all or has_med or has_ctr):
        return False, "", ""

    try:
        url = reverse("health-view-master", args=[prof.id])
    except NoReverseMatch:
        url = ""

    return True, url, "Есть важные данные в анкете здоровья — нажмите, чтобы посмотреть"

def _corner_badges_html(appt, appt_promocode):
    """
    Возвращает HTML «уголка» со значками (скидка + здоровье).
    Вызываем внутри карточки визита, у родителя должен быть position:relative.
    """
    # скидка
    promo_html = ""
    if appt_promocode or appt.final_price != appt.service.base_price:
        promo_html = "<span class='badge badge--promo' title='Applied discount'>%</span>"

    paid_total = getattr(appt, "_cached_paid_total", None)
    if paid_total is None:
        paid_total = _compute_paid_total(appt)
        setattr(appt, "_cached_paid_total", paid_total)

    final_price = getattr(appt, "final_price", None)
    grand_total = Decimal(final_price or Decimal("0.00"))
    if final_price is None:
        service_discounted_subtotal = Decimal("0.00")
        service_tax_total = Decimal("0.00")
        items_rel = getattr(appt, "items", None)
        if items_rel is not None:
            try:
                items_iterable = items_rel.all()
            except AttributeError:
                items_iterable = items_rel or []
            for appt_item in items_iterable or []:
                final_val = getattr(appt_item, "final_price", None)
                if final_val is None:
                    if appt_item.unit_price is not None:
                        final_val = appt_item.unit_price
                    else:
                        final_val = getattr(appt_item.service, "base_price", Decimal("0.00"))
                service_discounted_subtotal += Decimal(final_val or Decimal("0.00"))
                service_tax_total += Decimal(getattr(appt_item, "tax_amount", Decimal("0.00")) or Decimal("0.00"))
        service_discounted_subtotal = service_discounted_subtotal.quantize(TWOPLACES)
        service_tax_total = service_tax_total.quantize(TWOPLACES)

        products_total_with_tax = Decimal("0.00")
        product_sales_rel = getattr(appt, "product_sales", None)
        if product_sales_rel is not None:
            try:
                product_sales_iterable = product_sales_rel.all()
            except AttributeError:
                product_sales_iterable = product_sales_rel or []
            for sale in product_sales_iterable or []:
                subtotal = Decimal(getattr(sale, "total_amount", Decimal("0.00")) or Decimal("0.00"))
                tax_amount = Decimal(getattr(sale, "tax_amount", Decimal("0.00")) or Decimal("0.00"))
                products_total_with_tax += subtotal + tax_amount
        products_total_with_tax = products_total_with_tax.quantize(TWOPLACES)
        grand_total = (service_discounted_subtotal + service_tax_total + products_total_with_tax).quantize(TWOPLACES)
    else:
        grand_total = grand_total.quantize(TWOPLACES)

    payment_html = ""
    eps = Decimal("0.01")
    if grand_total is not None and (grand_total >= eps or paid_total >= eps):
            if paid_total >= grand_total - eps:
                payment_html = (
                    f"<span class='badge badge--paid'>"
                    f"<img src=\"{PAID_BADGE_ICON_URL}\" alt=\"Paid\" class=\"badge-icon\" />"
                    f"</span>"
                )
            elif paid_total >= eps and paid_total < grand_total - eps:
                payment_html = (
                    f"<span class='badge badge--partial'>"
                    f"<img src=\"{PARTIAL_BADGE_ICON_URL}\" alt=\"Partially paid\" class=\"badge-icon\" />"
                    f"</span>"
                )

    # здоровье
    show_flag, flag_url, flag_title = _health_flag_info(appt)
    health_html = ""
    if show_flag:
        ico = "⚕️"  # можно заменить на 🩺
        if flag_url:
            health_html = f'<a class="badge badge--health" href="{flag_url}" title="{flag_title}">{ico}</a>'
        else:
            health_html = f'<span class="badge badge--health" title="{flag_title}">{ico}</span>'

    note_html = ""
    note_value = getattr(appt, "notes", "") or ""
    if str(note_value).strip():
        note_html = "<span class='badge badge--note' title='Internal note'>??</span>"
    badges_html = "".join(filter(None, [promo_html, payment_html, health_html, note_html]))
    if not badges_html:
        return ""

    return f"<div class='corner-badges'>{badges_html}</div>"

# core/admin.py
from django.db import connection

def custom_index(request):
    from django.template.response import TemplateResponse

    tables = set(connection.introspection.table_names())
    userprof = None
    if "core_userprofile" in tables:
        try:
            userprof = getattr(request.user, "userprofile", None)
        except Exception:
            userprof = None

    ctx = {"userprof": userprof}
    return TemplateResponse(request, "admin/index.html", ctx)
