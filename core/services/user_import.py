"""Utilities for importing users from structured spreadsheets."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import StringIO
import csv
import re
from typing import Dict, Iterable, Iterator, List, Tuple

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction

from core.forms import UserImportRowForm
from core.models import Role, UserProfile, UserRole


class UserImportError(Exception):
    """Base error for user import issues."""


class UserImportSchemaError(UserImportError):
    """Raised when the uploaded file has unexpected structure."""


@dataclass
class ImportMessage:
    row_number: int
    username: str | None
    message: str


@dataclass
class UserImportResult:
    created: int = 0
    errors: List[ImportMessage] = field(default_factory=list)

    def add_error(self, row_number: int, username: str | None, message: str) -> None:
        self.errors.append(ImportMessage(row_number=row_number, username=username, message=message))


NORMALIZED_REQUIRED_HEADERS: Dict[str, str] = {
    "username": "username",
    "user": "username",
    "login": "username",
    "email": "email",
    "mail": "email",
    "password": "password",
    "pass": "password",
    "firstname": "first_name",
    "first_name": "first_name",
    "first name": "first_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "last name": "last_name",
    "phone": "phone",
    "phonenumber": "phone",
    "mobile": "phone",
    "telephone": "phone",
}

REQUIRED_FIELDS = {"username", "email", "password", "first_name", "last_name", "phone"}
SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xlsm")


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.strip().lower())


def _build_header_mapping(raw_headers: Iterable[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        if not header:
            continue
        normalized = normalize_header(str(header))
        target = NORMALIZED_REQUIRED_HEADERS.get(normalized)
        if target:
            mapping[idx] = target

    missing = REQUIRED_FIELDS.difference(mapping.values())
    if missing:
        readable_missing = ", ".join(sorted(missing))
        raise UserImportSchemaError(f"The file is missing required columns: {readable_missing}.")
    return mapping


def _iter_rows_from_csv(uploaded_file) -> Iterator[Tuple[int, Dict[str, str]]]:
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("utf-8")
    stream = StringIO(text)
    reader = csv.reader(stream)
    header_row = next(reader, None)
    if not header_row:
        raise UserImportSchemaError("The CSV file is empty.")

    header_mapping = _build_header_mapping(header_row)

    for row_index, row_values in enumerate(reader, start=2):
        row_data: Dict[str, str] = {}
        for idx, value in enumerate(row_values):
            if idx not in header_mapping:
                continue
            normalized_value = "" if value is None else str(value).strip()
            row_data[header_mapping[idx]] = normalized_value

        if all(not value for value in row_data.values()):
            continue
        yield row_index, row_data


def _iter_rows_from_xlsx(uploaded_file) -> Iterator[Tuple[int, Dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - guarded by requirements
        raise UserImportError("XLSX import requires the 'openpyxl' package.") from exc

    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_mapping: Dict[int, str] | None = None
        for excel_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_mapping is None:
                header_mapping = _build_header_mapping(row)
                continue

            row_data: Dict[str, str] = {}
            for idx, value in enumerate(row):
                if idx not in header_mapping:
                    continue
                normalized_value = "" if value is None else str(value).strip()
                row_data[header_mapping[idx]] = normalized_value

            if all(not value for value in row_data.values()):
                continue
            yield excel_index, row_data
    finally:
        workbook.close()


def iter_user_rows(uploaded_file) -> Iterator[Tuple[int, Dict[str, str]]]:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return _iter_rows_from_csv(uploaded_file)
    if name.endswith((".xlsx", ".xlsm")):
        return _iter_rows_from_xlsx(uploaded_file)
    raise UserImportError("Unsupported file type. Use CSV or XLSX.")


def _ensure_profile(user) -> UserProfile:
    profile, created = UserProfile.objects.get_or_create(user=user)
    if created:
        profile.source = "offline"
        profile.save(update_fields=["source"])
    else:
        if profile.source != "offline":
            profile.source = "offline"
            profile.save(update_fields=["source"])
    return profile


def _ensure_default_role(profile: UserProfile) -> None:
    client_role, _ = Role.objects.get_or_create(name="Client")
    UserRole.objects.get_or_create(user=profile, role=client_role)


def import_users_from_file(uploaded_file) -> UserImportResult:
    """Parse the uploaded file and create users.

    Rows that fail validation are collected in the result; successful rows are
    persisted in a single transaction per row.
    """
    result = UserImportResult()
    User = get_user_model()

    try:
        row_iterator = iter_user_rows(uploaded_file)
    except UserImportError:
        raise
    except Exception as exc:
        raise UserImportError(f"Unable to read the file: {exc}") from exc

    for row_number, payload in row_iterator:
        form = UserImportRowForm(data=payload)
        if not form.is_valid():
            error_data = form.errors.get_json_data()
            messages = [
                entry["message"]
                for field_errors in error_data.values()
                for entry in field_errors
            ]
            message = "; ".join(messages) or "Row validation failed"
            result.add_error(row_number, payload.get("username"), message)
            continue

        cleaned = form.cleaned_data
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=cleaned["username"],
                    email=cleaned["email"],
                    password=cleaned["password"],
                    first_name=cleaned["first_name"],
                    last_name=cleaned["last_name"],
                    is_active=True,
                )
                profile = _ensure_profile(user)
                profile.phone = cleaned["phone"]
                profile.save(update_fields=["phone"])
                _ensure_default_role(profile)
        except (ValidationError, IntegrityError) as exc:
            result.add_error(row_number, cleaned.get("username"), str(exc))
            continue
        result.created += 1

    return result

__all__ = [
    "import_users_from_file",
    "UserImportError",
    "UserImportSchemaError",
    "UserImportResult",
    "ImportMessage",
]
