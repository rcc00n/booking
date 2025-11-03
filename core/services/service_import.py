"""Utilities for importing services from structured spreadsheets."""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import StringIO
import re
from typing import Dict, Iterable, Iterator, List, Set, Tuple

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction

from core.forms import ServiceImportRowForm
from core.models import Service, ServiceCategory


class ServiceImportError(Exception):
    """Base error for service import failures."""


class ServiceImportSchemaError(ServiceImportError):
    """Raised when the uploaded file does not include required columns."""


@dataclass
class ImportMessage:
    row_number: int
    reference: str | None
    message: str


@dataclass
class ServiceImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[ImportMessage] = field(default_factory=list)

    def add_error(self, *, row_number: int, reference: str | None, message: str) -> None:
        self.errors.append(ImportMessage(row_number=row_number, reference=reference, message=message))
        self.skipped += 1


NORMALIZED_HEADERS: Dict[str, str] = {
    "servicename": "name",
    "service": "name",
    "name": "name",
    "retailprice": "base_price",
    "price": "base_price",
    "baseprice": "base_price",
    "duration": "duration",
    "durationmin": "duration",
    "durationminutes": "duration",
    "sessionduration": "duration",
    "extratime": "extra_time",
    "extratimeafter": "extra_time",
    "blockedtimeafter": "extra_time",
    "postbuffer": "extra_time",
    "buffer": "extra_time",
    "tax": "tax",
    "taxrate": "tax",
    "description": "description",
    "details": "description",
    "category": "category",
    "servicecategory": "category",
}

REQUIRED_FIELDS = {"name", "base_price", "duration"}


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _build_header_mapping(raw_headers: Iterable[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        if header is None:
            continue
        normalized = _normalize_header(header)
        target = NORMALIZED_HEADERS.get(normalized)
        if target:
            mapping[idx] = target
    missing = REQUIRED_FIELDS.difference(mapping.values())
    if missing:
        readable = ", ".join(sorted(missing))
        raise ServiceImportSchemaError(f"The file is missing required columns: {readable}.")
    return mapping


def _iter_rows_from_csv(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("utf-8")
    reader = csv.reader(StringIO(text))
    header_row = next(reader, None)
    if not header_row:
        raise ServiceImportSchemaError("The CSV file is empty.")

    header_mapping = _build_header_mapping(header_row)
    for row_index, row_values in enumerate(reader, start=2):
        row_data: Dict[str, str] = {}
        provided_fields: Set[str] = set()
        for idx, value in enumerate(row_values):
            if idx not in header_mapping:
                continue
            field_name = header_mapping[idx]
            provided_fields.add(field_name)
            row_data[field_name] = "" if value is None else str(value).strip()
        if not row_data:
            continue
        if all(not value for value in row_data.values()):
            continue
        yield row_index, row_data, provided_fields


def _iter_rows_from_xlsx(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - validated by requirements
        raise ServiceImportError("XLSX import requires the 'openpyxl' package.") from exc

    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_mapping: Dict[int, str] | None = None
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_mapping is None:
                header_mapping = _build_header_mapping(row)
                continue
            row_data: Dict[str, str] = {}
            provided_fields: Set[str] = set()
            for idx, value in enumerate(row):
                if idx not in header_mapping:
                    continue
                field_name = header_mapping[idx]
                provided_fields.add(field_name)
                if value is None:
                    row_data[field_name] = ""
                elif isinstance(value, str):
                    row_data[field_name] = value.strip()
                else:
                    row_data[field_name] = str(value).strip()
            if not row_data:
                continue
            if all(not value for value in row_data.values()):
                continue
            yield row_index, row_data, provided_fields
    finally:
        workbook.close()


def iter_service_rows(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        return _iter_rows_from_csv(uploaded_file)
    if name.endswith((".xlsx", ".xlsm")):
        return _iter_rows_from_xlsx(uploaded_file)
    raise ServiceImportError("Unsupported file type. Upload CSV or XLSX.")


def _resolve_category(name: str | None) -> ServiceCategory | None:
    if not name:
        return None
    normalized = name.strip()
    if not normalized:
        return None
    existing = ServiceCategory.objects.filter(name__iexact=normalized).first()
    if existing:
        return existing
    category = ServiceCategory(name=normalized)
    category.save()
    return category


def _lookup_service(*, name: str) -> Service | None:
    return Service.objects.filter(name__iexact=name).first()


def _apply_payload_to_service(
    *,
    service: Service,
    cleaned_data: Dict[str, object],
    provided_fields: Set[str],
) -> None:
    service.name = cleaned_data["name"]

    if "description" in provided_fields or not service.pk:
        service.description = cleaned_data.get("description") or ""

    if "base_price" in provided_fields or not service.pk:
        service.base_price = cleaned_data["base_price"]

    if "duration" in provided_fields or not service.pk:
        service.duration_min = int(cleaned_data["duration"])

    if "extra_time" in provided_fields or not service.pk:
        extra_minutes = cleaned_data.get("extra_time")
        service.extra_time_min = int(extra_minutes) if extra_minutes is not None else None

    if "tax" in provided_fields or not service.pk:
        service.is_taxable = bool(cleaned_data.get("tax"))

    if "category" in provided_fields or not service.pk:
        category_name = cleaned_data.get("category") or None
        service.category = _resolve_category(category_name)

    if not service.pk:
        service.is_active = True

    if not getattr(service, "image_alt_text", ""):
        service.image_alt_text = service.name


def _upsert_service(*, cleaned_data: Dict[str, object], provided_fields: Set[str]) -> Tuple[Service, bool]:
    name_value = cleaned_data["name"]
    service = _lookup_service(name=name_value)
    created = False
    if service is None:
        service = Service(name=name_value)
        created = True

    _apply_payload_to_service(service=service, cleaned_data=cleaned_data, provided_fields=provided_fields)
    service.full_clean()
    service.save()
    return service, created


def import_services_from_file(uploaded_file) -> ServiceImportResult:
    result = ServiceImportResult()

    try:
        row_iterator = iter_service_rows(uploaded_file)
    except ServiceImportError:
        raise
    except Exception as exc:
        raise ServiceImportError(f"Unable to read the file: {exc}") from exc

    for row_number, raw_payload, provided_fields in row_iterator:
        form = ServiceImportRowForm(data=raw_payload)
        if not form.is_valid():
            error_messages = [
                entry["message"]
                for field_errors in form.errors.get_json_data().values()
                for entry in field_errors
            ]
            reference = raw_payload.get("name")
            result.add_error(
                row_number=row_number,
                reference=reference,
                message="; ".join(error_messages) or "Row validation failed",
            )
            continue

        cleaned = form.cleaned_data
        reference = cleaned.get("name")

        try:
            with transaction.atomic():
                _, created = _upsert_service(cleaned_data=cleaned, provided_fields=provided_fields)
        except (ValidationError, IntegrityError) as exc:
            result.add_error(row_number=row_number, reference=reference, message=str(exc))
            continue

        if created:
            result.created += 1
        else:
            result.updated += 1

    return result
