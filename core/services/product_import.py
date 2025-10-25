"""Utilities for importing retail products from structured spreadsheets."""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from io import StringIO
import csv
import re
from typing import Dict, Iterable, Iterator, List, Set, Tuple

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction

from core.forms import ProductImportRowForm
from core.models import Product, ProductCategory


class ProductImportError(Exception):
    """Base error for product import failures."""


class ProductImportSchemaError(ProductImportError):
    """Raised when the uploaded file does not include required columns."""


@dataclass
class ImportMessage:
    row_number: int
    reference: str | None
    message: str


@dataclass
class ProductImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[ImportMessage] = field(default_factory=list)

    def add_error(self, *, row_number: int, reference: str | None, message: str) -> None:
        self.errors.append(ImportMessage(row_number=row_number, reference=reference, message=message))
        self.skipped += 1


SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xlsm")

NORMALIZED_HEADERS = {
    "productname": "name",
    "name": "name",
    "product": "name",
    "sku": "sku",
    "barcode": "sku",
    "description": "description",
    "measuretype": "measure_type",
    "measure_type": "measure_type",
    "measurevalue": "measure_value",
    "measure_value": "measure_value",
    "measurementvalue": "measure_value",
    "costprice": "cost_price",
    "cost_price": "cost_price",
    "wholesaleprice": "cost_price",
    "fullprice": "full_price",
    "full_price": "full_price",
    "price": "full_price",
    "retailprice": "full_price",
    "category": "category",
    "brand": "brand",
    "supplier": "supplier",
    "totalstock": "total_stock",
    "stock": "total_stock",
}

REQUIRED_FIELDS = {"name"}


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _build_header_mapping(raw_headers: Iterable[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        if header is None:
            continue
        normalized = _normalize_header(header)
        target = NORMALIZED_HEADERS.get(normalized)
        if target:
            mapping[idx] = target
    missing = REQUIRED_FIELDS.difference(mapping.values())
    if missing:
        readable = ", ".join(sorted(missing))
        raise ProductImportSchemaError(f"The file is missing required columns: {readable}.")
    return mapping


def _iter_rows_from_csv(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw_bytes.decode("utf-8")
    reader = csv.reader(StringIO(text))
    header_row = next(reader, None)
    if not header_row:
        raise ProductImportSchemaError("The CSV file is empty.")

    header_mapping = _build_header_mapping(header_row)
    for row_index, row_values in enumerate(reader, start=2):
        row_data: Dict[str, str] = {}
        provided_fields: Set[str] = set()
        for idx, value in enumerate(row_values):
            if idx not in header_mapping:
                continue
            field_name = header_mapping[idx]
            provided_fields.add(field_name)
            row_data[field_name] = "" if value is None else str(value).strip()
        if not row_data:
            continue
        if all(not value for value in row_data.values()):
            continue
        yield row_index, row_data, provided_fields


def _iter_rows_from_xlsx(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - validated by requirements
        raise ProductImportError("XLSX import requires the 'openpyxl' package.") from exc

    uploaded_file.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_mapping: Dict[int, str] | None = None
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_mapping is None:
                header_mapping = _build_header_mapping(row)
                continue
            row_data: Dict[str, str] = {}
            provided_fields: Set[str] = set()
            for idx, value in enumerate(row):
                if idx not in header_mapping:
                    continue
                field_name = header_mapping[idx]
                provided_fields.add(field_name)
                if value is None:
                    row_data[field_name] = ""
                elif isinstance(value, str):
                    row_data[field_name] = value.strip()
                else:
                    row_data[field_name] = str(value).strip()
            if not row_data:
                continue
            if all(not value for value in row_data.values()):
                continue
            yield row_index, row_data, provided_fields
    finally:
        workbook.close()


def iter_product_rows(uploaded_file) -> Iterator[Tuple[int, Dict[str, str], Set[str]]]:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return _iter_rows_from_csv(uploaded_file)
    if name.endswith((".xlsx", ".xlsm")):
        return _iter_rows_from_xlsx(uploaded_file)
    raise ProductImportError("Unsupported file type. Upload CSV or XLSX.")


def _resolve_category(name: str | None) -> ProductCategory | None:
    if not name:
        return None
    normalized = name.strip()
    if not normalized:
        return None
    existing = ProductCategory.objects.filter(name__iexact=normalized).first()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            existing.save(update_fields=["is_active"])
        return existing
    return ProductCategory.objects.create(name=normalized, description="")


def _lookup_product(*, sku: str | None, name: str) -> Product | None:
    if sku:
        product = Product.objects.filter(sku__iexact=sku).first()
        if product:
            return product
    return Product.objects.filter(name__iexact=name).first()


def _apply_payload_to_product(
    *,
    product: Product,
    cleaned_data: Dict[str, object],
    provided_fields: Set[str],
) -> None:
    sku_value = cleaned_data.get("sku") or None
    if "sku" in provided_fields or not product.pk:
        product.sku = sku_value

    product.name = cleaned_data["name"]

    if "category" in provided_fields or not product.pk:
        category_name = cleaned_data.get("category") or None
        product.category = _resolve_category(category_name)

    text_fields = ("description", "measure_type", "measure_value", "brand", "supplier")
    for field in text_fields:
        if field in provided_fields or not product.pk:
            value = cleaned_data.get(field) or ""
            setattr(product, field, value)

    if "cost_price" in provided_fields or not product.pk:
        product.cost_price = cleaned_data.get("cost_price")

    if "full_price" in provided_fields or not product.pk:
        price_value = cleaned_data.get("full_price")
        if price_value is None:
            price_value = Decimal("0.00") if not product.pk else product.price
        product.price = price_value

    if "total_stock" in provided_fields or not product.pk:
        stock_value = cleaned_data.get("total_stock")
        if stock_value is None:
            stock_value = 0 if not product.pk else product.quantity_in_stock
        product.quantity_in_stock = stock_value


def _upsert_product(*, cleaned_data: Dict[str, object], provided_fields: Set[str]) -> Tuple[Product, bool]:
    sku_value = cleaned_data.get("sku") or None
    name_value = cleaned_data["name"]
    product = _lookup_product(sku=sku_value, name=name_value)
    created = False
    if product is None:
        product = Product(name=name_value)
        created = True

    _apply_payload_to_product(product=product, cleaned_data=cleaned_data, provided_fields=provided_fields)
    product.full_clean()
    product.save()
    return product, created


def import_products_from_file(uploaded_file) -> ProductImportResult:
    result = ProductImportResult()

    try:
        row_iterator = iter_product_rows(uploaded_file)
    except ProductImportError:
        raise
    except Exception as exc:
        raise ProductImportError(f"Unable to read the file: {exc}") from exc

    for row_number, raw_payload, provided_fields in row_iterator:
        form = ProductImportRowForm(data=raw_payload)
        if not form.is_valid():
            error_messages = [
                entry["message"]
                for field_errors in form.errors.get_json_data().values()
                for entry in field_errors
            ]
            reference = raw_payload.get("sku") or raw_payload.get("name")
            result.add_error(
                row_number=row_number,
                reference=reference,
                message="; ".join(error_messages) or "Row validation failed",
            )
            continue

        cleaned = form.cleaned_data
        reference = cleaned.get("sku") or cleaned.get("name")

        try:
            with transaction.atomic():
                _product, created = _upsert_product(cleaned_data=cleaned, provided_fields=provided_fields)
        except (ValidationError, IntegrityError) as exc:
            result.add_error(row_number=row_number, reference=reference, message=str(exc))
            continue

        if created:
            result.created += 1
        else:
            result.updated += 1

    return result
